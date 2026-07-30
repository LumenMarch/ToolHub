import io
import os
import shutil
import tempfile
import threading
import traceback
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from time import perf_counter
from urllib.parse import quote

import polars as pl
import xlsxwriter
from dateutil.relativedelta import relativedelta
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from loguru import logger
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field
from PyPDF2 import PdfMerger

from app.core.auth import require_permission, require_tool_enabled
from app.models.user import User
from app.schemas.asset_comparison import (
    AssetComparisonAnnotationsUpdate,
    AssetComparisonJobCreate,
)
from app.services.asset_comparison.Customer_Customer import Customer_Customer
from app.services.asset_comparison.Customer_Notes import Customer_Notes
from app.services.asset_comparison.Finance_Finance import Finance_Finance
from app.services.asset_comparison.Finance_Notes import Finance_Notes
from app.services.asset_comparison.job_manager import (
    AssetComparisonJobConflictError,
    AssetComparisonJobExpiredError,
    AssetComparisonJobManager,
    AssetComparisonJobNotFoundError,
    AssetComparisonJobValidationError,
)
from app.services.asset_comparison.Notes_Notes import Notes_Notes
from app.services.asset_comparison.Notes_SFC import Notes_SFC
from app.services.asset_comparison.SFC_SFC import SFC_SFC
from app.services.upload.store import UploadStore

try:
    from app.services.asset_comparison.mod import create_excel_template
except ImportError:
    pass
try:
    from app.services.asset_comparison.pdf_generator import (
        create_pdf_from_sheets,
        excel_sheet_to_pdf,
    )
except ImportError:
    pass
try:
    from PyPDF2 import PdfMerger
except ImportError:
    pass

router = APIRouter()


class ComparisonRequest(BaseModel):
    thisFinance: str
    lastFinance: str
    thisSFC: str
    lastSFC: str
    thisNotes: str
    lastNotes: str
    thisCustomer: str
    lastCustomer: str
    departmentData: str
    custodianData: str
    driData: str
    remarks: dict = {}
    reviews: dict = {}


@router.get("/resolve-folder")
async def resolve_folder(
    name: str = "",
    current_user: User = Depends(require_permission("tool:use")),
    _: None = Depends(require_tool_enabled("asset-comparison")),
):
    """前端选文件夹后只能拿到文件夹名，用 find 搜索定位绝对路径"""
    import subprocess

    if not name or not name.strip():
        return {"status": "error", "message": "请提供文件夹名"}

    search_dirs = [
        str(Path.home() / "Desktop"),
        str(Path.home() / "Downloads"),
        str(Path.home() / "Documents"),
    ]
    found_paths = []

    for search_dir in search_dirs:
        if not os.path.isdir(search_dir):
            continue
        try:
            result = subprocess.run(
                [
                    "find",
                    search_dir,
                    "-maxdepth",
                    "3",
                    "-type",
                    "d",
                    "-name",
                    name.strip(),
                ],
                capture_output=True,
                text=True,
                timeout=5,
            )
            for line in result.stdout.strip().split("\n"):
                line = line.strip()
                if line:
                    found_paths.append(line)
        except Exception:
            continue

    if found_paths:
        # 如果有多个匹配，优先选文件数量最多的那个
        best_path = ""
        best_count = 0
        for p in found_paths:
            try:
                cnt = len(
                    [
                        f
                        for f in os.listdir(p)
                        if os.path.isfile(os.path.join(p, f)) and not f.startswith("~")
                    ]
                )
            except Exception:
                cnt = 0
            if cnt > best_count:
                best_count = cnt
                best_path = p
        return {
            "status": "success",
            "path": best_path,
            "file_count": best_count,
            "candidates": found_paths,
        }
    else:
        return {
            "status": "not_found",
            "message": f"未在 Desktop/Downloads/Documents 下找到名为 '{name}' 的文件夹",
        }


def _build_match_rules(this_month_str: str, last_month_str: str) -> dict:
    """构建文件名匹配规则"""
    return {
        "thisFinance": ("财务资产", "monthly", this_month_str),
        "lastFinance": ("财务资产", "monthly", last_month_str),
        "thisSFC": ("SFC资产", "monthly", this_month_str),
        "lastSFC": ("SFC资产", "monthly", last_month_str),
        "thisNotes": ("Notes资产", "monthly", this_month_str),
        "lastNotes": ("Notes资产", "monthly", last_month_str),
        "thisCustomer": ("客户资产", "monthly", this_month_str),
        "lastCustomer": ("客户资产", "monthly", last_month_str),
        "custodianData": ("财务保管人", "fixed", ""),
        "departmentData": ("财务保管部门", "fixed", ""),
        "driData": ("客户系统DRI", "fixed", ""),
    }


def _scan_and_match(folder: Path, match_rules: dict) -> tuple:
    """扫描文件夹中的文件并匹配，返回 (result_dict, found_files, matched_log)"""
    result = {k: "" for k in match_rules}
    found_files = []
    matched_log = []

    if folder.exists() and folder.is_dir():
        for f in folder.iterdir():
            if f.is_file() and not f.name.startswith("~"):
                found_files.append(f.name)
                stem = f.stem
                stem_clean = stem.replace(" ", "").replace("_", "").replace("-", "")
                for k, (keyword, rule_type, expected_date) in match_rules.items():
                    kw_clean = keyword.replace(" ", "")
                    if rule_type == "fixed":
                        if stem_clean == kw_clean:
                            result[k] = str(f)
                            matched_log.append(f"✓ {f.name} → {k}")
                    elif rule_type == "monthly":
                        if kw_clean in stem_clean and expected_date in stem_clean:
                            result[k] = str(f)
                            matched_log.append(f"✓ {f.name} → {k}")
    else:
        found_files.append("⚠️ 文件夹不存在或不是目录")

    return result, found_files, matched_log


ASSET_COMPARE_ROOT = Path(tempfile.gettempdir()) / "asset-compare"


class ScanByIdsRequest(BaseModel):
    """基于已上传文件 ID 的扫描请求。"""

    upload_ids: list[str] = Field(..., min_length=1, description="已上传文件 ID 列表")


@router.post("/scan")
async def scan_uploaded_files_by_ids(
    req: ScanByIdsRequest,
    current_user: User = Depends(require_permission("tool:use")),
    _: None = Depends(require_tool_enabled("asset-comparison")),
):
    """使用 upload_id 获取已上传文件，扫描匹配后返回路径。"""
    store = UploadStore()

    # 创建临时工作目录
    ASSET_COMPARE_ROOT.mkdir(parents=True, exist_ok=True)
    work_dir = Path(tempfile.mkdtemp(prefix="scan-", dir=ASSET_COMPARE_ROOT))

    logger.info(
        f"[scan] upload_ids={req.upload_ids} | work_dir={work_dir} | "
        f"user={current_user.username}"
    )

    # 将上传文件复制到工作目录
    for upload_id in req.upload_ids:
        info = store.get_info(upload_id)
        src = store.get_file_path(upload_id)
        safe_name = Path(info["filename"]).name
        if not safe_name:
            logger.warning(f"[scan] 跳过空文件名: upload_id={upload_id}")
            continue
        dst = work_dir / safe_name
        shutil.copy2(src, dst)
        logger.info(f"[scan] 复制 {info['filename']} -> {dst}")
    # 文件已复制到工作目录，删除源上传避免磁盘泄漏
    for upload_id in req.upload_ids:
        store.delete(upload_id)

    # 扫描匹配
    current_date = datetime.now()
    this_month_str = current_date.strftime("%Y%m")
    last_month_date = current_date - relativedelta(months=1)
    last_month_str = last_month_date.strftime("%Y%m")

    match_rules = _build_match_rules(this_month_str, last_month_str)
    result, found_files, matched_log = _scan_and_match(work_dir, match_rules)

    matched_count = sum(1 for v in result.values() if v)
    logger.info(f"[scan] 匹配 {matched_count}/{len(result)} 个数据表")
    for entry in matched_log:
        logger.info(f"[scan]   {entry}")

    return {
        "status": "success",
        "data": result,
        "scan_id": work_dir.name,
        "debug": {
            "folder": str(work_dir),
            "found_files": found_files,
            "matched": matched_log,
        },
    }


@router.get("/auto-paths")
async def get_auto_paths(
    folder: str = "",
    current_user: User = Depends(require_permission("tool:use")),
    _: None = Depends(require_tool_enabled("asset-comparison")),
):
    current_date = datetime.now()
    this_month_str = current_date.strftime("%Y%m")
    last_month_date = current_date - relativedelta(months=1)
    last_month_str = last_month_date.strftime("%Y%m")

    if folder:
        all_data_path = Path(folder)
    else:
        all_data_path = Path.home() / "Desktop" / "对比数据"

    match_rules = _build_match_rules(this_month_str, last_month_str)
    result, found_files, matched_log = _scan_and_match(all_data_path, match_rules)

    return {
        "status": "success",
        "data": result,
        "debug": {
            "folder": str(all_data_path),
            "found_files": found_files,
            "matched": matched_log,
        },
    }


def _safe_len(d):
    if d is None:
        return 0
    try:
        return len(d)
    except Exception:
        return 0


def task_ff(req: ComparisonRequest):
    ff = Finance_Finance()
    ff.this_Finance_path = req.thisFinance
    ff.last_Finance_path = req.lastFinance
    ff.Custodian_path = req.custodianData
    ff.Department_path = req.departmentData
    info = None
    if req.thisFinance and req.lastFinance:
        try:
            ff.read_Custodian_data()
            ff.read_Department_data()
            ff.read_this_Finance_data()
            ff.read_last_Finance_data()
            ff.Finance_check()
            diff_count = (
                _safe_len(ff.check_Custodian)
                + _safe_len(ff.check_Department)
                + _safe_len(ff.new_Custodian_assets)
                + _safe_len(ff.removed_Custodian_assets)
                + _safe_len(ff.new_Department_assets)
                + _safe_len(ff.removed_Department_assets)
            )
            # 依保管人 & 依部门 子维度明细
            cust_new = _safe_len(ff.new_Custodian_assets)
            cust_rm = _safe_len(ff.removed_Custodian_assets)
            cust_anomaly = _safe_len(ff.check_Custodian)
            dept_new = _safe_len(ff.new_Department_assets)
            dept_rm = _safe_len(ff.removed_Department_assets)
            dept_anomaly = _safe_len(ff.check_Department)
            info = {
                "key": "ff",
                "label": "【财务-财务】",
                "has_diff": diff_count > 0,
                "msg": f"保管人异常 {cust_anomaly} | 部门异常 {dept_anomaly}",
                "sub_groups": [
                    {
                        "label": "依保管人差异",
                        "new_count": cust_new,
                        "removed_count": cust_rm,
                        "anomaly_count": cust_anomaly,
                        "has_diff": (cust_new + cust_rm + cust_anomaly) > 0,
                    },
                    {
                        "label": "依部门差异",
                        "new_count": dept_new,
                        "removed_count": dept_rm,
                        "anomaly_count": dept_anomaly,
                        "has_diff": (dept_new + dept_rm + dept_anomaly) > 0,
                    },
                ],
            }
        except Exception as e:
            info = {
                "key": "ff",
                "label": "【财务-财务】",
                "has_diff": False,
                "msg": f"异常: {e}",
            }
    return "ff", ff, info


def task_sfc(req: ComparisonRequest):
    sfc = SFC_SFC()
    sfc.This_data_Path = req.thisSFC
    sfc.Last_data_path = req.lastSFC
    info = None
    if req.thisSFC and req.lastSFC:
        try:
            sfc.This_SFC_date()
            sfc.Last_SFC_date()
            sfc.SFC_SFC_Comparison()
            diff_count = _safe_len(sfc.new_assets) + _safe_len(sfc.removed_assets)
            info = {
                "key": "sfc",
                "label": "【SFC-SFC】",
                "has_diff": diff_count > 0,
                "msg": f"新增: {_safe_len(sfc.new_assets)}, 减少: {_safe_len(sfc.removed_assets)}",
            }
        except Exception as e:
            info = {
                "key": "sfc",
                "label": "【SFC-SFC】",
                "has_diff": False,
                "msg": f"异常: {e}",
            }
    return "sfc", sfc, info


def task_nn(req: ComparisonRequest):
    nn = Notes_Notes()
    nn.This_Notes_path = req.thisNotes
    nn.Last_Notes_path = req.lastNotes
    info = None
    if req.thisNotes and req.lastNotes:
        try:
            nn.This_Notes_date()
            nn.Last_Notes_date()
            nn.Notes_Notes_Comparison()
            diff_count = (
                _safe_len(nn.new_assets)
                + _safe_len(nn.removed_assets)
                + abs(_safe_len(nn.new_No_assets) - _safe_len(nn.removed_No_assets))
            )
            info = {
                "key": "nn",
                "label": "【Notes-Notes】",
                "has_diff": diff_count > 0,
                "msg": f"有资产新增: {_safe_len(nn.new_assets)}, 有资产减少: {_safe_len(nn.removed_assets)} | 无资产总差异: {abs(_safe_len(nn.new_No_assets) - _safe_len(nn.removed_No_assets))}",
            }
        except Exception as e:
            info = {
                "key": "nn",
                "label": "【Notes-Notes】",
                "has_diff": False,
                "msg": f"异常: {e}",
            }
    return "nn", nn, info


def task_cc(req: ComparisonRequest):
    cc = Customer_Customer()
    cc.this_Customer_path = req.thisCustomer
    cc.last_Customer_path = req.lastCustomer
    cc.Custodian_DRI_path = req.driData
    info = None
    if req.thisCustomer and req.lastCustomer and req.driData:
        try:
            cc.get_Customer_DRI()
            cc.read_this_Customer_data()
            cc.read_last_Customer_data()
            cc.Customer_Customer_Comparison()
            diff_count = _safe_len(cc.new_Customer_assets) + _safe_len(
                cc.removed_Customer_assets
            )
            info = {
                "key": "cc",
                "label": "【客户-客户】",
                "has_diff": diff_count > 0,
                "msg": f"新增: {_safe_len(cc.new_Customer_assets)}, 减少: {_safe_len(cc.removed_Customer_assets)}",
            }
        except Exception as e:
            info = {
                "key": "cc",
                "label": "【客户-客户】",
                "has_diff": False,
                "msg": f"异常: {e}",
            }
    return "cc", cc, info


def task_fn(req: ComparisonRequest):
    fn = Finance_Notes()
    fn.Finance_path = req.thisFinance
    fn.Notes_path = req.thisNotes
    fn.Custodian_path = req.custodianData
    info = None
    if req.thisFinance and req.thisNotes and req.custodianData:
        try:
            fn.read_Custodian_data()
            fn.read_Finance_data()
            fn.read_Notes_data()
            fn.Finance_Notes_Comparison()
            diff_count = _safe_len(fn.removed_assets) + _safe_len(fn.new_assets)
            info = {
                "key": "fn",
                "label": "【财务比Notes】",
                "has_diff": diff_count > 0,
                "msg": f"财务比Notes新增: {_safe_len(fn.removed_assets)}, 财务比Notes减少: {_safe_len(fn.new_assets)}",
            }
        except Exception as e:
            info = {
                "key": "fn",
                "label": "【财务比Notes】",
                "has_diff": False,
                "msg": f"异常: {e}",
            }
    return "fn", fn, info


def task_ns(req: ComparisonRequest):
    ns = Notes_SFC()
    ns.this_Notes_path = req.thisNotes
    ns.this_SFC_path = req.thisSFC
    info = None
    if req.thisNotes and req.thisSFC:
        try:
            ns.This_Notes_date()
            ns.This_SFC_date()
            ns.Notes_SFC_Comparison()
            diff_count = _safe_len(ns.Notes_new_assets) + _safe_len(
                ns.Notes_removed_assets
            )
            info = {
                "key": "ns",
                "label": "【Notes比SFC】",
                "has_diff": diff_count > 0,
                "msg": f"Notes比SFC新增: {_safe_len(ns.Notes_new_assets)}, Notes比SFC减少: {_safe_len(ns.Notes_removed_assets)}",
            }
        except Exception as e:
            info = {
                "key": "ns",
                "label": "【Notes比SFC】",
                "has_diff": False,
                "msg": f"异常: {e}",
            }
    return "ns", ns, info


def task_cn(req: ComparisonRequest):
    cn = Customer_Notes()
    cn.this_Customer_path = req.thisCustomer
    cn.this_Notes_path = req.thisNotes
    cn.this_Customer_DRI_path = req.driData
    info = None
    if req.thisCustomer and req.thisNotes and req.driData:
        try:
            cn.read_Customer_DRI()
            cn.read_this_Customer_data()
            cn.read_this_Notes_data()
            cn.Customer_Notes_Comparison()
            diff_count = _safe_len(cn.remove_assets) + _safe_len(cn.new_assets)
            info = {
                "key": "cn",
                "label": "【客户比Notes】",
                "has_diff": diff_count > 0,
                "msg": f"客户比Notes新增: {_safe_len(cn.remove_assets)}, 客户比Notes减少: {_safe_len(cn.new_assets)}",
            }
        except Exception as e:
            info = {
                "key": "cn",
                "label": "【客户比Notes】",
                "has_diff": False,
                "msg": f"异常: {e}",
            }
    return "cn", cn, info


def run_comparisons(req: ComparisonRequest, on_complete=None):
    summary = {}
    results_info = []

    # 使用普通线程池并发执行 7 个比对任务
    with ThreadPoolExecutor(max_workers=7) as executor:
        futures = [
            executor.submit(task_ff, req),
            executor.submit(task_sfc, req),
            executor.submit(task_nn, req),
            executor.submit(task_cc, req),
            executor.submit(task_fn, req),
            executor.submit(task_ns, req),
            executor.submit(task_cn, req),
        ]
        for future in as_completed(futures):
            key, instance, info = future.result()
            summary[key] = instance
            if info:
                results_info.append(info)
            if on_complete is not None:
                on_complete(key, instance, info)

    # 排序以保持输出顺序稳定
    order = ["ff", "sfc", "nn", "cc", "fn", "ns", "cn"]
    results_info.sort(key=lambda x: order.index(x["key"]) if x["key"] in order else 99)

    summary["results_info"] = results_info
    return summary


def apply_review_colors(ws, req_reviews):
    """根据前端的选择，给对应单元格涂色"""
    color_mapping = {
        "差異確認OK": "71D0F1",
        "待跟进": "FFEE00",
        "異常": "EC4337",
    }

    green_fill = PatternFill(
        start_color="D0F1AD", end_color="D0F1AD", fill_type="solid"
    )

    combo_to_cell = {
        "ff": ["D5"],
        "nn": ["D6"],
        "sfc": ["D7"],
        "cc": ["D8"],
        "fn": ["E5"],
        "ns": ["F6"],
        "cn": ["G8"],
    }

    for key, val in req_reviews.items():
        if key in combo_to_cell:
            coords = combo_to_cell[key]
            fill_color = color_mapping.get(val, "FFFFFF")
            fill_pattern = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type="solid"
            )
            for coord in coords:
                cell = ws[coord]
                cv = (
                    str(cell.value).strip().replace("/", "")
                    if cell.value is not None
                    else ""
                )
                numeric_val = float(cv) if cv and cv != "0" else 0
                if numeric_val == 0:
                    cell.fill = green_fill
                else:
                    cell.fill = fill_pattern


@router.post("/check")
def check_data(
    req: ComparisonRequest,
    current_user: User = Depends(require_permission("tool:use")),
    _: None = Depends(require_tool_enabled("asset-comparison")),
):
    try:
        summary = run_comparisons(req)
        return {
            "status": "success",
            "message": "交叉盘点全部完成。",
            "data": {"results": summary["results_info"]},
        }
    except Exception as e:
        return {
            "status": "error",
            "message": str(e),
            "errors": [traceback.format_exc()],
        }


_legacy_export_lock = threading.Lock()
_RAW_SOURCE_MODULES = {"ff", "nn", "sfc", "cc"}


def _module_export_definition(module_key: str):
    from app.services.asset_comparison.const import (
        CUSTOMER_CUSTOMER_SAVE_PATH,
        CUSTOMER_NOTES_SAVE_PATH,
        FINANCE_FINANCE_SAVE_PATH,
        FINANCE_NOTES_SAVE_PATH,
        NOTES_NOTES_SAVE_PATH,
        NOTES_SFC_SAVE_PATH,
        SFC_SFC_SAVE_PATH,
    )

    definitions = {
        "ff": ("Save_Check", Path(FINANCE_FINANCE_SAVE_PATH)),
        "nn": ("Save_Notes_Notes_Comparison", Path(NOTES_NOTES_SAVE_PATH)),
        "sfc": ("Save_SFC_SFC_Comparison", Path(SFC_SFC_SAVE_PATH)),
        "cc": (
            "Save_Customer_Customer_Comparison",
            Path(CUSTOMER_CUSTOMER_SAVE_PATH),
        ),
        "fn": ("Save_Finance_Notes_Comparison", Path(FINANCE_NOTES_SAVE_PATH)),
        "ns": ("Save_Notes_SFC_Comparison", Path(NOTES_SFC_SAVE_PATH)),
        "cn": ("Save_Customer_Notes_Comparison", Path(CUSTOMER_NOTES_SAVE_PATH)),
    }
    try:
        return definitions[module_key]
    except KeyError as exc:
        raise ValueError(f"未知的核对模块: {module_key}") from exc


def _write_no_difference_workbook(path: Path, label: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "對比結果"
    worksheet["A1"] = label
    worksheet["A2"] = "本次核对未发现差异"
    worksheet.column_dimensions["A"].width = 48
    workbook.save(path)


def _build_module_job_artifact(
    module_key: str,
    instance,
    result: dict,
    job_dir: Path,
) -> dict:
    method_name, legacy_path = _module_export_definition(module_key)
    target_path = job_dir / f"module-{module_key}.xlsx"
    temporary_path = target_path.with_suffix(".xlsx.tmp")

    with _legacy_export_lock:
        temporary_path.unlink(missing_ok=True)
        if result.get("has_diff"):
            legacy_path.unlink(missing_ok=True)
            getattr(instance, method_name)()
            if not legacy_path.is_file():
                raise RuntimeError(f"{result['label']}文件生成失败")
            shutil.copy2(legacy_path, temporary_path)
        else:
            _write_no_difference_workbook(temporary_path, result["label"])
        os.replace(temporary_path, target_path)

    return {
        "path": target_path.name,
        "filename": legacy_path.name,
        "content_type": (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        "size_bytes": target_path.stat().st_size,
    }


def _build_raw_job_artifact(summary: dict, job_dir: Path) -> dict:
    current_date = datetime.now()
    this_month_str = current_date.strftime("%Y%m")
    last_month_str = (current_date - relativedelta(months=1)).strftime("%Y%m")
    raw_buffer = _build_raw_data_xlsx(summary, this_month_str, last_month_str)
    if raw_buffer is None:
        raise RuntimeError("没有可供导出的原始数据")

    target_path = job_dir / "raw-data.xlsx"
    temporary_path = target_path.with_suffix(".xlsx.tmp")
    temporary_path.write_bytes(raw_buffer.getvalue())
    os.replace(temporary_path, target_path)
    return {
        "path": target_path.name,
        "filename": "原始数据.xlsx",
        "content_type": (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        "size_bytes": target_path.stat().st_size,
    }


def _execute_asset_comparison_job(
    job_id: str,
    inputs: dict[str, str],
    job_dir: Path,
    emit,
    is_cancel_requested,
):
    missing_inputs = [key for key, value in inputs.items() if not str(value).strip()]
    invalid_paths = [
        key
        for key, value in inputs.items()
        if str(value).strip() and not Path(value).is_file()
    ]
    if missing_inputs:
        raise ValueError(f"缺少输入文件: {', '.join(missing_inputs)}")
    if invalid_paths:
        raise ValueError(f"输入文件不存在: {', '.join(invalid_paths)}")

    emit("validation_ready")
    for module_key in ["ff", "sfc", "nn", "cc", "fn", "ns", "cn"]:
        emit("comparison_started", module_key=module_key)

    request = ComparisonRequest(**inputs)
    partial_summary = {}
    results_by_key = {}
    raw_future = None

    with ThreadPoolExecutor(
        max_workers=1,
        thread_name_prefix=f"asset-raw-{job_id[:8]}",
    ) as raw_executor:

        def on_complete(module_key, instance, info):
            nonlocal raw_future
            partial_summary[module_key] = instance
            result = info or {
                "key": module_key,
                "label": module_key,
                "has_diff": False,
                "msg": "异常: 核对模块没有返回结果",
            }
            results_by_key[module_key] = result
            if result.get("msg", "").startswith("异常:"):
                emit("comparison_failed", result=result)
            else:
                emit("comparison_ready", result=result)
                emit(
                    "artifact_building",
                    artifact_key=f"module_{module_key}",
                )
                try:
                    artifact = _build_module_job_artifact(
                        module_key,
                        instance,
                        result,
                        job_dir,
                    )
                    emit(
                        "artifact_ready",
                        artifact_key=f"module_{module_key}",
                        **artifact,
                    )
                except Exception as exc:
                    logger.exception(
                        f"asset module artifact failed: "
                        f"job_id={job_id} module={module_key} error={exc}"
                    )
                    emit(
                        "artifact_failed",
                        artifact_key=f"module_{module_key}",
                        error=str(exc),
                    )

            if raw_future is None and _RAW_SOURCE_MODULES.issubset(partial_summary):
                emit("artifact_building", artifact_key="raw_data_xlsx")
                raw_future = raw_executor.submit(
                    _build_raw_job_artifact,
                    {key: partial_summary[key] for key in _RAW_SOURCE_MODULES},
                    job_dir,
                )

        summary = run_comparisons(request, on_complete=on_complete)

        if is_cancel_requested():
            return {
                "summary": summary,
                "inputs": inputs,
                "results": results_by_key,
            }

        if raw_future is None:
            emit("artifact_building", artifact_key="raw_data_xlsx")
            raw_future = raw_executor.submit(
                _build_raw_job_artifact,
                summary,
                job_dir,
            )
        try:
            raw_artifact = raw_future.result()
            emit(
                "artifact_ready",
                artifact_key="raw_data_xlsx",
                **raw_artifact,
            )
        except Exception as exc:
            logger.exception(f"asset raw artifact failed: job_id={job_id} error={exc}")
            emit(
                "artifact_failed",
                artifact_key="raw_data_xlsx",
                error=str(exc),
            )

    return {
        "summary": summary,
        "inputs": inputs,
        "results": results_by_key,
    }


def _finalize_asset_comparison_job(
    job_id: str,
    runtime: dict,
    job_dir: Path,
    remarks: dict[str, str],
    reviews: dict[str, str],
) -> dict:
    effective_reviews = {
        key: reviews.get(key, "差異確認OK")
        for key in ["ff", "sfc", "nn", "cc", "fn", "ns", "cn"]
    }
    request = ComparisonRequest(
        **runtime["inputs"],
        remarks=remarks,
        reviews=effective_reviews,
    )
    raw_path = job_dir / "raw-data.xlsx"
    if not raw_path.is_file():
        raise RuntimeError("原始数据文件不存在")

    with _legacy_export_lock:
        content, filename = _build_complete_export(
            request,
            runtime["summary"],
            raw_data=raw_path.read_bytes(),
        )
    target_path = job_dir / "complete.zip"
    temporary_path = target_path.with_suffix(".zip.tmp")
    temporary_path.write_bytes(content)
    os.replace(temporary_path, target_path)
    return {
        "path": target_path.name,
        "filename": filename,
        "content_type": "application/zip",
        "size_bytes": target_path.stat().st_size,
    }


def _retry_asset_comparison_artifact(
    job_id: str,
    artifact_key: str,
    runtime: dict,
    job_dir: Path,
) -> dict:
    if artifact_key == "raw_data_xlsx":
        return _build_raw_job_artifact(runtime["summary"], job_dir)
    if artifact_key.startswith("module_"):
        module_key = artifact_key.removeprefix("module_")
        result = runtime["results"][module_key]
        instance = runtime["summary"][module_key]
        retried_result = None
        if result.get("msg", "").startswith("异常:"):
            task_by_key = {
                "ff": task_ff,
                "sfc": task_sfc,
                "nn": task_nn,
                "cc": task_cc,
                "fn": task_fn,
                "ns": task_ns,
                "cn": task_cn,
            }
            _, instance, result = task_by_key[module_key](
                ComparisonRequest(**runtime["inputs"])
            )
            if result is None or result.get("msg", "").startswith("异常:"):
                message = result.get("msg") if result else "核对模块没有返回结果"
                raise RuntimeError(message)
            runtime["summary"][module_key] = instance
            runtime["results"][module_key] = result
            retried_result = result

        artifact = _build_module_job_artifact(
            module_key,
            instance,
            result,
            job_dir,
        )
        if retried_result is not None:
            artifact["_comparison_result"] = retried_result
        return artifact
    raise ValueError(f"不支持重试的产物: {artifact_key}")


asset_comparison_job_manager = AssetComparisonJobManager(
    execute_job=_execute_asset_comparison_job,
    finalize_job=_finalize_asset_comparison_job,
    retry_artifact=_retry_asset_comparison_artifact,
)


def _raise_job_http_error(exc: Exception):
    if isinstance(exc, AssetComparisonJobExpiredError):
        raise HTTPException(status_code=410, detail="核对任务已过期") from exc
    if isinstance(exc, AssetComparisonJobNotFoundError):
        raise HTTPException(status_code=404, detail="核对任务不存在") from exc
    if isinstance(exc, AssetComparisonJobConflictError):
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    if isinstance(exc, AssetComparisonJobValidationError):
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    raise exc


@router.post("/jobs", status_code=202)
def create_asset_comparison_job(
    req: AssetComparisonJobCreate,
    current_user: User = Depends(require_permission("tool:use")),
    _: None = Depends(require_tool_enabled("asset-comparison")),
):
    payload = req.model_dump()
    client_request_id = payload.pop("clientRequestId")
    job, reused = asset_comparison_job_manager.create_job(
        user_id=current_user.id,
        client_request_id=client_request_id,
        inputs=payload,
    )
    return {
        **job,
        "reused": reused,
        "pollAfterMs": 1000,
    }


@router.get("/jobs/{job_id}")
def get_asset_comparison_job(
    job_id: str,
    current_user: User = Depends(require_permission("tool:use")),
    _: None = Depends(require_tool_enabled("asset-comparison")),
):
    try:
        return asset_comparison_job_manager.get_job(
            user_id=current_user.id,
            job_id=job_id,
        )
    except Exception as exc:
        _raise_job_http_error(exc)


@router.patch("/jobs/{job_id}/annotations")
def update_asset_comparison_annotations(
    job_id: str,
    req: AssetComparisonAnnotationsUpdate,
    current_user: User = Depends(require_permission("tool:use")),
    _: None = Depends(require_tool_enabled("asset-comparison")),
):
    try:
        return asset_comparison_job_manager.update_annotations(
            user_id=current_user.id,
            job_id=job_id,
            expected_revision=req.expectedRevision,
            remarks=req.remarks,
            reviews=req.reviews,
        )
    except Exception as exc:
        _raise_job_http_error(exc)


@router.post("/jobs/{job_id}/finalize", status_code=202)
def finalize_asset_comparison_job(
    job_id: str,
    current_user: User = Depends(require_permission("tool:use")),
    _: None = Depends(require_tool_enabled("asset-comparison")),
):
    try:
        return asset_comparison_job_manager.finalize(
            user_id=current_user.id,
            job_id=job_id,
        )
    except Exception as exc:
        _raise_job_http_error(exc)


@router.post("/jobs/{job_id}/artifacts/{artifact_key}/retry", status_code=202)
def retry_asset_comparison_artifact(
    job_id: str,
    artifact_key: str,
    current_user: User = Depends(require_permission("tool:use")),
    _: None = Depends(require_tool_enabled("asset-comparison")),
):
    try:
        return asset_comparison_job_manager.retry(
            user_id=current_user.id,
            job_id=job_id,
            artifact_key=artifact_key,
        )
    except Exception as exc:
        _raise_job_http_error(exc)


@router.get("/jobs/{job_id}/artifacts/{artifact_key}")
def download_asset_comparison_artifact(
    job_id: str,
    artifact_key: str,
    current_user: User = Depends(require_permission("tool:use")),
    _: None = Depends(require_tool_enabled("asset-comparison")),
):
    try:
        path, filename, content_type = asset_comparison_job_manager.open_artifact(
            user_id=current_user.id,
            job_id=job_id,
            artifact_key=artifact_key,
        )
        return FileResponse(
            path=path,
            media_type=content_type,
            filename=filename,
        )
    except Exception as exc:
        _raise_job_http_error(exc)


@router.delete("/jobs/{job_id}")
def cancel_asset_comparison_job(
    job_id: str,
    current_user: User = Depends(require_permission("tool:use")),
    _: None = Depends(require_tool_enabled("asset-comparison")),
):
    try:
        return asset_comparison_job_manager.cancel(
            user_id=current_user.id,
            job_id=job_id,
        )
    except Exception as exc:
        _raise_job_http_error(exc)


@router.delete("/jobs/{job_id}/purge")
def purge_asset_comparison_job(
    job_id: str,
    current_user: User = Depends(require_permission("tool:use")),
    _: None = Depends(require_tool_enabled("asset-comparison")),
):
    try:
        asset_comparison_job_manager.purge(
            user_id=current_user.id,
            job_id=job_id,
        )
        return {"status": "success"}
    except Exception as exc:
        _raise_job_http_error(exc)


def _convert_diff_dict_to_dataframe(diff_dict, comment: str = "", sheet_name: str = ""):
    """将 diff_dict 转换为 DataFrame 格式，用于 PDF 生成"""
    try:
        all_data = []
        for category, df in diff_dict.items():
            if df is None:
                continue

            if hasattr(df, "collect"):
                df = df.collect()

            if df is None or df.is_empty():
                continue

            # 添加分类列
            df_copy = df.with_columns(pl.lit(category).alias("分类"))
            df_copy = df_copy.select(
                ["分类", *[c for c in df_copy.columns if c != "分类"]]
            )

            # 检查是否为"7-Notes客户资产 VS 客户系统资产"模块，如果是则去掉"Model Number"和"资产编号"列
            if (
                "Notes客户资产" in category
                or "7-Notes客户资产" in sheet_name
                or "Notes客户资产" in sheet_name
            ):
                cols_to_drop = [
                    col
                    for col in ["Model Number", "资产编号", "資產編號"]
                    if col in df_copy.columns
                ]
                if cols_to_drop:
                    df_copy = df_copy.drop(cols_to_drop)

            all_data.append(df_copy)

        if all_data:
            combined_df = pl.concat(all_data, how="diagonal")
            return combined_df, comment
        else:
            logger.warning(f"没有任何有效数据可以合并: {sheet_name}")
            return None, comment

    except Exception as e:
        logger.error(f"转换 diff_dict 到 DataFrame 失败 ({sheet_name}): {e}")
        return None, comment


def estimate_width_by_font(value, font_size=12):
    value_str = str(value)
    width = 0
    for ch in value_str:
        if "\u4e00" <= ch <= "\u9fff":
            width += 2.1
        elif ch.isupper():
            width += 1.5
        else:
            width += 1
    return width * (font_size / 11)


def auto_adjust_row_height(ws, row, col, content, font_size=11, max_chars_per_line=50):
    lines_by_newline = content.split("\n")
    total_lines = 0
    for line in lines_by_newline:
        if len(line) == 0:
            total_lines += 1
        else:
            total_lines += (len(line) // max_chars_per_line) + 1
    ws.row_dimensions[row].height = max(20, total_lines * 15)


def write_comparison_to_sheet(diff_dict, comment: str, ws):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    start_col = 3
    row = 1
    font = Font(name="Arial", size=12)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side("thin"),
    )
    fill_title_bg = PatternFill(fill_type="solid", fgColor="D9E1F2")

    max_col_widths = dict()
    categories = list(diff_dict.keys())

    for idx, category in enumerate(categories):
        df = diff_dict[category]
        if df is None or df.is_empty():
            continue

        title_text = f"【{category}】（{len(df)}笔）"
        end_col = start_col + len(df.columns)
        ws.merge_cells(
            start_row=row, start_column=start_col, end_row=row, end_column=end_col
        )
        cell = ws.cell(row=row, column=start_col, value=title_text)
        cell.font = font
        cell.alignment = align_center
        cell.border = border
        cell.fill = fill_title_bg
        for col in range(start_col + 1, end_col + 1):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).fill = fill_title_bg
        max_col_widths[start_col] = max(
            max_col_widths.get(start_col, 0),
            estimate_width_by_font(title_text, font.size),
        )
        ws.row_dimensions[row].height = 25
        row += 1

        if comment.strip():
            remark_bg_fill = PatternFill(
                start_color="DEDEDE", end_color="DEDEDE", fill_type="solid"
            )
            remark_cell = ws.cell(row=row, column=start_col, value="備註：")
            remark_cell.border = border
            remark_cell.font = font
            remark_cell.alignment = align_center
            remark_cell.fill = remark_bg_fill
            max_col_widths[start_col] = max(
                max_col_widths.get(start_col, 0),
                estimate_width_by_font("備註：", font.size),
            )

            ws.merge_cells(
                start_row=row,
                start_column=start_col + 1,
                end_row=row,
                end_column=start_col + 3,
            )
            cell = ws.cell(row=row, column=start_col + 1, value=comment.strip())
            cell.font = font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = border
            cell.fill = remark_bg_fill

            for col in range(start_col + 1, start_col + 4):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.fill = remark_bg_fill
                max_col_widths[col] = max(
                    max_col_widths.get(col, 0),
                    estimate_width_by_font(comment.strip(), font.size),
                )

            auto_adjust_row_height(ws, row, start_col + 1, comment.strip())
            row += 1

        ws.cell(row=row, column=start_col, value="No.").font = font
        ws.cell(row=row, column=start_col).alignment = align_center
        ws.cell(row=row, column=start_col).border = border
        max_col_widths[start_col] = max(
            max_col_widths.get(start_col, 0), estimate_width_by_font("No.", font.size)
        )

        for col_idx, col_name in enumerate(df.columns, start=start_col + 1):
            cell = ws.cell(row=row, column=col_idx, value=col_name)
            cell.font = font
            cell.alignment = align_center
            cell.border = border
            max_col_widths[col_idx] = max(
                max_col_widths.get(col_idx, 0),
                estimate_width_by_font(col_name, font.size),
            )
        row += 1

        for i, row_vals in enumerate(df.iter_rows(named=False), start=1):
            cell = ws.cell(row=row, column=start_col, value=i)
            cell.font = font
            cell.alignment = align_center
            cell.border = border
            max_col_widths[start_col] = max(
                max_col_widths.get(start_col, 0),
                estimate_width_by_font(str(i), font.size),
            )

            for col_idx, value in enumerate(row_vals, start=start_col + 1):
                value_str = str(value)
                cell = ws.cell(row=row, column=col_idx, value=value_str)
                cell.font = font
                cell.alignment = align_center
                cell.border = border
                max_col_widths[col_idx] = max(
                    max_col_widths.get(col_idx, 0),
                    estimate_width_by_font(value_str, font.size),
                )
            row += 1

        if (
            category == "本月新增"
            and idx + 1 < len(categories)
            and categories[idx + 1] == "本月减少"
        ):
            row += 3
        else:
            row += 1

    for col_idx, width in max_col_widths.items():
        col_letter = get_column_letter(col_idx)
        adjusted_width = min(width * 1.2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width


def _normalize_raw_data_line_breaks(df: pl.DataFrame) -> pl.DataFrame:
    """规范化备注说明列的换行符，避免导出后显示 OOXML 控制字符转义"""
    if df is None or df.is_empty():
        return df
    remark_columns = {
        "备注说明",
        "备注説明",
        "備注说明",
        "備注説明",
        "備註說明",
        "備註説明",
    }
    target_columns = [
        col
        for col in remark_columns.intersection(df.columns)
        if df[col].dtype in (pl.String, pl.Utf8)
    ]
    if not target_columns:
        return df

    exprs = [
        pl.col(column).str.replace_all("\r\n", "\n").str.replace_all("\r", "\n")
        for column in target_columns
    ]
    return df.with_columns(exprs)


def _format_cell_value(v):
    if v is None:
        return ""
    if isinstance(v, (int, float, bool)):
        return v
    return str(v)


def _clean_df_for_rustpy(df: pl.DataFrame) -> pl.DataFrame:
    """清理字符串列中的 \\ufffd (替换字符) 避免 rustpy-xlsxwriter 字符串索引越界 panic"""
    if df is None or df.is_empty():
        return df
    exprs = [
        pl.col(col).str.replace_all("\ufffd", "?")
        for col in df.columns
        if df[col].dtype in (pl.String, pl.Utf8)
    ]
    if exprs:
        return df.with_columns(exprs)
    return df


def _extract_data_sheets(
    summary: dict, this_month_str: str, last_month_str: str
) -> dict:
    """从比对结果中提取各模块的原始数据 DataFrame 映射"""
    data_sheets = {}

    def _get_df(obj, attr_name):
        val = getattr(obj, attr_name, None)
        if val is None:
            return None
        if hasattr(val, "collect"):
            val = val.collect()
        return val

    modules = [
        ("ff", "this_Finance_data", "last_Finance_data", "财务"),
        ("nn", "this_Notes_data", "last_Notes_data", "Notes"),
        ("sfc", "this_SFC_data", "last_SFC_data", "SFC"),
        ("cc", "this_Customer_data", "last_Customer_data", "客户"),
    ]

    for key, this_attr, last_attr, label in modules:
        obj = summary.get(key)
        if not obj:
            continue
        df_this = _get_df(obj, this_attr)
        df_last = _get_df(obj, last_attr)
        if df_this is not None and not df_this.is_empty():
            data_sheets[f"{this_month_str}-{label}"] = df_this
        if df_last is not None and not df_last.is_empty():
            data_sheets[f"{last_month_str}-{label}"] = df_last

    return data_sheets


def _build_raw_data_xlsx_rustpy(
    summary: dict, this_month_str: str, last_month_str: str
) -> io.BytesIO | None:
    """使用 rustpy-xlsxwriter 从比对结果中提取原始数据，高效生成原始数据.xlsx 到 BytesIO 缓冲"""
    data_sheets = _extract_data_sheets(summary, this_month_str, last_month_str)
    if not data_sheets:
        return None

    import rustpy_xlsxwriter as rx

    buf = io.BytesIO()
    sheet_dfs = []
    for sheet_name, df in data_sheets.items():
        safe_name = sheet_name[:31]
        normalized_df = _normalize_raw_data_line_breaks(df)
        cleaned_df = _clean_df_for_rustpy(normalized_df)
        sheet_dfs.append((safe_name, cleaned_df))

    with rx.FastExcel(buf, autofit=False) as fe:
        for safe_name, df in sheet_dfs:
            fe.sheet(safe_name, df)

    buf.seek(0)
    return buf


def _build_raw_data_xlsx(
    summary: dict, this_month_str: str, last_month_str: str
) -> io.BytesIO | None:
    """从比对结果中提取原始数据，生成原始数据.xlsx 到 BytesIO 缓冲 (优先使用 rustpy-xlsxwriter)"""
    try:
        buf = _build_raw_data_xlsx_rustpy(summary, this_month_str, last_month_str)
        if buf is not None:
            return buf
    except Exception as e:
        logger.warning(f"rustpy_xlsxwriter 导出失败，降级回退到 xlsxwriter: {e}")

    data_sheets = _extract_data_sheets(summary, this_month_str, last_month_str)
    if not data_sheets:
        return None

    sheet_dfs = []
    for sheet_name, df in data_sheets.items():
        safe_name = sheet_name[:31]
        normalized_df = _normalize_raw_data_line_breaks(df)
        sheet_dfs.append((safe_name, normalized_df))

    buf = io.BytesIO()
    workbook = xlsxwriter.Workbook(buf, {"constant_memory": True})
    for safe_name, df in sheet_dfs:
        worksheet = workbook.add_worksheet(safe_name)
        worksheet.write_row(0, 0, list(df.columns))
        for r_idx, row in enumerate(df.iter_rows(named=False), start=1):
            worksheet.write_row(r_idx, 0, [_format_cell_value(v) for v in row])
    workbook.close()
    buf.seek(0)
    return buf


def _build_complete_export(
    req: ComparisonRequest,
    summary: dict,
    raw_data: bytes | None = None,
) -> tuple[bytes, str]:
    missing_remark_labels = [
        result.get("label", result.get("key", "未知模块"))
        for result in summary.get("results_info", [])
        if result.get("has_diff")
        and not str(req.remarks.get(result.get("key", ""), "")).strip()
    ]
    if missing_remark_labels:
        labels = "、".join(
            label.replace("【", "").replace("】", "") for label in missing_remark_labels
        )
        raise HTTPException(
            status_code=422,
            detail=f"请先填写异常原因再生成对比总结与 PDF：{labels}",
        )

    request_started_at = perf_counter()
    try:
        ff = summary.get("ff")
        nn = summary.get("nn")
        sfc = summary.get("sfc")
        cc = summary.get("cc")
        fn = summary.get("fn")
        ns = summary.get("ns")
        cn = summary.get("cn")

        from app.services.asset_comparison.const import SAVE_CHECK_PATH

        if not os.path.exists(SAVE_CHECK_PATH):
            os.makedirs(SAVE_CHECK_PATH, exist_ok=True)

        current_date = datetime.now()
        this_month_str = current_date.strftime("%Y%m")
        last_month_date = current_date - relativedelta(months=1)
        last_month_str = last_month_date.strftime("%Y%m")
        save_all_path = os.path.join(
            SAVE_CHECK_PATH, f"TE&PE资产对比_{this_month_str}对比总结.xlsx"
        )
        save_pdf_path = os.path.join(
            SAVE_CHECK_PATH, f"TE&PE资产对比_{this_month_str}对比总结.pdf"
        )

        if os.path.exists(save_all_path):
            os.remove(save_all_path)

        # Step 1: 结果表 Excel 模版导出
        step1_started_at = perf_counter()
        try:
            create_excel_template(save_all_path)
        except Exception as err:
            logger.warning(f"Template creation err: {err}")

        wb = load_workbook(save_all_path)
        ws = wb["差异总结"]

        this_Finance_Custodian_assets = _safe_len(
            getattr(ff, "this_Custodian_assets", [])
        )
        last_Finance_Custodian_assets = _safe_len(
            getattr(ff, "last_Custodian_assets", [])
        )
        this_Notes_assets = _safe_len(getattr(nn, "this_assets_filtered", []))
        last_Notes_assets = _safe_len(getattr(nn, "last_assets_filtered", []))
        this_Notes_NO_assets = getattr(nn, "this_invalid_all_rows", 0)
        last_Notes_NO_assets = getattr(nn, "last_invalid_all_rows", 0)
        this_SFC_assets = _safe_len(getattr(sfc, "this_SFC_assets", []))
        last_SFC_assets = _safe_len(getattr(sfc, "last_SFC_assets", []))
        this_Customer_assets = _safe_len(getattr(cc, "this_Customer_assets", []))
        last_Customer_assets = _safe_len(getattr(cc, "last_Customer_assets", []))
        this_Notes_Customer_assets = _safe_len(getattr(cn, "this_Notes_assets", []))

        data_rows = [
            [
                last_Finance_Custodian_assets,
                this_Finance_Custodian_assets,
                this_Finance_Custodian_assets - last_Finance_Custodian_assets,
                this_Finance_Custodian_assets - this_Notes_assets,
                "/",
                "/",
            ],
            [
                last_Notes_assets,
                this_Notes_assets,
                this_Notes_assets - last_Notes_assets,
                "",
                this_Notes_assets - this_SFC_assets,
                "/",
            ],
            [
                last_SFC_assets,
                this_SFC_assets,
                this_SFC_assets - last_SFC_assets,
                "/",
                "",
                "/",
            ],
            [
                last_Customer_assets,
                this_Customer_assets,
                this_Customer_assets - last_Customer_assets,
                "/",
                "/",
                this_Notes_Customer_assets - this_Customer_assets,
            ],
            ["/", this_Notes_Customer_assets, "/", "/", "/", ""],
            [
                last_Notes_NO_assets,
                this_Notes_NO_assets,
                this_Notes_NO_assets - last_Notes_NO_assets,
                "/",
                "/",
                "/",
            ],
        ]

        start_row = 5
        start_col = 2
        for row_idx, row_data in enumerate(data_rows):
            for col_idx, value in enumerate(row_data):
                col_letter = get_column_letter(start_col + col_idx)
                coord = f"{col_letter}{start_row + row_idx}"
                ws[coord] = value

        ws.merge_cells("A3:G3")
        ws.merge_cells("E5:E6")
        ws.merge_cells("F6:F7")
        ws.merge_cells("G8:G9")
        ws.merge_cells("A12:G12")

        apply_review_colors(ws, req.reviews)

        remark_mapping = [
            (13, "ff"),
            (14, "nn"),
            (15, "sfc"),
            (16, "cc"),
            (17, "fn"),
            (18, "ns"),
            (19, "cn"),
        ]
        for row_idx, r_key in remark_mapping:
            ws.merge_cells(f"B{row_idx}:G{row_idx}")
            cell = ws.cell(row=row_idx, column=2)
            cell.value = req.remarks.get(r_key, "")
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )

        wb.save(save_all_path)

        comparisons = []

        # 1-财务 VS 财务
        if ff:
            ff_dict = {}
            if getattr(ff, "new_Custodian_assets", []):
                df = ff.this_Finance_data
                if hasattr(df, "collect"):
                    df = df.collect()
                if df is not None and not df.is_empty():
                    ff_dict["依保管人新增"] = (
                        df.filter(
                            pl.col("資產編號").is_in(list(ff.new_Custodian_assets))
                        )
                        .select(["資產名稱", "資產編號", "保管人員"])
                        .unique(subset=["資產編號"])
                    )
            if getattr(ff, "removed_Custodian_assets", []):
                df = ff.last_Finance_data
                if hasattr(df, "collect"):
                    df = df.collect()
                if df is not None and not df.is_empty():
                    ff_dict["依保管人减少"] = (
                        df.filter(
                            pl.col("資產編號").is_in(list(ff.removed_Custodian_assets))
                        )
                        .select(["資產名稱", "資產編號", "保管人員"])
                        .unique(subset=["資產編號"])
                    )
            if getattr(ff, "new_Department_assets", []):
                df = ff.this_Finance_data
                if hasattr(df, "collect"):
                    df = df.collect()
                if df is not None and not df.is_empty():
                    ff_dict["依部门新增"] = (
                        df.filter(
                            pl.col("資產編號").is_in(list(ff.new_Department_assets))
                        )
                        .select(["資產名稱", "資產編號", "資產所屬部門代號"])
                        .unique(subset=["資產編號"])
                    )
            if getattr(ff, "removed_Department_assets", []):
                df = ff.last_Finance_data
                if hasattr(df, "collect"):
                    df = df.collect()
                if df is not None and not df.is_empty():
                    ff_dict["依部门减少"] = (
                        df.filter(
                            pl.col("資產編號").is_in(list(ff.removed_Department_assets))
                        )
                        .select(["資產名稱", "資產編號", "資產所屬部門代號"])
                        .unique(subset=["資產編號"])
                    )
            if ff_dict:
                comparisons.append(
                    ("1-财务 VS 财务", ff_dict, req.remarks.get("ff", ""))
                )

        # 2-Notes VS Notes
        if nn:
            nn_dict = {}
            if getattr(nn, "new_assets", []):
                df = nn.this_Notes_data
                if hasattr(df, "collect"):
                    df = df.collect()
                if df is not None and not df.is_empty():
                    nn_dict["本月新增"] = (
                        df.filter(pl.col("資產編號").is_in(list(nn.new_assets)))
                        .select(["資產名稱", "資產編號", "保管人"])
                        .unique()
                    )
            if getattr(nn, "removed_assets", []):
                df = nn.last_Notes_data
                if hasattr(df, "collect"):
                    df = df.collect()
                if df is not None and not df.is_empty():
                    nn_dict["本月减少"] = (
                        df.filter(pl.col("資產編號").is_in(list(nn.removed_assets)))
                        .select(["資產名稱", "資產編號", "保管人"])
                        .unique()
                    )
            if getattr(nn, "new_No_assets", []):
                df = nn.this_Notes_data
                if hasattr(df, "collect"):
                    df = df.collect()
                if df is not None and not df.is_empty():
                    try:
                        nn_dict["无资产记录-本月新增"] = df.filter(
                            pl.col("機身SN").is_in(list(nn.new_No_assets))
                        )
                    except Exception:
                        pass
            if getattr(nn, "removed_No_assets", []):
                df = nn.last_Notes_data
                if hasattr(df, "collect"):
                    df = df.collect()
                if df is not None and not df.is_empty():
                    try:
                        nn_dict["无资产记录-本月减少"] = df.filter(
                            pl.col("機身SN").is_in(list(nn.removed_No_assets))
                        )
                    except Exception:
                        pass
            if nn_dict:
                comparisons.append(
                    ("2-Notes VS Notes", nn_dict, req.remarks.get("nn", ""))
                )

        # 3-SFC VS SFC
        if sfc:
            ss_dict = {}
            if getattr(sfc, "new_assets", []):
                df = sfc.this_SFC_data
                if hasattr(df, "collect"):
                    df = df.collect()
                if df is not None and not df.is_empty():
                    ss_dict["本月新增"] = (
                        df.filter(pl.col("资产编号").is_in(list(sfc.new_assets)))
                        .select(["设备名称", "资产编号", "保管人"])
                        .unique()
                    )
            if getattr(sfc, "removed_assets", []):
                df = sfc.last_SFC_data
                if hasattr(df, "collect"):
                    df = df.collect()
                if df is not None and not df.is_empty():
                    ss_dict["本月减少"] = (
                        df.filter(pl.col("资产编号").is_in(list(sfc.removed_assets)))
                        .select(["设备名称", "资产编号", "保管人"])
                        .unique()
                    )
            if ss_dict:
                comparisons.append(
                    ("3-SFC VS SFC", ss_dict, req.remarks.get("sfc", ""))
                )

        # 4-客户资产 VS 客户资产
        if cc:
            cc_dict = {}
            if getattr(cc, "new_Customer_assets", []):
                df = cc.this_Customer_data
                if hasattr(df, "collect"):
                    df = df.collect()
                if df is not None and not df.is_empty():
                    cc_dict["本月新增"] = (
                        df.filter(
                            pl.col("Asset ID").is_in(list(cc.new_Customer_assets))
                        )
                        .select(["DRI", "Asset ID", "RFID"])
                        .unique()
                    )
            if getattr(cc, "removed_Customer_assets", []):
                df = cc.last_Customer_data
                if hasattr(df, "collect"):
                    df = df.collect()
                if df is not None and not df.is_empty():
                    cc_dict["本月减少"] = (
                        df.filter(
                            pl.col("Asset ID").is_in(list(cc.removed_Customer_assets))
                        )
                        .select(["DRI", "Asset ID", "RFID"])
                        .unique()
                    )
            if cc_dict:
                comparisons.append(
                    ("4-客户资产 VS 客户资产", cc_dict, req.remarks.get("cc", ""))
                )

        # 5-财务 VS Notes
        if fn:
            fn_dict = {}
            if getattr(fn, "new_assets", []):
                df = fn.Notes_data
                if hasattr(df, "collect"):
                    df = df.collect()
                if df is not None and not df.is_empty():
                    fn_dict["Notes比财务新增资产"] = (
                        df.filter(pl.col("資產編號").is_in(list(fn.new_assets)))
                        .select(["資產名稱", "資產編號", "保管人"])
                        .unique()
                    )
            if getattr(fn, "removed_assets", []):
                df = fn.Finance_data
                if hasattr(df, "collect"):
                    df = df.collect()
                if df is not None and not df.is_empty():
                    fn_dict["Notes比财务减少资产"] = (
                        df.filter(pl.col("資產編號").is_in(list(fn.removed_assets)))
                        .select(["資產名稱", "資產編號", "保管人員"])
                        .unique()
                    )
            if fn_dict:
                comparisons.append(
                    ("5-财务 VS Notes", fn_dict, req.remarks.get("fn", ""))
                )

        # 6-Notes VS SFC
        if ns:
            ns_dict = {}
            if getattr(ns, "Notes_new_assets", []):
                df = ns.this_Notes_data
                if hasattr(df, "collect"):
                    df = df.collect()
                if df is not None and not df.is_empty():
                    ns_dict["Notes有且SFC无"] = (
                        df.filter(pl.col("資產編號").is_in(list(ns.Notes_new_assets)))
                        .select(["資產名稱", "資產編號", "保管人"])
                        .unique()
                    )
            if getattr(ns, "Notes_removed_assets", []):
                df = ns.this_SFC_data
                if hasattr(df, "collect"):
                    df = df.collect()
                if df is not None and not df.is_empty():
                    ns_dict["SFC有且Notes无"] = (
                        df.filter(
                            pl.col("资产编号").is_in(list(ns.Notes_removed_assets))
                        )
                        .select(["设备名称", "资产编号", "保管人"])
                        .unique()
                    )
            if ns_dict:
                comparisons.append(
                    ("6-Notes VS SFC", ns_dict, req.remarks.get("ns", ""))
                )

        # 7-Notes客户资产 VS 客户系统资产
        if cn:
            cn_dict = {}
            if getattr(cn, "remove_assets", []):
                df = cn.this_Customer_data
                if hasattr(df, "collect"):
                    df = df.collect()
                if df is not None and not df.is_empty():
                    cn_dict["客户有且Notes无"] = (
                        df.filter(pl.col("RFID").is_in(list(cn.remove_assets)))
                        .select(["DRI", "Asset ID", "RFID"])
                        .unique()
                    )
            if getattr(cn, "new_assets", []):
                df = cn.this_Notes_data
                if hasattr(df, "collect"):
                    df = df.collect()
                if df is not None and not df.is_empty():
                    cn_dict["Notes有且客户无"] = df.filter(
                        pl.col("RFID（Tag）").is_in(list(cn.new_assets))
                    )
            if cn_dict:
                comparisons.append(
                    (
                        "7-Notes客户资产 VS 客户系统资产",
                        cn_dict,
                        req.remarks.get("cn", ""),
                    )
                )

        sheet_data_dict = {}
        wb = load_workbook(save_all_path)
        for sheet_name, diff_dict, comment in comparisons:
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
            ws_comp = wb[sheet_name]

            if sheet_name == "1-财务 VS 财务":
                excel_diff_dict = {
                    key: value
                    for key, value in diff_dict.items()
                    if key.startswith("依保管人")
                }
                write_comparison_to_sheet(excel_diff_dict, comment, ws_comp)
            else:
                write_comparison_to_sheet(diff_dict, comment, ws_comp)

            # 收集 sheet 数据用于 PDF 生成
            if diff_dict:
                merged_df, comment_text = _convert_diff_dict_to_dataframe(
                    diff_dict, comment, sheet_name
                )
                if merged_df is not None and not merged_df.is_empty():
                    sheet_data_dict[sheet_name] = (merged_df, comment_text)

        wb.save(save_all_path)
        step1_elapsed = perf_counter() - step1_started_at
        logger.info(f"Step 1: 结果表 Excel 模版导出完成，耗时 {step1_elapsed:.3f}s")

        # Step 2: PDF 导出会签表
        step2_started_at = perf_counter()
        pdf_ok = False
        summary_pdf_path = os.path.join(
            SAVE_CHECK_PATH, f"TE&PE资产对比_{this_month_str}对比总结_summary.pdf"
        )
        detail_pdf_path = os.path.join(
            SAVE_CHECK_PATH, f"TE&PE资产对比_{this_month_str}对比总结_detail.pdf"
        )

        try:
            summary_ok = False
            try:
                summary_ok = excel_sheet_to_pdf(
                    save_all_path, "差异总结", summary_pdf_path
                )
            except Exception as sum_e:
                logger.warning(f"总结PDF生成失败: {sum_e}")

            detail_ok = False
            if sheet_data_dict:
                try:
                    detail_ok = create_pdf_from_sheets(
                        sheet_data_dict,
                        detail_pdf_path,
                        f"TE&PE资产对比_{this_month_str}详细差异",
                    )
                except Exception as det_e:
                    logger.warning(f"明细PDF生成失败: {det_e}")

            if (
                summary_ok
                and detail_ok
                and os.path.exists(summary_pdf_path)
                and os.path.exists(detail_pdf_path)
            ):
                try:
                    merger = PdfMerger()
                    merger.append(summary_pdf_path)  # 差异总结
                    merger.append(detail_pdf_path)  # 详细差异
                    merger.write(save_pdf_path)
                    merger.close()
                    pdf_ok = True
                except Exception as merge_e:
                    logger.warning(f"PDF合并失败: {merge_e}")

            if not pdf_ok and summary_ok and os.path.exists(summary_pdf_path):
                import shutil

                shutil.copy2(summary_pdf_path, save_pdf_path)
                pdf_ok = True

        except Exception as pdf_e:
            logger.warning(f"PDF生成处理出错: {pdf_e}")

        finally:
            for tmp in [summary_pdf_path, detail_pdf_path]:
                if os.path.exists(tmp):
                    try:
                        os.remove(tmp)
                    except Exception as clean_e:
                        logger.warning(f"清理临时文件失败 ({tmp}): {clean_e}")

        step2_elapsed = perf_counter() - step2_started_at
        logger.info(
            f"Step 2: PDF 导出会签表完成，耗时 {step2_elapsed:.3f}s (ok={pdf_ok})"
        )

        # Step 3: 原始数据.xlsx 生成
        step3_started_at = perf_counter()
        raw_buf = (
            io.BytesIO(raw_data)
            if raw_data is not None
            else _build_raw_data_xlsx(summary, this_month_str, last_month_str)
        )
        step3_elapsed = perf_counter() - step3_started_at
        raw_size = raw_buf.getbuffer().nbytes if raw_buf else 0
        logger.info(
            f"Step 3: 原始数据.xlsx 生成完成，耗时 {step3_elapsed:.3f}s，文件大小 {raw_size} 字节"
        )

        # Step 4: 导出包 ZIP 压缩打包
        step4_started_at = perf_counter()
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            # 对比总结 XLSX
            zf.write(save_all_path, f"TE&PE资产对比_{this_month_str}对比总结.xlsx")
            # PDF（如有）
            if pdf_ok:
                zf.write(save_pdf_path, f"TE&PE资产对比_{this_month_str}对比总结.pdf")
            # 原始数据 XLSX
            if raw_buf:
                zf.writestr("原始数据.xlsx", raw_buf.getvalue())
        zip_buf.seek(0)
        step4_elapsed = perf_counter() - step4_started_at
        zip_size = zip_buf.getbuffer().nbytes
        logger.info(
            f"Step 4: 导出包 ZIP 压缩打包完成，耗时 {step4_elapsed:.3f}s，文件大小 {zip_size} 字节"
        )

        zip_filename = f"TE&PE资产对比_{this_month_str}.zip"
        total_elapsed = perf_counter() - request_started_at
        logger.info(f"导出包处理完成，总耗时 {total_elapsed:.3f}s")
        return zip_buf.getvalue(), zip_filename

    except Exception as e:
        elapsed = perf_counter() - request_started_at
        logger.error(
            f"导出包处理出错，耗时 {elapsed:.3f}s: {e!r}\n{traceback.format_exc()}"
        )
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/save")
def save_results(
    req: ComparisonRequest,
    current_user: User = Depends(require_permission("tool:use")),
    _: None = Depends(require_tool_enabled("asset-comparison")),
):
    stage_started_at = perf_counter()
    logger.info("资产对比 save 开始: run_comparisons")
    summary = run_comparisons(req)
    logger.info(
        f"资产对比 run_comparisons 完成，耗时 {perf_counter() - stage_started_at:.3f}s"
    )
    content, filename = _build_complete_export(req, summary)
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/zip",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"
        },
    )


@router.post("/export/{module}")
def export_single_module(
    module: str,
    req: ComparisonRequest,
    current_user: User = Depends(require_permission("tool:use")),
    _: None = Depends(require_tool_enabled("asset-comparison")),
):
    try:
        summary = run_comparisons(req)
        from app.services.asset_comparison.const import (
            CUSTOMER_CUSTOMER_SAVE_PATH,
            CUSTOMER_NOTES_SAVE_PATH,
            FINANCE_FINANCE_SAVE_PATH,
            FINANCE_NOTES_SAVE_PATH,
            NOTES_NOTES_SAVE_PATH,
            NOTES_SFC_SAVE_PATH,
            SFC_SFC_SAVE_PATH,
        )

        if module == "ff":
            summary["ff"].Save_Check()
            return {
                "status": "success",
                "message": f"财务对比单独导出成功:\n{FINANCE_FINANCE_SAVE_PATH}",
            }

        elif module == "nn":
            summary["nn"].Save_Notes_Notes_Comparison()
            return {
                "status": "success",
                "message": f"Notes对比单独导出成功:\n{NOTES_NOTES_SAVE_PATH}",
            }

        elif module == "sfc":
            summary["sfc"].Save_SFC_SFC_Comparison()
            return {
                "status": "success",
                "message": f"SFC对比单独导出成功:\n{SFC_SFC_SAVE_PATH}",
            }

        elif module == "cc":
            summary["cc"].Save_Customer_Customer_Comparison()
            return {
                "status": "success",
                "message": f"客户对比单独导出成功:\n{CUSTOMER_CUSTOMER_SAVE_PATH}",
            }

        elif module == "fn":
            summary["fn"].Save_Finance_Notes_Comparison()
            return {
                "status": "success",
                "message": f"财务-Notes对比单独导出成功:\n{FINANCE_NOTES_SAVE_PATH}",
            }

        elif module == "ns":
            summary["ns"].Save_Notes_SFC_Comparison()
            return {
                "status": "success",
                "message": f"Notes-SFC对比单独导出成功:\n{NOTES_SFC_SAVE_PATH}",
            }

        elif module == "cn":
            summary["cn"].Save_Customer_Notes_Comparison()
            return {
                "status": "success",
                "message": f"客户-Notes对比单独导出成功:\n{CUSTOMER_NOTES_SAVE_PATH}",
            }
        else:
            return {"status": "error", "message": "未知的导出模块"}

    except Exception as e:
        return {
            "status": "error",
            "message": str(e),
            "errors": [traceback.format_exc()],
        }

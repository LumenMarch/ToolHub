import os  # noqa: E402, I001, UP015, F401
from datetime import datetime  # noqa: E402, I001, UP015, F401

import openpyxl  # noqa: E402, I001, UP015, F401
import polars as pl  # noqa: E402, I001, UP015, F401
from app.services.asset_comparison.const import NOTES_NOTES_SAVE_PATH  # noqa: E402, I001, UP015, F401
from app.services.asset_comparison.excel_writer import new_workbook, safe_cell
from loguru import logger  # noqa: E402, I001, UP015, F401

No_CheckRFID = ["A1300011C5C3", "A13000103933", "A1300010E606"]


class Notes_Notes:
    def __init__(self):
        self.input_catalog = None
        self.this_invalid_all_rows = None
        self.last_assets_filtered = None
        self.last_All_Notes_ = None
        self.this_assets_filtered = None
        self.this_All_Notes_ = None
        self.last_invalid_all_rows = None

        self.removed_No_assets = []
        self.new_No_assets = []
        self.removed_assets = []
        self.new_assets = []
        self.last_Notes_data = []
        self.this_Notes_data = []

        self.This_Notes_path = None
        self.Last_Notes_path = None

    def safe_read_excel(self, path):
        """安全讀取Excel並返回 DataFrame（不是 LazyFrame）"""
        required_columns = {"資產編號"}

        try:
            # 嘗試使用 Polars 讀取 Excel
            # 注意：设置 infer_schema_length=0 让Polars将所有列读取为字符串，避免类型推断错误
            if self.input_catalog is not None:
                df_polars = self.input_catalog.read_excel(path, infer_schema_length=0)
            else:
                df_polars = pl.read_excel(path, infer_schema_length=0)

            # 檢查是否包含必需的列
            header_columns = set(df_polars.columns)
            if required_columns.intersection(header_columns):
                # 自动去除资产编号列末尾的点号
                if "資產編號" in df_polars.columns:
                    df_polars = df_polars.with_columns(
                        [pl.col("資產編號").str.strip_chars_end(".").alias("資產編號")]
                    )
                # 移除全空列並返回 DataFrame
                non_null_cols = [
                    col
                    for col in df_polars.columns
                    if not df_polars[col].is_null().all()
                ]
                return df_polars.select(non_null_cols)

            # 如果表頭不正確，嘗試查找正確的表頭行
            for i in range(min(5, len(df_polars))):
                row_values = df_polars.row(i)
                clean_columns = {
                    str(col).strip() for col in row_values if col is not None
                }
                if required_columns.intersection(clean_columns):
                    new_columns = [
                        str(col).strip() if col is not None else f"col_{j}"
                        for j, col in enumerate(row_values)
                    ]
                    df_polars = df_polars.slice(i + 1).rename(
                        {
                            old: new
                            for old, new in zip(
                                df_polars.columns, new_columns, strict=False
                            )
                        }
                    )
                    # 自动去除资产编号列末尾的点号
                    if "資產編號" in df_polars.columns:
                        df_polars = df_polars.with_columns(
                            [
                                pl.col("資產編號")
                                .str.strip_chars_end(".")
                                .alias("資產編號")
                            ]
                        )
                    # 移除全空列並返回 DataFrame
                    non_null_cols = [
                        col
                        for col in df_polars.columns
                        if not df_polars[col].is_null().all()
                    ]
                    return df_polars.select(non_null_cols)

            raise ValueError(f"未在文件 {os.path.basename(path)} 中找到完整表頭")

        except Exception as e:
            logger.error(f"读取 {os.path.basename(path)} 失败: {e}")
            raise

    def This_Notes_date(self):
        """讀取本月Notes數據（使用LazyFrame）"""
        self.this_Notes_data = None
        if not self.This_Notes_path:
            return
        try:
            df_lazy = self.safe_read_excel(self.This_Notes_path)
            self.this_Notes_data = df_lazy
        except Exception as e:
            logger.error(e)

    def Last_Notes_date(self):
        """讀取上月Notes數據（使用LazyFrame）"""
        self.last_Notes_data = None
        if not self.Last_Notes_path:
            return
        try:
            df_lazy = self.safe_read_excel(self.Last_Notes_path)
            self.last_Notes_data = df_lazy
        except Exception as e:
            logger.error(e)

    def Notes_Notes_Comparison(self):
        """Notes與Notes數據比較（使用LazyFrame優化）"""
        if self.this_Notes_data is None or self.last_Notes_data is None:
            logger.error("文件讀取錯誤，請檢查文件")
            return
        else:
            try:
                # 使用 DataFrame 操作，獲取所有資產名稱
                this_all_notes_query = self.this_Notes_data.filter(
                    pl.col("資產名稱").is_not_null()
                ).select("資產名稱")
                self.this_All_Notes_ = this_all_notes_query.to_series().to_list()

                last_all_notes_query = self.last_Notes_data.filter(
                    pl.col("資產名稱").is_not_null()
                ).select("資產名稱")
                self.last_All_Notes_ = last_all_notes_query.to_series().to_list()

                # 過濾和匹配資產編號（使用DataFrame）- 只识别以18-或13-开头的为有效资产编号
                this_assets_query = self.this_Notes_data.filter(
                    pl.col("資產編號").is_not_null()
                    & (
                        pl.col("資產編號")
                        .cast(pl.String)
                        .str.strip_chars()
                        .str.len_chars()
                        > 0
                    )
                    & (
                        pl.col("資產編號").cast(pl.String).str.starts_with("18-")
                        | pl.col("資產編號").cast(pl.String).str.starts_with("13-")
                    )
                ).select("資產編號")
                self.this_assets_filtered = this_assets_query.to_series().to_list()

                last_assets_query = self.last_Notes_data.filter(
                    pl.col("資產編號").is_not_null()
                    & (
                        pl.col("資產編號")
                        .cast(pl.String)
                        .str.strip_chars()
                        .str.len_chars()
                        > 0
                    )
                    & (
                        pl.col("資產編號").cast(pl.String).str.starts_with("18-")
                        | pl.col("資產編號").cast(pl.String).str.starts_with("13-")
                    )
                ).select("資產編號")
                self.last_assets_filtered = last_assets_query.to_series().to_list()

                self.new_assets = set(self.this_assets_filtered) - set(
                    self.last_assets_filtered
                )
                self.removed_assets = set(self.last_assets_filtered) - set(
                    self.this_assets_filtered
                )

                # 處理無資產記錄（資產編號為空或不以18-或13-开头的记录）
                # 使用機身SN来匹配无资产记录的变化

                # 獲取本月無資產記錄（带机身SN）
                this_no_asset_df = self.this_Notes_data.filter(
                    pl.col("資產編號").is_null()
                    | (
                        pl.col("資產編號")
                        .cast(pl.String)
                        .str.strip_chars()
                        .str.len_chars()
                        == 0
                    )
                    | (
                        ~(
                            pl.col("資產編號").cast(pl.String).str.starts_with("18-")
                            | pl.col("資產編號").cast(pl.String).str.starts_with("13-")
                        )
                    )
                ).select(["資產名稱", "機身SN", "保管人"])

                # 獲取上月無資產記錄（带机身SN）
                last_no_asset_df = self.last_Notes_data.filter(
                    pl.col("資產編號").is_null()
                    | (
                        pl.col("資產編號")
                        .cast(pl.String)
                        .str.strip_chars()
                        .str.len_chars()
                        == 0
                    )
                    | (
                        ~(
                            pl.col("資產編號").cast(pl.String).str.starts_with("18-")
                            | pl.col("資產編號").cast(pl.String).str.starts_with("13-")
                        )
                    )
                ).select(["資產名稱", "機身SN", "保管人"])

                self.this_invalid_all_rows = this_no_asset_df.height
                self.last_invalid_all_rows = last_no_asset_df.height
                logger.info(
                    f"Notes_Notes本月无资产记录数量: {self.this_invalid_all_rows}"
                )
                logger.info(
                    f"Notes_Notes上月无资产记录数量: {self.last_invalid_all_rows}"
                )

                # 使用機身SN来识别新增和减少的无资产记录
                # 创建用于比对的唯一标识（资产名称 + 机身SN）
                this_no_asset_keys = (
                    this_no_asset_df.select(
                        pl.concat_str(
                            [pl.col("資產名稱"), pl.col("機身SN")], separator="||"
                        ).alias("key")
                    )
                    .to_series()
                    .to_list()
                )

                last_no_asset_keys = (
                    last_no_asset_df.select(
                        pl.concat_str(
                            [pl.col("資產名稱"), pl.col("機身SN")], separator="||"
                        ).alias("key")
                    )
                    .to_series()
                    .to_list()
                )

                # 找出新增和减少的无资产记录
                new_no_asset_keys_set = set(this_no_asset_keys) - set(
                    last_no_asset_keys
                )
                removed_no_asset_keys_set = set(last_no_asset_keys) - set(
                    this_no_asset_keys
                )

                # 保存新增和减少的无资产记录的key列表（用于后续保存时过滤）
                self.new_No_assets = list(new_no_asset_keys_set)
                self.removed_No_assets = list(removed_no_asset_keys_set)

                logger.info(f"Notes_Notes新增无资产记录: {len(self.new_No_assets)} 笔")
                logger.info(
                    f"Notes_Notes减少无资产记录: {len(self.removed_No_assets)} 笔"
                )

            except Exception as e:
                logger.error(e)

    def Save_Notes_Notes_Comparison(self):
        """保存Notes與Notes比較結果為Excel文件"""
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if (
            not self.new_assets
            and not self.removed_assets
            and not self.new_No_assets
            and not self.removed_No_assets
        ):
            return
        else:
            try:
                # 創建Excel寫入器
                wb, worksheet = new_workbook("對比結果")
                # 使用 DataFrame 構建查詢
                new_df = (
                    self.this_Notes_data.filter(
                        pl.col("資產編號").is_in(list(self.new_assets))
                    )
                    .select(["資產名稱", "資產編號", "保管人"])
                    .unique()
                )

                removed_df = (
                    self.last_Notes_data.filter(
                        pl.col("資產編號").is_in(list(self.removed_assets))
                    )
                    .select(["資產名稱", "資產編號", "保管人"])
                    .unique()
                )

                # 寫入標題信息
                worksheet.merge_cells("A1:D1")
                worksheet["A1"] = f"本月Notes_VS_上月Notes (对比时间{current_time})"
                worksheet["A1"].font = openpyxl.styles.Font(bold=True, size=14)
                worksheet["A1"].alignment = openpyxl.styles.Alignment(
                    horizontal="center"
                )
                # 合并并居中 A2:D2 和 A3:D3
                worksheet.merge_cells("A2:D2")
                worksheet.merge_cells("A3:D3")
                worksheet["A2"].alignment = openpyxl.styles.Alignment(
                    horizontal="center"
                )
                worksheet["A3"].alignment = openpyxl.styles.Alignment(
                    horizontal="center"
                )

                current_row = 3

                # 寫入新增資產數據
                if not new_df.is_empty():
                    # 標題
                    worksheet[f"A{current_row}"] = (
                        f"本月比上月新增资产 {len(self.new_assets)}笔"
                    )
                    worksheet[f"A{current_row}"].font = openpyxl.styles.Font(
                        bold=True, size=12
                    )
                    current_row += 1

                    # 列標題
                    headers = ["No.", "資產名稱", "資產編號", "保管人"]
                    for i, header in enumerate(headers, 1):
                        cell = worksheet.cell(row=current_row, column=i, value=header)
                        cell.font = openpyxl.styles.Font(bold=True)
                        cell.fill = openpyxl.styles.PatternFill(
                            start_color="E6F3FF",
                            end_color="E6F3FF",
                            fill_type="solid",
                        )
                    current_row += 1

                    # 數據
                    for i, row_dict in enumerate(new_df.iter_rows(named=True), 1):
                        worksheet.cell(row=current_row, column=1, value=i)
                        safe_cell(worksheet, current_row, 2, row_dict["資產名稱"])
                        safe_cell(worksheet, current_row, 3, row_dict["資產編號"])
                        safe_cell(worksheet, current_row, 4, row_dict["保管人"])
                        current_row += 1
                    current_row += 1

                # 寫入減少資產數據
                if not removed_df.is_empty():
                    worksheet[f"A{current_row}"] = (
                        f"本月比上月减少资产 {len(self.removed_assets)}笔"
                    )
                    worksheet[f"A{current_row}"].font = openpyxl.styles.Font(
                        bold=True, size=12
                    )
                    current_row += 1

                    headers = ["No.", "資產名稱", "資產編號", "保管人"]
                    for i, header in enumerate(headers, 1):
                        cell = worksheet.cell(row=current_row, column=i, value=header)
                        cell.font = openpyxl.styles.Font(bold=True)
                        cell.fill = openpyxl.styles.PatternFill(
                            start_color="FFE6E6",
                            end_color="FFE6E6",
                            fill_type="solid",
                        )
                    current_row += 1

                    for i, row_dict in enumerate(removed_df.iter_rows(named=True), 1):
                        worksheet.cell(row=current_row, column=1, value=i)
                        safe_cell(worksheet, current_row, 2, row_dict["資產名稱"])
                        safe_cell(worksheet, current_row, 3, row_dict["資產編號"])
                        safe_cell(worksheet, current_row, 4, row_dict["保管人"])
                        current_row += 1
                    current_row += 1

                # 處理無資產記錄的變化
                if len(self.new_No_assets) > 0:
                    # 根据key列表过滤出新增的无资产记录
                    new_No_df = (
                        self.this_Notes_data.filter(
                            pl.col("資產編號").is_null()
                            | (
                                pl.col("資產編號")
                                .cast(pl.String)
                                .str.strip_chars()
                                .str.len_chars()
                                == 0
                            )
                            | (
                                ~(
                                    pl.col("資產編號")
                                    .cast(pl.String)
                                    .str.starts_with("18-")
                                    | pl.col("資產編號")
                                    .cast(pl.String)
                                    .str.starts_with("13-")
                                )
                            )
                        )
                        .select(["資產名稱", "機身SN", "保管人"])
                        .with_columns(
                            pl.concat_str(
                                [pl.col("資產名稱"), pl.col("機身SN")],
                                separator="||",
                            ).alias("key")
                        )
                        .filter(pl.col("key").is_in(self.new_No_assets))
                        .drop("key")
                    )

                    if not new_No_df.is_empty():
                        worksheet[f"A{current_row}"] = (
                            f"本月比上月新增无资产记录 {len(self.new_No_assets)}笔"
                        )
                        worksheet[f"A{current_row}"].font = openpyxl.styles.Font(
                            bold=True, size=12
                        )
                        current_row += 1

                        headers = ["No.", "資產名稱", "機身SN", "保管人"]
                        for i, header in enumerate(headers, 1):
                            cell = worksheet.cell(
                                row=current_row, column=i, value=header
                            )
                            cell.font = openpyxl.styles.Font(bold=True)
                            cell.fill = openpyxl.styles.PatternFill(
                                start_color="E6F3FF",
                                end_color="E6F3FF",
                                fill_type="solid",
                            )
                        current_row += 1

                        for i, row_dict in enumerate(
                            new_No_df.iter_rows(named=True), 1
                        ):
                            worksheet.cell(row=current_row, column=1, value=i)
                            safe_cell(worksheet, current_row, 2, row_dict["資產名稱"])
                            safe_cell(worksheet, current_row, 3, row_dict["機身SN"])
                            safe_cell(worksheet, current_row, 4, row_dict["保管人"])
                            current_row += 1
                        current_row += 1

                if len(self.removed_No_assets) > 0:
                    # 根据key列表过滤出减少的无资产记录
                    removed_No_df = (
                        self.last_Notes_data.filter(
                            pl.col("資產編號").is_null()
                            | (
                                pl.col("資產編號")
                                .cast(pl.String)
                                .str.strip_chars()
                                .str.len_chars()
                                == 0
                            )
                            | (
                                ~(
                                    pl.col("資產編號")
                                    .cast(pl.String)
                                    .str.starts_with("18-")
                                    | pl.col("資產編號")
                                    .cast(pl.String)
                                    .str.starts_with("13-")
                                )
                            )
                        )
                        .select(["資產名稱", "機身SN", "保管人"])
                        .with_columns(
                            pl.concat_str(
                                [pl.col("資產名稱"), pl.col("機身SN")],
                                separator="||",
                            ).alias("key")
                        )
                        .filter(pl.col("key").is_in(self.removed_No_assets))
                        .drop("key")
                    )

                    if not removed_No_df.is_empty():
                        worksheet[f"A{current_row}"] = (
                            f"本月比上月减少无资产记录 {len(self.removed_No_assets)}笔"
                        )
                        worksheet[f"A{current_row}"].font = openpyxl.styles.Font(
                            bold=True, size=12
                        )
                        current_row += 1

                        headers = ["No.", "資產名稱", "機身SN", "保管人"]
                        for i, header in enumerate(headers, 1):
                            cell = worksheet.cell(
                                row=current_row, column=i, value=header
                            )
                            cell.font = openpyxl.styles.Font(bold=True)
                            cell.fill = openpyxl.styles.PatternFill(
                                start_color="FFE6E6",
                                end_color="FFE6E6",
                                fill_type="solid",
                            )
                        current_row += 1

                        for i, row_dict in enumerate(
                            removed_No_df.iter_rows(named=True), 1
                        ):
                            worksheet.cell(row=current_row, column=1, value=i)
                            safe_cell(worksheet, current_row, 2, row_dict["資產名稱"])
                            safe_cell(worksheet, current_row, 3, row_dict["機身SN"])
                            safe_cell(worksheet, current_row, 4, row_dict["保管人"])
                            current_row += 1
                        current_row += 1

                # 設置列寬和邊框
                from openpyxl.styles import Border, Side  # noqa: E402, I001, UP015, F401

                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

                # 設置列寬
                column_widths = [8, 40, 25, 20]  # 根據內容調整
                for i, width in enumerate(column_widths, 1):
                    worksheet.column_dimensions[
                        openpyxl.utils.get_column_letter(i)
                    ].width = width

                # 添加邊框
                for row in worksheet.iter_rows(
                    min_row=1,
                    max_row=worksheet.max_row,
                    min_col=1,
                    max_col=worksheet.max_column,
                ):
                    for cell in row:
                        cell.border = thin_border

                wb.save(NOTES_NOTES_SAVE_PATH)
            except Exception as e:
                logger.error(e)

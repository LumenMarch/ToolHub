import os  # noqa: E402, I001, UP015, F401
from datetime import datetime  # noqa: E402, I001, UP015, F401

import openpyxl  # noqa: E402, I001, UP015, F401
import polars as pl  # noqa: E402, I001, UP015, F401
import xlrd  # noqa: E402, I001, UP015, F401
from app.services.asset_comparison.excel_writer import (  # noqa: E402, I001, UP015, F401
    new_workbook,
    write_section,
)
from app.services.asset_comparison.TableParser import (  # noqa: E402
    TableParser,
)  # 使用自定義的 HTML 表格解析器  # noqa: E402, I001, UP015, F401
from loguru import logger  # noqa: E402, I001, UP015, F401


class SFC_SFC:
    def __init__(self):
        self.input_catalog = None
        self.removed_count = None
        self.new_count = None
        self.removed_assets = None
        self.new_assets = None
        self.this_SFC_assets = None
        self.last_SFC_assets = None
        self.this_SFC_data = None
        self.last_SFC_data = None

        self.This_data_Path = None
        self.Last_data_path = None

    def safe_read_excel(self, path):
        """安全讀取Excel或HTML偽裝的Excel，並返回 Polars DataFrame"""
        required_columns = {"设备编号", "资产编号", "資產編號"}

        # 智能檢測文件格式並相應處理
        def is_html_file(file_path):
            """檢測文件是否為HTML格式"""
            try:
                with open(file_path, encoding="utf-8-sig", errors="ignore") as file:
                    for line in file:
                        line = line.strip()
                        if not line:
                            continue  # skip blank lines (including bare BOM)
                        line = line.lower()
                        return line.startswith("<") or "html" in line or "table" in line
                return False
            except Exception:
                return False

        # 首先檢測文件格式
        if is_html_file(path):
            logger.info(f"檢測到HTML格式文件，使用HTML解析: {os.path.basename(path)}")
            # 使用 HTML 解析方式
            try:
                if self.input_catalog is not None:
                    html_content = self.input_catalog.read_text(
                        path, encoding="utf-8-sig", errors="ignore"
                    )
                else:
                    with open(path, encoding="utf-8-sig", errors="ignore") as f:
                        html_content = f.read()

                # 使用自定義的 TableParser 解析 HTML
                parser = TableParser()
                parser.feed(html_content)

                if not parser.tables:
                    raise ValueError(
                        f"未在 HTML 文件中找到表格: {os.path.basename(path)}"
                    )

                # 使用第一個表格
                table_data = parser.tables[0]
                if not table_data:
                    raise ValueError(f"表格為空: {os.path.basename(path)}")

                # 查找表頭行
                header_row = None
                for idx, row in enumerate(table_data):
                    # 處理空值：將None轉換為空字符串，然後清理
                    clean_columns = {
                        str(col).strip() for col in row if col and str(col).strip()
                    }
                    if required_columns.intersection(clean_columns):
                        header_row = idx
                        break

                if header_row is not None:
                    # 使用找到的行作為表頭
                    headers = table_data[header_row]
                    data_rows = table_data[header_row + 1 :]
                else:
                    # 默認使用第一行為表頭
                    logger.warning(
                        f"HTML中未找到完整表頭，默認使用第一行: {os.path.basename(path)}"
                    )
                    headers = table_data[0] if table_data else []
                    data_rows = table_data[1:] if len(table_data) > 1 else []

                # 確保所有行都有相同的列數
                max_cols = len(headers) if headers else 0
                if max_cols == 0:
                    raise ValueError(f"無法確定表格結構: {os.path.basename(path)}")

                # 補齊缺失的列並處理空值
                normalized_data = []
                for row in data_rows:
                    # 處理每一行的空值：將None轉換為空字符串
                    cleaned_row = [
                        str(cell).strip() if cell is not None else "" for cell in row
                    ]
                    cleaned_row + [""] * (max_cols - len(cleaned_row))
                    normalized_data.append(cleaned_row[:max_cols])

                # 創建 Polars DataFrame
                if normalized_data:
                    df_dict = {}
                    for i, header in enumerate(headers):
                        # 處理表頭的空值
                        column_name = (
                            str(header).strip()
                            if header is not None and str(header).strip()
                            else f"col_{i}"
                        )
                        # 處理數據列的空值
                        column_data = []
                        for row in normalized_data:
                            if i < len(row):
                                cell_value = row[i]
                                # 確保空值統一處理為空字符串
                                column_data.append(
                                    cell_value
                                    if cell_value and str(cell_value).strip()
                                    else ""
                                )
                            else:
                                column_data.append("")
                        df_dict[column_name] = column_data

                    df = pl.DataFrame(df_dict)
                else:
                    # 如果沒有數據行，創建空的 DataFrame
                    df_dict = {}
                    for i, header in enumerate(headers):
                        column_name = (
                            str(header).strip()
                            if header is not None and str(header).strip()
                            else f"col_{i}"
                        )
                        df_dict[column_name] = []
                    df = pl.DataFrame(df_dict)

                logger.info(f"HTML解析成功，原始数据行数: {len(df)}")

            except Exception as html_error:
                logger.exception(f"HTML解析失敗: {html_error}")
                return None

        else:
            logger.info(f"檢測到Excel格式文件，嘗試讀取: {os.path.basename(path)}")
            # 策略：.xls 文件使用 xlrd，.xlsx 用 calamine
            try:
                if os.path.splitext(path)[1].lower() == ".xls":
                    logger.info(f".xls 文件使用 xlrd 讀取: {os.path.basename(path)}")
                    wb = xlrd.open_workbook(path)
                    sheet = wb.sheet_by_index(0)
                    headers = [sheet.cell_value(0, c) for c in range(sheet.ncols)]
                    data_rows = []
                    for r in range(1, sheet.nrows):
                        row = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
                        data_rows.append(row)
                    df = pl.DataFrame(data_rows, schema=headers, orient="row")
                else:
                    if self.input_catalog is not None:
                        df = self.input_catalog.read_excel(
                            path, engine="calamine", infer_schema_length=0
                        )
                    else:
                        df = pl.read_excel(
                            path, engine="calamine", infer_schema_length=0
                        )
                logger.info(f"Excel文件讀取成功，原始数据行数: {len(df)}")

            except Exception as excel_error:
                logger.exception(f"Excel讀取失敗: {excel_error}，嘗試 HTML 回退解析")
                # is_html_file 可能因 BOM / 空行漏检，回退到 HTML 解析路徑
                try:
                    if self.input_catalog is not None:
                        html_content = self.input_catalog.read_text(
                            path, encoding="utf-8-sig", errors="ignore"
                        )
                    else:
                        with open(path, encoding="utf-8-sig", errors="ignore") as file:
                            html_content = file.read()

                    parser = TableParser()
                    parser.feed(html_content)

                    if not parser.tables:
                        raise ValueError(
                            f"未在 HTML 文件中找到表格: {os.path.basename(path)}"
                        )

                    table_data = parser.tables[0]
                    if not table_data:
                        raise ValueError(f"表格為空: {os.path.basename(path)}")

                    # 查找表頭行
                    header_row = None
                    for idx, row in enumerate(table_data):
                        clean_columns = {
                            str(col).strip() for col in row if col and str(col).strip()
                        }
                        if required_columns.intersection(clean_columns):
                            header_row = idx
                            break

                    if header_row is not None:
                        headers = table_data[header_row]
                        data_rows = table_data[header_row + 1 :]
                    else:
                        logger.warning(
                            f"HTML中未找到完整表頭，默認使用第一行: {os.path.basename(path)}"
                        )
                        headers = table_data[0] if table_data else []
                        data_rows = table_data[1:] if len(table_data) > 1 else []

                    max_cols = len(headers) if headers else 0
                    if max_cols == 0:
                        raise ValueError(f"無法確定表格結構: {os.path.basename(path)}")

                    normalized_data = []
                    for row in data_rows:
                        cleaned_row = [
                            str(cell).strip() if cell is not None else ""
                            for cell in row
                        ]
                        cleaned_row + [""] * (max_cols - len(cleaned_row))
                        normalized_data.append(cleaned_row[:max_cols])

                    if normalized_data:
                        df_dict = {}
                        for i, header in enumerate(headers):
                            column_name = (
                                str(header).strip()
                                if header is not None and str(header).strip()
                                else f"col_{i}"
                            )
                            column_data = []
                            for row in normalized_data:
                                if i < len(row):
                                    cell_value = row[i]
                                    column_data.append(
                                        cell_value
                                        if cell_value and str(cell_value).strip()
                                        else ""
                                    )
                                else:
                                    column_data.append("")
                            df_dict[column_name] = column_data
                        df = pl.DataFrame(df_dict)
                    else:
                        df_dict = {}
                        for i, header in enumerate(headers):
                            column_name = (
                                str(header).strip()
                                if header is not None and str(header).strip()
                                else f"col_{i}"
                            )
                            df_dict[column_name] = []
                        df = pl.DataFrame(df_dict)

                    logger.info(f"HTML回退解析成功，原始数据行数: {len(df)}")
                except Exception as html_error:
                    logger.exception(f"HTML回退解析也失敗: {html_error}")
                    return None

        # 統一的数据清理和过滤逻辑
        try:
            # 檢查是否包含必需的列
            df_columns = set(df.columns)
            if not required_columns.intersection(df_columns):
                logger.warning(
                    f"文件中未找到必需的列，嘗試查找包含關鍵字的列: {os.path.basename(path)}"
                )
                # 嘗試查找包含關鍵字的列
                found_columns = []
                for col in df.columns:
                    col_str = str(col).lower()
                    if any(
                        keyword in col_str
                        for keyword in ["设备编号", "资产编号", "資產編號", "設備編號"]
                    ):
                        found_columns.append(col)

                if not found_columns:
                    logger.warning(
                        f"未找到資產編號相關列，但繼續處理: {os.path.basename(path)}"
                    )
                else:
                    logger.info(f"找到相關列: {found_columns}")

            # 将所有列转换为字符串类型以避免数据类型推断错误
            df = df.with_columns([pl.col(c).cast(pl.Utf8) for c in df.columns])

            # 🎯 加强空数据过滤
            # 自动去除资产编号列末尾的点号
            for col in ["资产编号", "資產編號"]:
                if col in df.columns:
                    df = df.with_columns(
                        pl.col(col).cast(pl.Utf8).str.strip_chars_end(".").alias(col)
                    )

            # 🎯 过滤完全空白的行
            df = df.filter(
                ~pl.all_horizontal(
                    [
                        pl.col(c)
                        .cast(pl.Utf8)
                        .fill_null("")
                        .str.strip_chars()
                        .is_in(["", "nan", "none", "null"])
                        for c in df.columns
                    ]
                )
            )

            # 🎯 过滤资产编号为空的行
            asset_cols = ["资产编号", "資產編號"]
            for col in asset_cols:
                if col in df.columns:
                    df = df.filter(
                        ~pl.col(col)
                        .cast(pl.Utf8)
                        .fill_null("")
                        .str.strip_chars()
                        .str.to_lowercase()
                        .is_in(["", "nan", "none", "null"])
                    )

            logger.info(f"过滤后数据行数: {len(df)}")
            if len(df) > 0:
                logger.info(f"数据样本: {df.head(3).to_dicts()}")

            return df

        except Exception as e:
            logger.exception(f"数据清理和过滤失败: {e}")
            return None

    def This_SFC_date(self):
        """讀取本期SFC數據"""
        self.this_SFC_data = None
        try:
            self.this_SFC_data = self.safe_read_excel(self.This_data_Path)
        except Exception as e:
            logger.exception(e)

    def Last_SFC_date(self):
        """讀取上期SFC數據"""
        self.last_SFC_data = None
        try:
            self.last_SFC_data = self.safe_read_excel(self.Last_data_path)
        except Exception as e:
            logger.error(e)

    def _get_comparison_keys(self, df):
        """获取用于对比的唯一标识键。
        资产编号有效时用资产编号，为 NA 时降级使用设备编号。
        返回 (keys_set, asset_col, device_col)"""
        if df is None or df.is_empty():
            return set(), None, None

        asset_col = next((c for c in ["资产编号", "資產編號"] if c in df.columns), None)
        device_col = next(
            (c for c in ["设备编号", "設備編號"] if c in df.columns), None
        )

        invalid_values = {"nan", "none", "null", "na", ""}
        keys = set()
        for row in df.iter_rows(named=True):
            asset_val = str(row.get(asset_col) or "").strip() if asset_col else ""
            device_val = str(row.get(device_col) or "").strip() if device_col else ""

            if asset_val and asset_val.lower() not in invalid_values:
                keys.add(f"A:{asset_val}")
            elif device_val and device_val.lower() not in invalid_values:
                keys.add(f"D:{device_val}")

        return keys, asset_col, device_col

    def SFC_SFC_Comparison(self):
        """SFC數據比較"""
        self.new_assets = None
        self.removed_assets = None
        self.this_SFC_assets = None
        self.last_SFC_assets = None
        if self.last_SFC_data is None or self.this_SFC_data is None:
            return
        else:
            try:
                # 获取对比标识键（资产编号降级设备编号）
                this_keys, self._asset_col, self._device_col = (
                    self._get_comparison_keys(self.this_SFC_data)
                )
                last_keys, _, _ = self._get_comparison_keys(self.last_SFC_data)

                self.this_SFC_assets = list(this_keys)
                self.last_SFC_assets = list(last_keys)
                logger.info(f"本月有效资产数量: {len(this_keys)}")
                logger.info(f"上月有效资产数量: {len(last_keys)}")
                if this_keys:
                    logger.info(f"本月资产样本: {list(this_keys)[:3]}")
                if last_keys:
                    logger.info(f"上月资产样本: {list(last_keys)[:3]}")

                self.new_assets = this_keys - last_keys
                self.removed_assets = last_keys - this_keys

                new_count = len(self.new_assets)
                removed_count = len(self.removed_assets)

                logger.info(f"新增资产数量: {new_count}")
                logger.info(f"减少资产数量: {removed_count}")

            except Exception as e:
                logger.exception(e)

    def Save_SFC_SFC_Comparison(self, output_path):
        """保存SFC比較結果為Excel文件"""
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            if not self.new_assets and not self.removed_assets:
                return
            else:
                # 检测列名
                asset_col = getattr(self, "_asset_col", None)
                device_col = getattr(self, "_device_col", None)
                if not asset_col:
                    for c in ["资产编号", "資產編號"]:
                        if c in self.this_SFC_data.columns:
                            asset_col = c
                            break
                if not device_col:
                    for c in ["设备编号", "設備編號"]:
                        if c in self.this_SFC_data.columns:
                            device_col = c
                            break

                # 分离 A:/D: 前缀标识
                def split_keys(keys):
                    by_asset = {k[2:] for k in keys if k.startswith("A:")}
                    by_device = {k[2:] for k in keys if k.startswith("D:")}
                    return by_asset, by_device

                new_by_asset, new_by_device = split_keys(self.new_assets)
                removed_by_asset, removed_by_device = split_keys(self.removed_assets)

                invalid_values = ["nan", "none", "null", "na", ""]

                def filter_by_keys(df, asset_keys, device_keys):
                    """用资产编号或设备编号过滤 DataFrame"""
                    conds = []
                    if asset_col and asset_col in df.columns and asset_keys:
                        am = (
                            pl.col(asset_col)
                            .cast(pl.Utf8)
                            .str.strip_chars()
                            .is_in(list(asset_keys))
                        )
                        valid = ~pl.col(asset_col).cast(
                            pl.Utf8
                        ).str.strip_chars().str.to_lowercase().is_in(invalid_values)
                        conds.append(am & valid)
                    if device_col and device_col in df.columns and device_keys:
                        dm = (
                            pl.col(device_col)
                            .cast(pl.Utf8)
                            .str.strip_chars()
                            .is_in(list(device_keys))
                        )
                        conds.append(dm)
                    if not conds:
                        return df.clear()
                    return df.filter(pl.any_horizontal(conds))

                # 选列：name, asset, keeper + device(用于NA替补)
                name_col = None
                for c in ["设备名称", "設備名称", "資產名稱", "资产名称"]:
                    if c in self.this_SFC_data.columns:
                        name_col = c
                        break
                if not name_col:
                    name_col = self.this_SFC_data.columns[0]

                keeper_col = "保管人"
                if keeper_col not in self.this_SFC_data.columns:
                    for c in self.this_SFC_data.columns:
                        if "保管" in c or "管理" in c or "负责" in c:
                            keeper_col = c
                            break
                    else:
                        keeper_col = self.this_SFC_data.columns[-1]

                select_cols = [name_col, asset_col, keeper_col]
                if (
                    device_col
                    and device_col not in select_cols
                    and device_col in self.this_SFC_data.columns
                ):
                    select_cols.append(device_col)

                new_df = filter_by_keys(self.this_SFC_data, new_by_asset, new_by_device)
                if not new_df.is_empty():
                    new_df = new_df.select(select_cols).unique()

                removed_df = filter_by_keys(
                    self.last_SFC_data, removed_by_asset, removed_by_device
                )
                if not removed_df.is_empty():
                    removed_df = removed_df.select(select_cols).unique()

                # 资产编号为 NA 时用设备编号替代显示
                def fill_na_asset(df):
                    if (
                        device_col
                        and asset_col
                        and asset_col in df.columns
                        and device_col in df.columns
                    ):
                        df = df.with_columns(
                            pl.when(
                                pl.col(asset_col)
                                .cast(pl.Utf8)
                                .str.strip_chars()
                                .str.to_lowercase()
                                .is_in(invalid_values)
                            )
                            .then(pl.col(device_col))
                            .otherwise(pl.col(asset_col))
                            .alias(asset_col)
                        )
                    return df

                new_df = fill_na_asset(new_df)
                removed_df = fill_na_asset(removed_df)

                # 统一输出列名
                out_cols_map = {
                    name_col: "设备名称",
                    asset_col: "资产编号",
                    keeper_col: "保管人",
                }
                out_cols_map = {k: v for k, v in out_cols_map.items() if k != v}

                if not new_df.is_empty():
                    new_df = new_df.rename(out_cols_map).select(
                        ["设备名称", "资产编号", "保管人"]
                    )
                else:
                    new_df = pl.DataFrame(schema=["设备名称", "资产编号", "保管人"])

                if not removed_df.is_empty():
                    removed_df = removed_df.rename(out_cols_map).select(
                        ["设备名称", "资产编号", "保管人"]
                    )
                else:
                    removed_df = pl.DataFrame(schema=["设备名称", "资产编号", "保管人"])

                # 創建Excel寫入器
                wb, ws = new_workbook("對比結果")

                # 寫入標題信息
                ws.merge_cells("A1:D1")
                ws["A1"] = f"本月SFC资产_VS_上月SFC资产 (对比时间{current_time})"
                ws["A1"].font = openpyxl.styles.Font(bold=True, size=14)
                ws["A1"].alignment = openpyxl.styles.Alignment(horizontal="center")
                # 合并并居中 A2:D2 和 A3:D3
                ws.merge_cells("A2:D2")
                ws.merge_cells("A3:D3")
                ws["A2"].alignment = openpyxl.styles.Alignment(horizontal="center")
                ws["A3"].alignment = openpyxl.styles.Alignment(horizontal="center")

                current_row = 3

                # 寫入新增資產數據
                if not new_df.is_empty():
                    current_row = write_section(
                        ws,
                        new_df,
                        f"本月比上月新增资产 {len(self.new_assets)}笔",
                        current_row,
                        "new",
                    )

                # 寫入減少資產數據
                if not removed_df.is_empty():
                    current_row = write_section(
                        ws,
                        removed_df,
                        f"本月比上月减少资产 {len(self.removed_assets)}笔",
                        current_row,
                        "removed",
                    )

                # 設置列寬
                from openpyxl.utils import get_column_letter

                column_widths = [8, 40, 25, 20]  # 根據內容調整
                for i, width in enumerate(column_widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = width

                wb.save(output_path)

        except Exception as e:
            logger.exception(e)

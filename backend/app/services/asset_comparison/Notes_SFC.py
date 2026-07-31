import os  # noqa: E402, I001, UP015, F401
from datetime import datetime  # noqa: E402, I001, UP015, F401

import openpyxl  # noqa: E402, I001, UP015, F401
import polars as pl  # noqa: E402, I001, UP015, F401
import xlrd  # noqa: E402, I001, UP015, F401
from app.services.asset_comparison.excel_writer import (  # noqa: E402, I001, UP015, F401
    new_workbook,
    write_section,
)
from app.services.asset_comparison.TableParser import (  # noqa: E402
    TableParser,
)  # 使用自定義的 HTML 表格解析器  # noqa: E402, I001, UP015, F401
from loguru import logger  # noqa: E402, I001, UP015, F401


class Notes_SFC:
    def __init__(self):
        self.input_catalog = None
        self.Notes_new_assets = None
        self.Notes_removed_assets = None
        self.this_Notes_assets = None
        self.this_SFC_assets = None

        self.this_Notes_path = None
        self.this_SFC_path = None

        self.removed_assets = None
        self.new_assets = None
        self.this_Notes_data = None
        self.this_SFC_data = None

        # 列名变量
        self.notes_asset_col = "資產編號"  # 默认值
        self.sfc_asset_col = "资产编号"  # 默认值
        self.notes_device_col = None
        self.sfc_device_col = None

    def safe_read_excel(self, path):
        """安全讀取Excel或HTML偽裝的Excel，並轉為 Polars DataFrame"""
        required_columns = {"设备编号", "资产编号", "資產編號"}

        try:
            # 嘗試使用 Polars 讀取 Excel
            # 注意：设置 infer_schema_length=0 让Polars将所有列读取为字符串，避免类型推断错误
            if self.input_catalog is not None:
                df_polars = self.input_catalog.read_excel(path, infer_schema_length=0)
            else:
                df_polars = pl.read_excel(path, infer_schema_length=0)

            # 檢查是否包含必需的列
            header_columns = set(df_polars.columns)
            if required_columns.intersection(header_columns):
                # 自动去除资产编号列末尾的点号
                if "资产编号" in df_polars.columns:
                    df_polars = df_polars.with_columns(
                        [pl.col("资产编号").str.strip_chars_end(".").alias("资产编号")]
                    )
                if "資產編號" in df_polars.columns:
                    df_polars = df_polars.with_columns(
                        [pl.col("資產編號").str.strip_chars_end(".").alias("資產編號")]
                    )
                # 移除全空列並返回 DataFrame
                non_null_cols = [
                    col
                    for col in df_polars.columns
                    if not df_polars[col].is_null().all()
                ]
                return df_polars.select(non_null_cols)

            # 如果表頭不正確，嘗試查找正確的表頭行
            for i in range(min(5, len(df_polars))):
                row_values = df_polars.row(i)
                clean_columns = {
                    str(col).strip() for col in row_values if col is not None
                }
                if required_columns.intersection(clean_columns):
                    new_columns = [
                        str(col).strip() if col is not None else f"col_{j}"
                        for j, col in enumerate(row_values)
                    ]
                    df_polars = df_polars.slice(i + 1).rename(
                        {
                            old: new
                            for old, new in zip(
                                df_polars.columns, new_columns, strict=False
                            )
                        }
                    )
                    # 自动去除资产编号列末尾的点号
                    if "资产编号" in df_polars.columns:
                        df_polars = df_polars.with_columns(
                            [
                                pl.col("资产编号")
                                .str.strip_chars_end(".")
                                .alias("资产编号")
                            ]
                        )
                    if "資產編號" in df_polars.columns:
                        df_polars = df_polars.with_columns(
                            [
                                pl.col("資產編號")
                                .str.strip_chars_end(".")
                                .alias("資產編號")
                            ]
                        )
                    # 移除全空列並返回 DataFrame
                    non_null_cols = [
                        col
                        for col in df_polars.columns
                        if not df_polars[col].is_null().all()
                    ]
                    return df_polars.select(non_null_cols)

            # 如果都不行，拋出錯誤
            raise ValueError(f"未在文件 {os.path.basename(path)} 中找到完整表頭")

        except Exception as e:
            # 嘗試用 xlrd 讀取 .xls 文件（calamine 對舊格式 .xls 支援不佳）
            if os.path.splitext(path)[1].lower() == ".xls":
                logger.info(f".xls 文件嘗試使用 xlrd 讀取: {os.path.basename(path)}")
                try:
                    wb = xlrd.open_workbook(path)
                    sheet = wb.sheet_by_index(0)
                    headers = [sheet.cell_value(0, c) for c in range(sheet.ncols)]
                    data_rows = []
                    for r in range(1, sheet.nrows):
                        row = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
                        data_rows.append(row)
                    df_polars = pl.DataFrame(data_rows, schema=headers, orient="row")
                    # 自动去除资产编号列末尾的点号
                    if "资产编号" in df_polars.columns:
                        df_polars = df_polars.with_columns(
                            [
                                pl.col("资产编号")
                                .str.strip_chars_end(".")
                                .alias("资产编号")
                            ]
                        )
                    if "資產編號" in df_polars.columns:
                        df_polars = df_polars.with_columns(
                            [
                                pl.col("資產編號")
                                .str.strip_chars_end(".")
                                .alias("資產編號")
                            ]
                        )
                    # 移除全空列
                    non_null_cols = [
                        col
                        for col in df_polars.columns
                        if not df_polars[col].is_null().all()
                    ]
                    logger.info(f"xlrd 讀取 .xls 成功，行數: {len(df_polars)}")
                    return df_polars.select(non_null_cols)
                except Exception as xlrd_error:
                    logger.error(f"xlrd 讀取 .xls 也失敗: {xlrd_error}")

            logger.error(
                f"Polars Excel讀取失敗: {e}，嘗試使用HTML方式讀取: {os.path.basename(path)}"
            )

            # 嘗試按 HTML 表格方式讀取（應對偽裝的 .xls 文件）
            try:
                if self.input_catalog is not None:
                    html_content = self.input_catalog.read_text(
                        path, encoding="utf-8-sig", errors="ignore"
                    )
                else:
                    with open(path, encoding="utf-8-sig", errors="ignore") as f:
                        html_content = f.read()

                # 使用自定義的 TableParser 解析 HTML
                parser = TableParser()
                parser.feed(html_content)

                if not parser.tables:
                    raise ValueError(
                        f"未在 HTML 文件中找到表格: {os.path.basename(path)}"
                    )

                # 使用第一個表格
                table_data = parser.tables[0]
                if not table_data:
                    raise ValueError(f"表格為空: {os.path.basename(path)}")

                # 查找表頭行
                header_row = None
                for idx, row in enumerate(table_data):
                    clean_columns = {str(col).strip() for col in row if col}
                    if required_columns.intersection(clean_columns):
                        header_row = idx
                        break

                if header_row is not None:
                    # 使用找到的行作為表頭
                    headers = table_data[header_row]
                    data_rows = table_data[header_row + 1 :]
                else:
                    # 默認使用第一行為表頭
                    logger.warning(
                        f"HTML中未找到完整表頭，默認使用第一行: {os.path.basename(path)}"
                    )
                    headers = table_data[0] if table_data else []
                    data_rows = table_data[1:] if len(table_data) > 1 else []

                # 確保所有行都有相同的列數
                max_cols = len(headers) if headers else 0
                if max_cols == 0:
                    raise ValueError(f"無法確定表格結構: {os.path.basename(path)}")

                # 補齊缺失的列
                normalized_data = []
                for row in data_rows:
                    normalized_row = list(row) + [""] * (max_cols - len(row))
                    normalized_data.append(normalized_row[:max_cols])

                # 創建 Polars DataFrame，显式指定所有列为字符串类型避免类型推断错误
                if normalized_data:
                    df_dict = {}
                    for i, header in enumerate(headers):
                        column_name = str(header).strip() if header else f"col_{i}"
                        # 确保所有值都转为字符串，避免类型推断问题
                        df_dict[column_name] = [
                            str(row[i]) if i < len(row) and row[i] is not None else ""
                            for row in normalized_data
                        ]

                    # 创建 schema，所有列都指定为字符串类型
                    schema = {col: pl.String for col in df_dict}
                    df_polars = pl.DataFrame(df_dict, schema=schema)
                else:
                    # 如果沒有數據行，創建空的 DataFrame
                    df_dict = {
                        str(header).strip() if header else f"col_{i}": []
                        for i, header in enumerate(headers)
                    }
                    schema = {col: pl.String for col in df_dict.keys()}
                    df_polars = pl.DataFrame(df_dict, schema=schema)

                # 自动去除资产编号列末尾的点号
                if "资产编号" in df_polars.columns:
                    df_polars = df_polars.with_columns(
                        [pl.col("资产编号").str.strip_chars_end(".").alias("资产编号")]
                    )
                if "資產編號" in df_polars.columns:
                    df_polars = df_polars.with_columns(
                        [pl.col("資產編號").str.strip_chars_end(".").alias("資產編號")]
                    )

                # 移除全空列
                non_null_cols = [
                    col
                    for col in df_polars.columns
                    if not df_polars[col].is_null().all()
                ]
                return df_polars.select(non_null_cols)

            except Exception as e2:
                logger.exception(f"HTML解析也失敗: {e2}")
                raise e2

    def This_Notes_date(self):
        """讀取Notes數據"""
        self.this_Notes_data = None
        if not self.this_Notes_path:
            return
        try:
            self.this_Notes_data = self.safe_read_excel(self.this_Notes_path)
            if self.this_Notes_data is not None:
                logger.info(f"Notes数据列名: {list(self.this_Notes_data.columns)}")
                logger.info(f"Notes数据形状: {self.this_Notes_data.shape}")
        except Exception as e:
            logger.exception(e)

    def This_SFC_date(self):
        """讀取SFC數據"""
        self.this_SFC_data = None
        if not self.this_SFC_path:
            return
        try:
            self.this_SFC_data = self.safe_read_excel(self.this_SFC_path)
            if self.this_SFC_data is not None:
                logger.info(f"SFC数据列名: {list(self.this_SFC_data.columns)}")
                logger.info(f"SFC数据形状: {self.this_SFC_data.shape}")
        except Exception as e:
            logger.exception(e)

    def Notes_SFC_Comparison(self):
        """Notes與SFC數據比較"""
        if self.this_Notes_data is None or self.this_SFC_data is None:
            logger.error("文件讀取錯誤，請檢查文件")
            return
        else:
            try:
                # 智能检测SFC资产编号列名
                sfc_asset_col = None
                if "资产编号" in self.this_SFC_data.columns:
                    sfc_asset_col = "资产编号"
                elif "資產編號" in self.this_SFC_data.columns:
                    sfc_asset_col = "資產編號"
                else:
                    logger.error("SFC数据中未找到资产编号列")
                    return

                # 检测SFC设备编号列（用于资产编号为NA时降级）
                sfc_device_col = None
                for c in ["设备编号", "設備編號"]:
                    if c in self.this_SFC_data.columns:
                        sfc_device_col = c
                        break

                # 提取SFC对比标识（资产编号有效用资产编号，为NA时降级设备编号）
                invalid_values = {"nan", "none", "null", "na", ""}
                sfc_keys = set()
                for row in self.this_SFC_data.iter_rows(named=True):
                    asset_val = str(row.get(sfc_asset_col) or "").strip()
                    device_val = (
                        str(row.get(sfc_device_col) or "").strip()
                        if sfc_device_col
                        else ""
                    )
                    if asset_val and asset_val.lower() not in invalid_values:
                        sfc_keys.add(f"A:{asset_val}")
                    elif device_val and device_val.lower() not in invalid_values:
                        sfc_keys.add(f"D:{device_val}")

                self.this_SFC_assets = list(sfc_keys)

                # 智能检测Notes资产编号列名
                notes_asset_col = None
                if "资产编号" in self.this_Notes_data.columns:
                    notes_asset_col = "资产编号"
                elif "資產編號" in self.this_Notes_data.columns:
                    notes_asset_col = "資產編號"
                else:
                    logger.error("Notes数据中未找到资产编号列")
                    return

                # 检测Notes设备编号列
                notes_device_col = None
                for c in ["设备编号", "設備編號"]:
                    if c in self.this_Notes_data.columns:
                        notes_device_col = c
                        break

                # 提取Notes对比标识（资产编号有效用资产编号，为NA时降级设备编号）
                notes_keys = set()
                for row in self.this_Notes_data.iter_rows(named=True):
                    asset_val = str(row.get(notes_asset_col) or "").strip()
                    device_val = (
                        str(row.get(notes_device_col) or "").strip()
                        if notes_device_col
                        else ""
                    )
                    # 只保留以18-或13-开头的有效资产编号
                    if (
                        asset_val
                        and asset_val.lower() not in invalid_values
                        and (asset_val.startswith("18-") or asset_val.startswith("13-"))
                    ):
                        notes_keys.add(f"A:{asset_val}")
                    elif device_val and device_val.lower() not in invalid_values:
                        notes_keys.add(f"D:{device_val}")

                self.this_Notes_assets = list(notes_keys)

                # 保存列名供后续使用
                self.sfc_asset_col = sfc_asset_col
                self.notes_asset_col = notes_asset_col
                self.sfc_device_col = sfc_device_col
                self.notes_device_col = notes_device_col

                self.Notes_new_assets = notes_keys - sfc_keys
                self.Notes_removed_assets = sfc_keys - notes_keys

            except Exception as e:
                logger.exception(e)

    def Save_Notes_SFC_Comparison(self, output_path):
        """保存Notes與SFC比較結果為Excel文件"""
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            if not self.Notes_new_assets and not self.Notes_removed_assets:
                return
            else:
                # 智能检测列名并使用 Polars 操作
                # 检测Notes数据的列名
                notes_name_col = None
                if "资产名称" in self.this_Notes_data.columns:
                    notes_name_col = "资产名称"
                elif "資產名稱" in self.this_Notes_data.columns:
                    notes_name_col = "資產名稱"
                elif "设备名称" in self.this_Notes_data.columns:
                    notes_name_col = "设备名称"
                else:
                    notes_name_col = self.this_Notes_data.columns[0]

                notes_keeper_col = "保管人"
                if "保管人" not in self.this_Notes_data.columns:
                    for col in self.this_Notes_data.columns:
                        if "保管" in col or "管理" in col or "负责" in col:
                            notes_keeper_col = col
                            break
                    else:
                        notes_keeper_col = self.this_Notes_data.columns[-1]

                # 检测SFC数据的列名
                sfc_name_col = None
                if "设备名称" in self.this_SFC_data.columns:
                    sfc_name_col = "设备名称"
                elif "資產名稱" in self.this_SFC_data.columns:
                    sfc_name_col = "資產名稱"
                elif "资产名称" in self.this_SFC_data.columns:
                    sfc_name_col = "资产名称"
                else:
                    sfc_name_col = self.this_SFC_data.columns[0]

                sfc_keeper_col = "保管人"
                if "保管人" not in self.this_SFC_data.columns:
                    for col in self.this_SFC_data.columns:
                        if "保管" in col or "管理" in col or "负责" in col:
                            sfc_keeper_col = col
                            break
                    else:
                        sfc_keeper_col = self.this_SFC_data.columns[-1]

                invalid_values = ["nan", "none", "null", "na", ""]

                def split_keys(keys):
                    by_asset = {k[2:] for k in keys if k.startswith("A:")}
                    by_device = {k[2:] for k in keys if k.startswith("D:")}
                    return by_asset, by_device

                def filter_by_keys(df, asset_col, device_col, asset_keys, device_keys):
                    conds = []
                    if asset_col and asset_col in df.columns and asset_keys:
                        am = (
                            pl.col(asset_col)
                            .cast(pl.Utf8)
                            .str.strip_chars()
                            .is_in(list(asset_keys))
                        )
                        valid = ~pl.col(asset_col).cast(
                            pl.Utf8
                        ).str.strip_chars().str.to_lowercase().is_in(invalid_values)
                        conds.append(am & valid)
                    if device_col and device_col in df.columns and device_keys:
                        dm = (
                            pl.col(device_col)
                            .cast(pl.Utf8)
                            .str.strip_chars()
                            .is_in(list(device_keys))
                        )
                        conds.append(dm)
                    if not conds:
                        return df.clear()
                    return df.filter(pl.any_horizontal(conds))

                def fill_na_asset(df, asset_col, device_col):
                    if (
                        device_col
                        and asset_col
                        and asset_col in df.columns
                        and device_col in df.columns
                    ):
                        df = df.with_columns(
                            pl.when(
                                pl.col(asset_col)
                                .cast(pl.Utf8)
                                .str.strip_chars()
                                .str.to_lowercase()
                                .is_in(invalid_values)
                            )
                            .then(pl.col(device_col))
                            .otherwise(pl.col(asset_col))
                            .alias(asset_col)
                        )
                    return df

                # --- 新增资产（来自Notes） ---
                new_by_asset, new_by_device = split_keys(self.Notes_new_assets)
                new_select = [
                    notes_name_col,
                    self.notes_asset_col,
                    notes_keeper_col,
                ]
                if (
                    self.notes_device_col
                    and self.notes_device_col not in new_select
                    and self.notes_device_col in self.this_Notes_data.columns
                ):
                    new_select.append(self.notes_device_col)

                new_df = filter_by_keys(
                    self.this_Notes_data,
                    self.notes_asset_col,
                    self.notes_device_col,
                    new_by_asset,
                    new_by_device,
                )
                if not new_df.is_empty():
                    new_df = new_df.select(new_select).unique()
                new_df = fill_na_asset(
                    new_df, self.notes_asset_col, self.notes_device_col
                )
                out_rename_notes = {
                    k: v
                    for k, v in {
                        notes_name_col: "资产名称",
                        self.notes_asset_col: "资产编号",
                        notes_keeper_col: "保管人",
                    }.items()
                    if k != v
                }
                if not new_df.is_empty():
                    new_df = new_df.rename(out_rename_notes).select(
                        ["资产名称", "资产编号", "保管人"]
                    )
                else:
                    new_df = pl.DataFrame(schema=["资产名称", "资产编号", "保管人"])

                # --- 减少资产（来自SFC） ---
                removed_by_asset, removed_by_device = split_keys(
                    self.Notes_removed_assets
                )
                removed_select = [sfc_name_col, self.sfc_asset_col, sfc_keeper_col]
                if (
                    self.sfc_device_col
                    and self.sfc_device_col not in removed_select
                    and self.sfc_device_col in self.this_SFC_data.columns
                ):
                    removed_select.append(self.sfc_device_col)

                removed_df = filter_by_keys(
                    self.this_SFC_data,
                    self.sfc_asset_col,
                    self.sfc_device_col,
                    removed_by_asset,
                    removed_by_device,
                )
                if not removed_df.is_empty():
                    removed_df = removed_df.select(removed_select).unique()
                removed_df = fill_na_asset(
                    removed_df, self.sfc_asset_col, self.sfc_device_col
                )
                out_rename_sfc = {
                    k: v
                    for k, v in {
                        sfc_name_col: "设备名称",
                        self.sfc_asset_col: "资产编号",
                        sfc_keeper_col: "保管人",
                    }.items()
                    if k != v
                }
                if not removed_df.is_empty():
                    removed_df = removed_df.rename(out_rename_sfc).select(
                        ["设备名称", "资产编号", "保管人"]
                    )
                else:
                    removed_df = pl.DataFrame(schema=["设备名称", "资产编号", "保管人"])

                # 初始化工作表
                wb, ws = new_workbook("對比結果")

                # 標題
                ws.merge_cells("A1:D1")
                ws["A1"] = f"本月Notes资产_VS_本月SFC资产 (对比时间{current_time})"
                ws["A1"].font = openpyxl.styles.Font(bold=True, size=14)
                ws["A1"].alignment = openpyxl.styles.Alignment(horizontal="center")
                # 合并并居中 A2:D2 和 A3:D3
                ws.merge_cells("A2:D2")
                ws.merge_cells("A3:D3")
                ws["A2"].alignment = openpyxl.styles.Alignment(horizontal="center")
                ws["A3"].alignment = openpyxl.styles.Alignment(horizontal="center")

                current_row = 3
                if not new_df.is_empty():
                    current_row = write_section(
                        ws,
                        new_df,
                        f"本月比上月新增资产 {len(self.Notes_new_assets)}笔",
                        current_row,
                        "new",
                    )

                if not removed_df.is_empty():
                    current_row = write_section(
                        ws,
                        removed_df,
                        f"本月比上月减少资产 {len(self.Notes_removed_assets)}笔",
                        current_row,
                        "removed",
                    )

                # 列寬與邊框
                from openpyxl.utils import get_column_letter

                widths = [8, 35, 25, 20]
                for i, w in enumerate(widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = w

                wb.save(output_path)

        except Exception as e:
            logger.exception(e)

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

SECTION_FILL = {
    "new": PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),
    "removed": PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"),
    "error": PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid"),
}

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def new_workbook(sheet_name: str) -> tuple[Workbook, Worksheet]:
    """創建 Workbook，激活 sheet 並命名為 sheet_name。"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    return wb, ws


def write_section(
    ws: Worksheet,
    df: pl.DataFrame,
    title: str,
    start_row: int,
    section_type: str,
) -> int:
    """向 ws 寫入一個數據段：段標題 → 空行 → 表頭(No.+df.columns，帶 SECTION_FILL[section_type] 與 thin Border) → 數據行(用 df.iter_rows(named=False) 逐行寫入，空值寫 "")。

    df 為空時寫「（無數據）」。返回下一段的起始行號。
    """
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(bold=True)
    start_row += 2

    if df.is_empty():
        ws.cell(row=start_row, column=1, value="（無數據）")
        return start_row + 2

    headers = ["No."] + list(df.columns)
    fill = SECTION_FILL.get(section_type)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        if fill:
            cell.fill = fill
        cell.border = THIN_BORDER

    start_row += 1

    for row_idx, row_values in enumerate(df.iter_rows(named=False), 1):
        no_cell = ws.cell(row=start_row, column=1, value=row_idx)
        no_cell.border = THIN_BORDER

        for col_idx, value in enumerate(row_values, 2):
            val_str = "" if value is None else str(value)
            cell = ws.cell(row=start_row, column=col_idx, value=val_str)
            cell.border = THIN_BORDER

        start_row += 1

    return start_row + 1


def safe_cell(
    ws: Worksheet, row: int, col: int, value: str | int | float | None
) -> None:
    """寫入單元格，value 為 None 時寫空字串。"""
    ws.cell(row=row, column=col, value="" if value is None else str(value))

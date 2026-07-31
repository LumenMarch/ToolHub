import os  # noqa: E402, I001, UP015, F401
from datetime import datetime  # noqa: E402, I001, UP015, F401

import openpyxl  # noqa: E402, I001, UP015, F401
import polars as pl  # noqa: E402, I001, UP015, F401
from app.services.asset_comparison.excel_writer import new_workbook, safe_cell
from loguru import logger  # noqa: E402, I001, UP015, F401


class Finance_Notes:
    def __init__(self):
        self.input_catalog = None
        self.processed_notes_data = None
        self.date_Notes_assets = None
        self.Finance_path = None
        self.Custodian_path = None
        self.Notes_path = None

        self.removed_assets = []
        self.new_assets = []
        self.Finance_assets = None
        self.Notes_assets = None

        self.Finance_data = None
        self.Notes_data = None
        self.Custodian_txt = None

    def safe_read_excel(self, path):
        """安全讀取Excel並轉為LazyFrame"""
        required_columns = {"資產編號"}

        try:
            # 嘗試使用 Polars 讀取 Excel
            # 注意：设置 infer_schema_length=0 让Polars将所有列读取为字符串，避免类型推断错误
            if self.input_catalog is not None:
                df_polars = self.input_catalog.read_excel(path, infer_schema_length=0)
            else:
                df_polars = pl.read_excel(path, infer_schema_length=0)

            # 檢查是否包含必需的列
            header_columns = set(df_polars.columns)
            if required_columns.intersection(header_columns):
                # 自动去除资产编号列末尾的点号
                if "資產編號" in df_polars.columns:
                    df_polars = df_polars.with_columns(
                        [pl.col("資產編號").str.strip_chars_end(".").alias("資產編號")]
                    )
                # 移除全空列並返回 DataFrame
                non_null_cols = [
                    col
                    for col in df_polars.columns
                    if not df_polars[col].is_null().all()
                ]
                return df_polars.select(non_null_cols)

            # 如果表頭不正確，嘗試查找正確的表頭行
            for i in range(min(5, len(df_polars))):
                row_values = df_polars.row(i)
                clean_columns = {
                    str(col).strip() for col in row_values if col is not None
                }
                if required_columns.intersection(clean_columns):
                    new_columns = [
                        str(col).strip() if col is not None else f"col_{j}"
                        for j, col in enumerate(row_values)
                    ]
                    df_polars = df_polars.slice(i + 1).rename(
                        {
                            old: new
                            for old, new in zip(
                                df_polars.columns, new_columns, strict=False
                            )
                        }
                    )
                    # 自动去除资产编号列末尾的点号
                    if "資產編號" in df_polars.columns:
                        df_polars = df_polars.with_columns(
                            [
                                pl.col("資產編號")
                                .cast(pl.String)
                                .str.strip_chars_end(".")
                                .alias("資產編號")
                            ]
                        )
                    # 移除全空列並返回 DataFrame
                    non_null_cols = [
                        col
                        for col in df_polars.columns
                        if not df_polars[col].is_null().all()
                    ]
                    return df_polars.select(non_null_cols)

            raise ValueError(f"未在文件 {os.path.basename(path)} 中找到完整表頭")

        except Exception as e:
            logger.error(f"读取 {os.path.basename(path)} 失败: {e}")
            raise

    def read_Custodian_data(self):
        self.Custodian_txt = None
        if self.input_catalog is not None:
            self.Custodian_txt = self.input_catalog.read_text_lines(self.Custodian_path)
        else:
            with open(self.Custodian_path, encoding="utf-8") as file:
                lines = file.readlines()  # 读取所有行
                self.Custodian_txt = [line.strip() for line in lines]

    def read_Finance_data(self):
        """讀取財務數據（使用LazyFrame）"""
        self.Finance_data = None
        try:
            df_lazy = self.safe_read_excel(self.Finance_path)
            self.Finance_data = df_lazy
        except Exception as e:
            logger.error(e)

    def read_Notes_data(self):
        """讀取Notes數據（使用LazyFrame）"""
        self.Notes_data = None
        try:
            df_lazy = self.safe_read_excel(self.Notes_path)
            self.Notes_data = df_lazy
        except Exception as e:
            logger.error(e)

    def Finance_Notes_Comparison(self):
        """財務與Notes數據比較（使用DataFrame優化）"""
        # 檢查 DataFrame 是否為空
        if (
            self.Finance_data.head(1).is_empty()
            or self.Notes_data.head(1).is_empty()
            or self.Custodian_txt is None
        ):
            return
        else:
            try:
                # 使用 DataFrame 操作，過濾 Notes 資產（只识别以18-或13-开头的为有效资产编号）
                notes_assets_query = self.Notes_data.filter(
                    pl.col("資產編號").is_not_null()
                    & (
                        pl.col("資產編號")
                        .cast(pl.String)
                        .str.strip_chars()
                        .str.len_chars()
                        > 0
                    )
                    & (
                        pl.col("資產編號").cast(pl.String).str.starts_with("18-")
                        | pl.col("資產編號").cast(pl.String).str.starts_with("13-")
                    )
                ).select("資產編號")

                self.date_Notes_assets = notes_assets_query.to_series().to_list()

                # 過濾財務數據
                finance_assets_query = self.Finance_data.filter(
                    pl.col("保管人員").is_in(self.Custodian_txt)
                ).select("資產編號")

                self.Finance_assets = finance_assets_query.to_series().to_list()

                self.new_assets = set(self.date_Notes_assets) - set(self.Finance_assets)
                self.removed_assets = set(self.Finance_assets) - set(
                    self.date_Notes_assets
                )

                # Notes数据已在读取时处理，直接使用
                self.processed_notes_data = self.Notes_data

            except Exception as e:
                logger.error(e)

    def Save_Finance_Notes_Comparison(self, output_path):
        """保存財務與Notes比較結果為Excel文件"""
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            if not self.new_assets and not self.removed_assets:
                return
            else:
                wb, ws = new_workbook("對比結果")
                # 構建查詢
                new_df = (
                    self.processed_notes_data.filter(
                        pl.col("資產編號").is_in(list(self.new_assets))
                    )
                    .select(["資產名稱", "資產編號", "保管人"])
                    .unique()
                )

                removed_df = (
                    self.Finance_data.filter(
                        pl.col("資產編號").is_in(list(self.removed_assets))
                    )
                    .select(
                        [
                            pl.col("資產名稱"),
                            pl.col("資產編號"),
                            pl.col("保管人員").alias("保管人"),
                        ]
                    )
                    .unique()
                )

                # 標題
                ws.merge_cells("A1:D1")
                ws["A1"] = f"本月Notes资产_VS_本月财务资产 (对比时间{current_time})"
                ws["A1"].font = openpyxl.styles.Font(bold=True, size=14)
                ws["A1"].alignment = openpyxl.styles.Alignment(horizontal="center")
                # 合并并居中 A2:D2 和 A3:D3
                ws.merge_cells("A2:D2")
                ws.merge_cells("A3:D3")
                ws["A2"].alignment = openpyxl.styles.Alignment(horizontal="center")
                ws["A3"].alignment = openpyxl.styles.Alignment(horizontal="center")

                row = 3
                # 新增
                if not new_df.is_empty():
                    ws[f"A{row}"] = f"本月Notes比财务新增资产 {len(self.new_assets)}笔"
                    ws[f"A{row}"].font = openpyxl.styles.Font(bold=True, size=12)
                    row += 1
                    headers = ["No.", "資產名稱", "資產編號", "保管人"]
                    for i, h in enumerate(headers, 1):
                        c = ws.cell(row=row, column=i, value=h)
                        c.font = openpyxl.styles.Font(bold=True)
                        c.fill = openpyxl.styles.PatternFill(
                            start_color="E6F3FF",
                            end_color="E6F3FF",
                            fill_type="solid",
                        )
                    row += 1
                    for i, row_dict in enumerate(new_df.iter_rows(named=True), 1):
                        ws.cell(row=row, column=1, value=i)
                        safe_cell(ws, row, 2, row_dict["資產名稱"])
                        safe_cell(ws, row, 3, row_dict["資產編號"])
                        safe_cell(ws, row, 4, row_dict["保管人"])
                        row += 1
                    row += 1

                # 減少
                if not removed_df.is_empty():
                    ws[f"A{row}"] = (
                        f"本月Notes比财务减少资产 {len(self.removed_assets)}笔"
                    )
                    ws[f"A{row}"].font = openpyxl.styles.Font(bold=True, size=12)
                    row += 1
                    headers = ["No.", "資產名稱", "資產編號", "保管人"]
                    for i, h in enumerate(headers, 1):
                        c = ws.cell(row=row, column=i, value=h)
                        c.font = openpyxl.styles.Font(bold=True)
                        c.fill = openpyxl.styles.PatternFill(
                            start_color="FFE6E6",
                            end_color="FFE6E6",
                            fill_type="solid",
                        )
                    row += 1
                    for i, row_dict in enumerate(removed_df.iter_rows(named=True), 1):
                        ws.cell(row=row, column=1, value=i)
                        safe_cell(ws, row, 2, row_dict["資產名稱"])
                        safe_cell(ws, row, 3, row_dict["資產編號"])
                        safe_cell(ws, row, 4, row_dict["保管人"])
                        row += 1

                # 列寬與邊框
                from openpyxl.styles import Border, Side  # noqa: E402, I001, UP015, F401

                thin = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
                widths = [8, 35, 25, 20]
                for i, w in enumerate(widths, 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
                for r in ws.iter_rows(
                    min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
                ):
                    for cell in r:
                        cell.border = thin
                wb.save(output_path)
        except Exception as e:
            logger.error(e)

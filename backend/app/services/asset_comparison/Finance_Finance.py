import os  # noqa: E402, I001, UP015, F401
from datetime import datetime  # noqa: E402, I001, UP015, F401

import openpyxl  # noqa: E402, I001, UP015, F401
import polars as pl  # noqa: E402, I001, UP015, F401
from app.services.asset_comparison.excel_writer import new_workbook, write_section
from loguru import logger  # noqa: E402, I001, UP015, F401


class Finance_Finance:
    def __init__(self):
        self.input_catalog = None
        self.Department_Difference = []
        self.Custodian_Difference = []
        self.check_Custodian = []
        self.check_Department = []

        self.this_Finance_path = None
        self.last_Finance_path = None
        self.Custodian_path = None
        self.Department_path = None

        self.this_Finance_data = None
        self.last_Finance_data = None
        self.Custodian_data = None
        self.Department_data = None

        self.this_Custodian_assets = None
        self.last_Custodian_assets = None
        self.this_Department_assets = None
        self.last_Department_assets = None

        self.new_Custodian_assets = []
        self.new_Department_assets = []
        self.removed_Custodian_assets = []
        self.removed_Department_assets = []

    def read_Custodian_data(self):
        self.Custodian_data = []
        if self.Custodian_path is None:
            self.Custodian_data = []
        elif self.input_catalog is not None:
            self.Custodian_data = self.input_catalog.read_text_lines(
                self.Custodian_path
            )
        else:
            with open(self.Custodian_path, encoding="utf-8") as file:
                lines = file.readlines()  # 读取所有行
                self.Custodian_data = [line.strip() for line in lines]

    def read_Department_data(self):
        self.Department_data = []
        if self.Department_path is None:
            self.Department_data = []
        elif self.input_catalog is not None:
            self.Department_data = self.input_catalog.read_text_lines(
                self.Department_path
            )
        else:
            with open(self.Department_path, encoding="utf-8") as file:
                lines = file.readlines()  # 读取所有行
                self.Department_data = [line.strip() for line in lines]

    def safe_read_excel(self, path):
        """使用 polars 讀取 Excel 並返回 DataFrame（不是 LazyFrame）"""
        try:
            # 先用 read_excel 讀取數據
            # 注意：设置 infer_schema_length=0 让Polars将所有列读取为字符串，避免类型推断错误
            if self.input_catalog is not None:
                df_eager = self.input_catalog.read_excel(path, infer_schema_length=0)
            else:
                df_eager = pl.read_excel(path, infer_schema_length=0)

            # 檢查表頭中是否包含空值
            header_row = df_eager.columns
            if any(
                cell is None or "__UNNAMED__" in str(cell).strip()
                for cell in header_row
            ):
                # 如果需要重新設置表頭，使用第一行數據作為表頭
                new_columns = [
                    str(col).strip() if col is not None else f"col_{i}"
                    for i, col in enumerate(df_eager.row(0))
                ]
                df_eager = df_eager.slice(1).rename(
                    {
                        old: new
                        for old, new in zip(df_eager.columns, new_columns, strict=False)
                    }
                )
                print(f"文件 {os.path.basename(path)} 的表頭包含空值，已重新設置")

            # 自动去除资产编号列末尾的点号
            if "資產編號" in df_eager.columns:
                df_eager = df_eager.with_columns(
                    [pl.col("資產編號").str.strip_chars_end(".").alias("資產編號")]
                )

            # 返回 DataFrame 而不是 LazyFrame
            return df_eager

        except Exception as e:
            logger.error(f"读取 {os.path.basename(path)} 失败: {e}")
            raise

    def read_this_Finance_data(self):
        """讀取本月財務數據（使用 DataFrame）"""
        self.this_Finance_data = None
        if not self.this_Finance_path:
            return
        self.this_Finance_data = self.safe_read_excel(self.this_Finance_path)

    def read_last_Finance_data(self):
        """讀取上月財務數據（使用 DataFrame）"""
        self.last_Finance_data = None
        if not self.last_Finance_path:
            return
        self.last_Finance_data = self.safe_read_excel(self.last_Finance_path)

    def Finance_check(self):
        try:
            # 先檢查 DataFrame 是否為空
            if self.last_Finance_data.is_empty() or self.this_Finance_data.is_empty():
                logger.error("数据为空")
                return

            print("開始處理財務數據比較...")

            # 使用 DataFrame 構建查詢，在最後一次性執行
            this_Custodian_query = self.this_Finance_data.filter(
                pl.col("保管人員").is_in(self.Custodian_data)
            )

            this_Department_query = self.this_Finance_data.filter(
                pl.col("資產所屬部門代號")
                .cast(pl.String)
                .str.zfill(4)
                .is_in(self.Department_data)
            )

            last_Custodian_query = self.last_Finance_data.filter(
                pl.col("保管人員").is_in(self.Custodian_data)
            )

            last_Department_query = self.last_Finance_data.filter(
                pl.col("資產所屬部門代號")
                .cast(pl.String)
                .str.zfill(4)
                .is_in(self.Department_data)
            )

            # 直接獲得過濾後的數據（不需要 collect）
            this_Custodian_filtered = this_Custodian_query
            this_Department_filtered = this_Department_query
            last_Custodian_filtered = last_Custodian_query
            last_Department_filtered = last_Department_query

            if len(this_Custodian_filtered) > 0:
                # 獲取部門數據並檢查異常
                current_departments = this_Custodian_filtered.select(
                    pl.col("資產所屬部門代號").cast(pl.String).str.zfill(4)
                ).to_series()

                # 找出不在部門列表中的記錄
                invalid_department_mask = ~current_departments.is_in(
                    self.Department_data
                )
                invalid_department_df = this_Custodian_filtered.filter(
                    invalid_department_mask
                )

                self.check_Department = (
                    invalid_department_df.select("資產編號").to_series().to_list()
                    if len(invalid_department_df) > 0
                    else []
                )
                logger.info(f"发现{len(invalid_department_df)}条异常部门记录")
            else:
                self.check_Department = []

            if len(this_Department_filtered) > 0:
                # 獲取保管人數據並檢查異常
                current_custodians = this_Department_filtered.select(
                    "保管人員"
                ).to_series()

                # 找出不在保管人列表中的記錄
                invalid_custodian_mask = ~current_custodians.is_in(self.Custodian_data)
                invalid_custodian_df = this_Department_filtered.filter(
                    invalid_custodian_mask
                )

                self.check_Custodian = (
                    invalid_custodian_df.select("資產編號").to_series().to_list()
                    if len(invalid_custodian_df) > 0
                    else []
                )
                logger.info(f"发现{len(invalid_custodian_df)}条异常保管人记录")
            else:
                self.check_Custodian = []

            # 獲取資產編號列表
            self.this_Custodian_assets = (
                this_Custodian_filtered.select("資產編號").to_series().to_list()
            )
            self.last_Custodian_assets = (
                last_Custodian_filtered.select("資產編號").to_series().to_list()
            )

            self.this_Department_assets = (
                this_Department_filtered.select("資產編號").to_series().to_list()
            )
            self.last_Department_assets = (
                last_Department_filtered.select("資產編號").to_series().to_list()
            )

            self.new_Custodian_assets = set(self.this_Custodian_assets) - set(
                self.last_Custodian_assets
            )
            self.new_Department_assets = set(self.this_Department_assets) - set(
                self.last_Department_assets
            )
            self.removed_Custodian_assets = set(self.last_Custodian_assets) - set(
                self.this_Custodian_assets
            )
            self.removed_Department_assets = set(self.last_Department_assets) - set(
                self.this_Department_assets
            )

            self.Custodian_Difference = abs(
                len(self.new_Custodian_assets) - len(self.removed_Custodian_assets)
            )
            self.Department_Difference = abs(
                len(self.new_Department_assets) - len(self.removed_Department_assets)
            )

        except Exception as e:
            logger.exception(e)

    def Save_Check(self, output_path):
        """保存財務對比結果為Excel文件"""
        try:
            if (
                not self.new_Custodian_assets
                and not self.removed_Custodian_assets
                and not self.new_Department_assets
                and not self.removed_Department_assets
                and not self.check_Custodian
                and not self.check_Department
            ):
                return

            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # 使用 DataFrame 構建查詢
            new_df_query = (
                self.this_Finance_data.filter(
                    pl.col("資產編號").is_in(list(self.new_Custodian_assets))
                )
                .select(["資產名稱", "資產編號", "資產所屬部門代號", "保管人員"])
                .unique()
            )

            removed_df_query = (
                self.last_Finance_data.filter(
                    pl.col("資產編號").is_in(list(self.removed_Custodian_assets))
                )
                .select(["資產名稱", "資產編號", "資產所屬部門代號", "保管人員"])
                .unique()
            )

            new_df_Department_query = (
                self.this_Finance_data.filter(
                    pl.col("資產編號").is_in(list(self.new_Department_assets))
                )
                .select(["資產名稱", "資產編號", "資產所屬部門代號", "保管人員"])
                .unique()
            )

            removed_df_Department_query = (
                self.last_Finance_data.filter(
                    pl.col("資產編號").is_in(list(self.removed_Department_assets))
                )
                .select(["資產名稱", "資產編號", "資產所屬部門代號", "保管人員"])
                .unique()
            )

            Error_Custodian_query = (
                self.this_Finance_data.filter(
                    pl.col("資產編號").is_in(self.check_Custodian)
                )
                .select(["資產名稱", "資產所屬部門代號", "資產編號", "保管人員"])
                .unique()
            )

            Error_Department_query = (
                self.this_Finance_data.filter(
                    pl.col("資產編號").is_in(self.check_Department)
                )
                .select(["資產名稱", "資產所屬部門代號", "資產編號", "保管人員"])
                .unique()
            )

            # 直接獲取數據
            new_df = new_df_query
            removed_df = removed_df_query
            new_df_Department = new_df_Department_query
            removed_df_Department = removed_df_Department_query
            Error_Custodian = Error_Custodian_query
            Error_Department = Error_Department_query

            # 创建Excel文件
            wb, worksheet = new_workbook("對比結果")

            # 设置标题
            worksheet.merge_cells("A1:E1")
            worksheet["A1"] = f"本月财务_VS_上月财务 (对比时间{current_time})"
            worksheet["A1"].font = openpyxl.styles.Font(bold=True, size=14)
            worksheet["A1"].alignment = openpyxl.styles.Alignment(horizontal="center")
            # 合并并居中 A2:D2 和 A3:D3
            worksheet.merge_cells("A2:D2")
            worksheet.merge_cells("A3:D3")
            worksheet["A2"].alignment = openpyxl.styles.Alignment(horizontal="center")
            worksheet["A3"].alignment = openpyxl.styles.Alignment(horizontal="center")

            current_row = 3

            # 处理保管人新增资产
            if not new_df.is_empty():
                current_row = write_section(
                    worksheet,
                    new_df,
                    f"依保管人本月比上月新增资产 {len(self.new_Custodian_assets)}笔",
                    current_row,
                    "new",
                )

            # 处理保管人减少资产
            if not removed_df.is_empty():
                current_row = write_section(
                    worksheet,
                    removed_df,
                    f"依保管人本月比上月减少资产 {len(self.removed_Custodian_assets)}笔",
                    current_row,
                    "removed",
                )

            # 处理部门新增资产
            if not new_df_Department.is_empty():
                current_row = write_section(
                    worksheet,
                    new_df_Department,
                    f"依保管部门本月比上月新增资产 {len(self.new_Department_assets)}笔",
                    current_row,
                    "new",
                )

            # 处理部门减少资产
            if not removed_df_Department.is_empty():
                current_row = write_section(
                    worksheet,
                    removed_df_Department,
                    f"依保管部门本月比上月减少资产 {len(self.removed_Department_assets)}笔",
                    current_row,
                    "removed",
                )

            # 处理保管人错误资产
            if not Error_Custodian.is_empty():
                current_row = write_section(
                    worksheet,
                    Error_Custodian,
                    f"本月财务_VS_上月财务_保管人错误资产 {len(self.check_Custodian)}笔",
                    current_row,
                    "error",
                )

            # 处理部门错误资产
            if not Error_Department.is_empty():
                current_row = write_section(
                    worksheet,
                    Error_Department,
                    f"本月财务_VS_上月财务_部门错误资产 {len(self.check_Department)}笔",
                    current_row,
                    "error",
                )

            # 设置列宽
            for col in ["A", "B", "C", "D", "E"]:
                worksheet.column_dimensions[col].width = 20

            wb.save(output_path)

        except Exception as e:
            logger.exception(e)

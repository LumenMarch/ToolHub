from datetime import datetime  # noqa: E402, I001, UP015, F401

import pandas as pd  # noqa: E402, I001, UP015, F401
from dateutil.relativedelta import relativedelta  # noqa: E402, I001, UP015, F401
from openpyxl.reader.excel import load_workbook  # noqa: E402, I001, UP015, F401
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402, I001, UP015, F401

current_date = datetime.now()
this_month_str = current_date.strftime("%Y%m")
last_month_date = current_date - relativedelta(months=1)
last_month_str = last_month_date.strftime("%Y%m")


def create_excel_template(SAVE_ALL_PATH):
    columnes_a = ["", "", "", "", "", "", ""]
    Default_data = [
        ["", "", "OK", "差異確認OK", "待跟进", "異常", ""],
        [f"{this_month_str!s}", "", "", "", "", "", ""],
        [
            "Item",
            f"{last_month_str!s}",
            f"{this_month_str!s}",
            "與上月差異",
            "5-財務 VS Notes有資產",
            "6-Notes有資產VS SFC",
            "7-Notes客户资产VS客户系统资产",
        ],
        ["1-財務", "", "", "", "", "/", "/"],
        ["2-Notes 有資產", "", "", "", "", "", "/"],
        ["3-SFC", "", "", "", "/", "", "/"],
        ["4-客户系统资产(RFID)", "", "", "", "/", "/", ""],
        ["2a-Notes 客户资产", "", "", "", "/", "/", ""],
        ["2b-Notes 無資產", "", "", "", "/", "/", "/"],
        ["", "", "", "", "", "", ""],  # 空行分隔
        ["比对工具总结说明:", "", "", "", "", "", ""],  # 备注区域标题
        ["1-财务 VS 财务", "", "", "", "", "", ""],
        ["2-Notes VS Notes", "", "", "", "", "", ""],
        ["3-SFC VS SFC", "", "", "", "", "", ""],
        ["4-客户资产 VS 客户资产", "", "", "", "", "", ""],
        ["5-财务 VS Notes", "", "", "", "", "", ""],
        ["6-Notes VS SFC", "", "", "", "", "", ""],
        ["7-Notes客户资产 VS 客户系统资产", "", "", "", "", "", ""],
    ]

    df = pd.DataFrame(Default_data, columns=columnes_a)
    df.to_excel(SAVE_ALL_PATH, index=False, sheet_name="差异总结")

    wb = load_workbook(SAVE_ALL_PATH)
    ws = wb.active
    ws.merge_cells("A3:G3")

    # 创建其他 sheet
    sheet_names = [
        "1-财务 VS 财务",
        "2-Notes VS Notes",
        "3-SFC VS SFC",
        "4-客户资产 VS 客户资产",
        "5-财务 VS Notes",
        "6-Notes VS SFC",
        "7-Notes客户资产 VS 客户系统资产",
    ]
    for name in sheet_names:
        wb.create_sheet(title=name)

    # 填充样式
    fill_color = PatternFill(
        start_color="DADADA", end_color="DADADA", fill_type="solid"
    )
    for row in ws["A4:A10"]:
        for cell in row:
            cell.fill = fill_color
    ws["A3"].fill = fill_color
    for cell in ws["A4:G4"][0]:
        cell.fill = fill_color

    # 备注区域标题样式
    ws["A12"].fill = PatternFill(
        start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"
    )
    ws["A12"].font = Font(name="Arial", size=13, bold=True)
    ws["A12"].alignment = Alignment(horizontal="left", vertical="center")  # 左对齐
    # 备注标题左对齐
    for row_idx in range(13, 20):  # A13到A19
        cell = ws.cell(row=row_idx, column=1)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.font = Font(name="Arial", size=11, bold=True)
    # 备注内容左对齐并启用自动换行
    for row_idx in range(13, 20):
        for col_idx in range(2, 8):  # B到G列
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )

    colors = {
        3: "D0F1AD",  # OK
        4: "71D0F1",  # 差異確認OK
        5: "FFEE00",  # 待確定
        6: "EC4337",  # 異常
    }
    for col in range(3, 7):
        cell = ws.cell(row=2, column=col)
        cell.fill = PatternFill(
            start_color=colors[col], end_color=colors[col], fill_type="solid"
        )

    # 字體和對齊
    font_style = Font(name="Arial", size=13)
    alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        row_index = row[0].row
        for cell in row:
            cell.font = font_style
            cell.alignment = alignment
        # 所有行保持相同高度25
        ws.row_dimensions[row_index].height = 25

    # 欄寬設定
    column_widths = {
        "A": 32.64,
        "B": 12,
        "C": 12,
        "D": 15,
        "E": 27.31,
        "F": 27.31,
        "G": 45,
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 邊框設定
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = thin_border

    wb.save(SAVE_ALL_PATH)

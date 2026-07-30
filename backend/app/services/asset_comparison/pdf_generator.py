#!/usr/bin/env python3

"""
PDF生成模块
从Excel文件转换为PDF格式，支持多个sheet页面
新增：从原始数据直接生成PDF的方法
"""

import logging  # noqa: E402, I001, UP015, F401
import os  # noqa: E402, I001, UP015, F401
from datetime import datetime  # noqa: E402, I001, UP015, F401
from pathlib import Path  # noqa: E402, I001, UP015, F401
from threading import Lock  # noqa: E402, I001, UP015, F401
from time import perf_counter  # noqa: E402, I001, UP015, F401
from unicodedata import east_asian_width  # noqa: E402, I001, UP015, F401
from xml.sax.saxutils import escape  # noqa: E402, I001, UP015, F401

import openpyxl  # noqa: E402, I001, UP015, F401
from PyPDF2 import PdfMerger  # noqa: E402, I001, UP015, F401
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT  # noqa: E402, I001, UP015, F401
from reportlab.pdfbase import pdfmetrics  # noqa: E402, I001, UP015, F401
from reportlab.pdfbase.ttfonts import TTFont  # noqa: E402, I001, UP015, F401

# 新增：从原始数据生成PDF的库
try:
    from PyPDF2 import PdfMerger  # noqa: E402, I001, UP015, F401
    from reportlab.lib import colors  # noqa: E402, I001, UP015, F401
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT  # noqa: E402, I001, UP015, F401
    from reportlab.lib.pagesizes import (  # noqa: E402, I001, UP015, F401
        A4,
        landscape,
        letter,
    )
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet  # noqa: E402, I001, UP015, F401
    from reportlab.lib.units import inch, mm  # noqa: E402, I001, UP015, F401
    from reportlab.pdfbase import pdfmetrics  # noqa: E402, I001, UP015, F401
    from reportlab.pdfbase.ttfonts import TTFont  # noqa: E402, I001, UP015, F401
    from reportlab.platypus import (  # noqa: E402, I001, UP015, F401
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    print("Warning: reportlab not available. Install with: pip install reportlab")

MAPLE_MONO_FONT_NAME = "MapleMonoCN"
MAPLE_MONO_BOLD_FONT_NAME = "MapleMonoCN-Bold"
FONT_DIRECTORY = Path(__file__).with_name("fonts")
MAPLE_MONO_FONT_PATH = FONT_DIRECTORY / "MapleMono-CN-Regular.ttf"
MAPLE_MONO_BOLD_FONT_PATH = FONT_DIRECTORY / "MapleMono-CN-Bold.ttf"
FONT_REGISTRATION_LOCK = Lock()


class RawDataToPDFConverter:
    """从原始数据直接生成PDF的转换器"""

    def __init__(self):
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab库未安装，无法使用RawDataToPDFConverter")

        self.styles = getSampleStyleSheet()
        self.setup_chinese_font()
        self.setup_custom_styles()

    def setup_chinese_font(self):
        """设置中文字体支持"""
        missing_paths = [
            str(font_path)
            for font_path in (MAPLE_MONO_FONT_PATH, MAPLE_MONO_BOLD_FONT_PATH)
            if not font_path.is_file()
        ]
        if missing_paths:
            raise FileNotFoundError(
                f"Maple Mono CN 字体文件缺失: {', '.join(missing_paths)}"
            )

        with FONT_REGISTRATION_LOCK:
            registered_fonts = set(pdfmetrics.getRegisteredFontNames())
            if MAPLE_MONO_FONT_NAME not in registered_fonts:
                pdfmetrics.registerFont(
                    TTFont(MAPLE_MONO_FONT_NAME, str(MAPLE_MONO_FONT_PATH))
                )
            if MAPLE_MONO_BOLD_FONT_NAME not in registered_fonts:
                pdfmetrics.registerFont(
                    TTFont(MAPLE_MONO_BOLD_FONT_NAME, str(MAPLE_MONO_BOLD_FONT_PATH))
                )
            pdfmetrics.registerFontFamily(
                MAPLE_MONO_FONT_NAME,
                normal=MAPLE_MONO_FONT_NAME,
                bold=MAPLE_MONO_BOLD_FONT_NAME,
                italic=MAPLE_MONO_FONT_NAME,
                boldItalic=MAPLE_MONO_BOLD_FONT_NAME,
            )

        self.chinese_font = MAPLE_MONO_FONT_NAME
        self.chinese_bold_font = MAPLE_MONO_BOLD_FONT_NAME

    def setup_custom_styles(self):
        """设置自定义样式"""
        # 标题样式
        self.title_style = ParagraphStyle(
            "CustomTitle",
            parent=self.styles["Heading1"],
            fontSize=18,
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName=self.chinese_font,
        )

        # 副标题样式
        self.subtitle_style = ParagraphStyle(
            "CustomSubtitle",
            parent=self.styles["Heading2"],
            fontSize=14,
            spaceAfter=15,
            alignment=TA_CENTER,
            fontName=self.chinese_font,
        )

        # 正文样式
        self.body_style = ParagraphStyle(
            "CustomBody",
            parent=self.styles["Normal"],
            fontSize=10,
            spaceAfter=6,
            fontName=self.chinese_font,
        )

    @staticmethod
    def _text_width_units(value):
        """按字符显示宽度估算文本占用空间"""
        text = str(value)
        lines = text.splitlines() or [""]
        return max(
            sum(2 if east_asian_width(char) in {"W", "F"} else 1 for char in line)
            for line in lines
        )

    def _calculate_column_widths(self, table_data, available_width):
        """根据内容权重分配列宽，并确保总宽度不超过页面内容区"""
        column_count = len(table_data[0])
        desired_widths = []
        for column_index in range(column_count):
            content_width = max(
                self._text_width_units(row[column_index]) for row in table_data
            )
            desired_widths.append(min(max(content_width * 4.2 + 12, 30), 100))

        desired_total = sum(desired_widths)
        if desired_total <= available_width:
            return desired_widths

        minimum_width = min(30, available_width / column_count)
        remaining_width = available_width - minimum_width * column_count
        flexible_widths = [max(width - minimum_width, 1) for width in desired_widths]
        flexible_total = sum(flexible_widths)
        return [
            minimum_width + remaining_width * flexible_width / flexible_total
            for flexible_width in flexible_widths
        ]

    @staticmethod
    def _table_paragraph(value, style):
        """将单元格内容转换为支持自动换行的安全段落"""
        text = escape(str(value))
        text = text.replace("\r\n", "\n").replace("\r", "\n")
        return Paragraph(text.replace("\n", "<br/>"), style)

    def create_data_table(
        self, sheet_name, data, available_width=None, comment: str = ""
    ):
        """创建数据表格"""
        elements = []

        # 生成统计信息标题
        stats_title = self.generate_stats_title(sheet_name, data)
        elements.append(Paragraph(stats_title, self.subtitle_style))
        elements.append(Spacer(1, 10 * mm))

        # 如果有注释，在统计信息下面显示
        if comment and comment.strip():
            elements.append(Paragraph(f"注释：{comment.strip()}", self.body_style))
            elements.append(Spacer(1, 10 * mm))

        # 处理数据
        df = data.collect() if hasattr(data, "collect") else data

        # 检查数据是否为空
        is_empty = df is None or df.is_empty()

        if is_empty:
            elements.append(Paragraph("暂无数据", self.body_style))
            return elements

        # 准备表格数据
        table_data = []

        # 添加表头
        headers = [str(col) for col in df.columns]
        table_data.append(headers)

        # 添加数据行（限制行数以避免PDF过大）
        max_rows = 99999  # 每个表格最多显示100行
        for row_dict in df.head(max_rows).iter_rows(named=True):
            row_data = []
            for value in row_dict.values():
                if value is None:
                    row_data.append("")
                else:
                    # 处理长文本
                    text = str(value)
                    if len(text) > 50:
                        text = text[:47] + "..."
                    row_data.append(text)
            table_data.append(row_data)

        # 如果数据被截断，添加说明
        if len(df) > max_rows:
            elements.append(
                Paragraph(
                    f"注意：仅显示前{max_rows}行数据，共{len(df)}行", self.body_style
                )
            )
            elements.append(Spacer(1, 5 * mm))

        if available_width is None:
            available_width = landscape(A4)[0] - 40 * mm

        column_count = len(headers)
        if column_count <= 8:
            header_font_size, body_font_size = 9, 8
        elif column_count <= 12:
            header_font_size, body_font_size = 8, 7
        elif column_count <= 18:
            header_font_size, body_font_size = 7, 6.5
        else:
            header_font_size, body_font_size = 6, 5.5

        header_style = ParagraphStyle(
            "TableHeader",
            fontName=self.chinese_font,
            fontSize=header_font_size,
            leading=header_font_size + 1,
            alignment=TA_CENTER,
            textColor=colors.whitesmoke,
            wordWrap="CJK",
            splitLongWords=True,
        )
        cell_style = ParagraphStyle(
            "TableCell",
            fontName=self.chinese_font,
            fontSize=body_font_size,
            leading=body_font_size + 1,
            alignment=TA_CENTER,
            wordWrap="CJK",
            splitLongWords=True,
        )
        column_widths = self._calculate_column_widths(table_data, available_width)
        wrapped_table_data = [
            [
                self._table_paragraph(
                    value, header_style if row_index == 0 else cell_style
                )
                for value in row
            ]
            for row_index, row in enumerate(table_data)
        ]

        # 创建表格
        table = Table(
            wrapped_table_data,
            colWidths=column_widths,
            repeatRows=1,
            hAlign="CENTER",
        )

        # 设置表格样式
        table_style = TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )

        table.setStyle(table_style)
        elements.append(table)

        return elements

    def generate_stats_title(self, sheet_name, data):
        """生成统计信息标题"""
        try:
            # 处理数据
            df = data.collect() if hasattr(data, "collect") else data

            # 检查数据是否为空
            is_empty = df is None or df.is_empty()

            if is_empty:
                return "暂无数据"

            # 统计各分类的数量
            if "分类" in df.columns:
                category_counts = df["分类"].value_counts()

                # 构建统计信息
                stats_parts = []
                for category, count in category_counts.iter_rows():
                    # 跳过注释分类，不显示"注释多少笔"
                    if category == "注释":
                        continue
                    elif "新增" in category or "增加" in category:
                        if "依保管人" in category:
                            stats_parts.append(f"依保管人增加{count}笔")
                        elif "依部门" in category:
                            stats_parts.append(f"依部门增加{count}笔")
                        else:
                            stats_parts.append(f"增加{count}笔")
                    elif "减少" in category or "删除" in category:
                        if "依保管人" in category:
                            stats_parts.append(f"依保管人减少{count}笔")
                        elif "依部门" in category:
                            stats_parts.append(f"依部门减少{count}笔")
                        else:
                            stats_parts.append(f"减少{count}笔")
                    else:
                        stats_parts.append(f"{category}{count}笔")

                if stats_parts:
                    return " | ".join(stats_parts)

            # 如果没有分类列，返回总行数
            total_rows = len(df)
            return f"总计{total_rows}笔数据"

        except Exception as e:
            logging.error(f"生成统计标题失败: {e!s}")
            return self.translate_sheet_name(sheet_name)

    def translate_sheet_name(self, sheet_name):
        """将sheet名称转换为中文显示名称"""
        translations = {
            "差异总结": "差异总结",
            "1-财务 VS 财务": "1-财务 VS 财务",
            "2-Notes VS Notes": "2-Notes VS Notes",
            "3-SFC VS SFC": "3-SFC VS SFC",
            "4-客户资产 VS 客户资产": "4-客户资产 VS 客户资产",
            "5-财务 VS Notes": "5-财务 VS Notes",
            "6-Notes VS SFC": "6-Notes VS SFC",
            "7-Notes客户资产 VS 客户系统资产": "7-Notes客户资产 VS 客户系统资产",
        }
        return translations.get(sheet_name, sheet_name)

    def create_single_sheet_pdf(
        self, sheet_name, data, pdf_path, title=None, comment: str = ""
    ):
        """为单个sheet创建PDF"""
        if not REPORTLAB_AVAILABLE:
            logging.error("reportlab库未安装，无法生成PDF")
            return False

        try:
            # 创建PDF文档
            doc = SimpleDocTemplate(
                pdf_path,
                pagesize=landscape(A4),
                rightMargin=20 * mm,
                leftMargin=20 * mm,
                topMargin=20 * mm,
                bottomMargin=20 * mm,
            )

            # 构建PDF内容
            story = []

            # 添加标题
            if title is None:
                title = self.translate_sheet_name(sheet_name)

            story.append(Paragraph(title, self.title_style))
            story.append(Spacer(1, 10 * mm))

            # 添加生成时间
            current_time = datetime.now().strftime("%Y年%m月%d日 %H:%M:%S")
            time_text = f"生成时间：{current_time}"
            story.append(Paragraph(time_text, self.body_style))
            story.append(Spacer(1, 15 * mm))

            # 添加数据表格
            story.extend(
                self.create_data_table(sheet_name, data, doc.width, comment=comment)
            )

            # 生成PDF
            doc.build(story)
            logging.info(f"单个sheet PDF文件已生成: {pdf_path}")
            return True

        except Exception as e:
            logging.error(f"单个sheet PDF生成失败: {e!s}")
            return False

    def merge_pdfs(self, pdf_paths, output_path):
        """合并多个PDF文件"""
        try:
            if not pdf_paths:
                logging.error("没有PDF文件路径提供")
                return False

            merger = PdfMerger()
            valid_pdfs = []

            for pdf_path in pdf_paths:
                if os.path.exists(pdf_path):
                    try:
                        merger.append(pdf_path)
                        valid_pdfs.append(pdf_path)
                        logging.info(f"添加PDF文件: {pdf_path}")
                    except Exception as e:
                        logging.error(f"添加PDF文件失败 {pdf_path}: {e!s}")
                else:
                    logging.warning(f"PDF文件不存在: {pdf_path}")

            if not valid_pdfs:
                logging.error("没有有效的PDF文件可以合并")
                return False

            merger.write(output_path)
            merger.close()

            logging.info(f"PDF文件合并成功: {output_path}")
            return True

        except Exception as e:
            logging.error(f"PDF合并失败: {e!s}")
            return False

    def create_pdf_from_sheets(
        self, sheet_data_dict, output_pdf_path, title="资产对比报告"
    ):
        """从多个sheet数据创建合并的PDF"""
        if not REPORTLAB_AVAILABLE:
            logging.error("reportlab库未安装，无法生成PDF")
            return False

        try:
            # 创建临时目录
            temp_dir = os.path.join(os.path.dirname(output_pdf_path), "temp_pdfs")
            os.makedirs(temp_dir, exist_ok=True)

            # 生成每个sheet的PDF
            pdf_paths = []
            for sheet_name, (data, comment) in sheet_data_dict.items():
                if data is not None:
                    sheet_started_at = perf_counter()
                    # 检查数据是否为空
                    is_empty = False
                    if hasattr(data, "is_empty"):
                        is_empty = data.is_empty()
                    elif hasattr(data, "height"):
                        is_empty = data.height == 0

                    if not is_empty:
                        # 生成单个sheet的PDF
                        temp_pdf_path = os.path.join(temp_dir, f"{sheet_name}.pdf")
                        sheet_title = self.translate_sheet_name(sheet_name)

                        success = self.create_single_sheet_pdf(
                            sheet_name,
                            data,
                            temp_pdf_path,
                            sheet_title,
                            comment=comment,
                        )
                        row_count = (
                            len(data)
                            if hasattr(data, "__len__")
                            else getattr(data, "height", None)
                        )
                        elapsed = perf_counter() - sheet_started_at
                        logging.info(
                            "pdf: stage=generate_sheet elapsed=%.3fs "
                            "sheet=%r rows=%r ok=%r",
                            elapsed,
                            sheet_name,
                            row_count,
                            success,
                        )
                        if success:
                            pdf_paths.append(temp_pdf_path)
                            logging.info(f"Sheet PDF生成成功: {sheet_name}")
                        else:
                            logging.warning(f"Sheet PDF生成失败: {sheet_name}")

            # 合并所有PDF
            if pdf_paths:
                success = self.merge_pdfs(pdf_paths, output_pdf_path)

                # 清理临时文件
                for pdf_path in pdf_paths:
                    try:
                        os.remove(pdf_path)
                    except Exception:
                        pass

                try:
                    os.rmdir(temp_dir)
                except Exception:
                    pass

                return success
            else:
                logging.warning("没有可用的PDF文件进行合并")
                return False

        except Exception as e:
            logging.error(f"从sheets创建PDF失败: {e!s}")
            return False


def create_pdf_from_sheets(sheet_data_dict, output_pdf_path, title="资产对比报告"):
    """便捷函数：从多个sheet数据创建合并的PDF"""
    if not REPORTLAB_AVAILABLE:
        logging.error("reportlab库未安装，请运行: pip install reportlab")
        return False

    converter = RawDataToPDFConverter()
    return converter.create_pdf_from_sheets(sheet_data_dict, output_pdf_path, title)


def excel_sheet_to_pdf(excel_path, sheet_name, pdf_path=None):
    """从Excel文件指定sheet生成PDF（保留Excel样式）"""
    if not REPORTLAB_AVAILABLE:
        logging.error("reportlab库未安装，请运行: pip install reportlab")
        return False

    try:
        # 如果没有指定pdf_path，自动生成带日期和sheet名称的文件名
        if pdf_path is None:
            current_date = datetime.now().strftime("%Y%m%d")
            base_name = os.path.splitext(os.path.basename(excel_path))[0]
            pdf_path = f"{base_name}_{sheet_name}_{current_date}.pdf"

        # 创建转换器实例
        converter = RawDataToPDFConverter()

        # 打开 Excel
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"❌ 没找到 sheet: {sheet_name}")
        ws = wb[sheet_name]

        # 收集数据和样式
        data = []
        table_style_cmds = []

        # 只读取到最后一个有内容的行（最多20行，避免读取过多空白行）
        max_row = min(ws.max_row, 20)  # 限制最大行数为20行

        for r_idx, row in enumerate(ws.iter_rows(max_row=max_row)):
            row_data = []
            # 检查这一行是否完全为空
            is_empty_row = all(
                cell.value is None or str(cell.value).strip() == "" for cell in row
            )

            # 如果是备注区域之后的完全空行，跳过
            if r_idx >= 19 and is_empty_row:
                continue

            for c_idx, cell in enumerate(row):
                val = "" if cell.value is None else str(cell.value)
                row_data.append(val)

                # 样式 - 背景色
                fill = cell.fill
                if (
                    fill
                    and fill.fgColor
                    and fill.fgColor.type == "rgb"
                    and fill.fgColor.rgb not in ("00000000", "FFFFFFFF")
                ):
                    hex_color = fill.fgColor.rgb  # e.g. 'FF00FF00'
                    rgb = tuple(
                        int(hex_color[i : i + 2], 16) / 255 for i in (2, 4, 6)
                    )  # 转换为 0-1
                    table_style_cmds.append(
                        ("BACKGROUND", (c_idx, r_idx), (c_idx, r_idx), rgb)
                    )

                # 样式 - 字体加粗
                if cell.font and cell.font.bold:
                    table_style_cmds.append(
                        (
                            "FONTNAME",
                            (c_idx, r_idx),
                            (c_idx, r_idx),
                            converter.chinese_bold_font,
                        )
                    )

                # 样式 - 字体颜色
                if cell.font and cell.font.color and cell.font.color.type == "rgb":
                    hex_color = cell.font.color.rgb
                    if hex_color not in ("00000000", "FFFFFFFF"):
                        rgb = tuple(
                            int(hex_color[i : i + 2], 16) / 255 for i in (2, 4, 6)
                        )
                        table_style_cmds.append(
                            ("TEXTCOLOR", (c_idx, r_idx), (c_idx, r_idx), rgb)
                        )

            data.append(row_data)

        # 处理合并单元格
        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            # 转换为0基索引
            start_col = min_col - 1
            start_row = min_row - 1
            end_col = max_col - 1
            end_row = max_row - 1

            # 添加SPAN样式，合并单元格
            table_style_cmds.append(
                ("SPAN", (start_col, start_row), (end_col, end_row))
            )

            # 备注标题行（第12行，A12:G12）设置为左对齐
            if start_row == 11:  # 第12行（0基索引为11）
                table_style_cmds.append(
                    ("ALIGN", (start_col, start_row), (end_col, end_row), "LEFT")
                )
            # 备注内容区域（第13-19行的B-G列）设置为左对齐，顶部对齐
            elif (
                start_row >= 12 and start_row <= 18 and start_col >= 1
            ):  # 第13-19行，B列起
                table_style_cmds.append(
                    ("ALIGN", (start_col, start_row), (end_col, end_row), "LEFT")
                )
                table_style_cmds.append(
                    ("VALIGN", (start_col, start_row), (end_col, end_row), "TOP")
                )

        # 生成 PDF
        from reportlab.lib.enums import TA_CENTER  # noqa: E402, I001, UP015, F401
        from reportlab.lib.pagesizes import A4, landscape  # noqa: E402, I001, UP015, F401
        from reportlab.lib.styles import ParagraphStyle  # noqa: E402, I001, UP015, F401
        from reportlab.platypus import Paragraph  # noqa: E402, I001, UP015, F401

        # 创建PDF文档
        pdf = SimpleDocTemplate(pdf_path, pagesize=landscape(A4))

        # 创建内容列表
        story = []

        # 添加表头（sheet名称）
        title_style = ParagraphStyle(
            "CustomTitle",
            fontSize=16,
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName=converter.chinese_font,
        )
        story.append(Paragraph(sheet_name, title_style))

        # 添加生成时间
        current_time = datetime.now().strftime("%Y年%m月%d日 %H:%M:%S")
        time_style = ParagraphStyle(
            "TimeStyle",
            fontSize=10,
            spaceAfter=15,
            alignment=TA_CENTER,
            fontName=converter.chinese_font,
        )
        story.append(Paragraph(f"生成时间：{current_time}", time_style))

        # 添加表格
        table = Table(data)

        # 默认样式 + 单元格样式
        base_style = [
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTNAME", (0, 0), (-1, -1), converter.chinese_font),  # 设置中文字体
            ("FONTSIZE", (0, 0), (-1, -1), 10),  # 设置字体大小
        ]
        table.setStyle(TableStyle(base_style + table_style_cmds))

        story.append(table)

        # 生成PDF
        pdf.build(story)
        logging.info(f"✅ PDF 已生成：{pdf_path}")
        return True

    except Exception as e:
        logging.error(f"Excel转PDF失败: {e!s}")
        return False

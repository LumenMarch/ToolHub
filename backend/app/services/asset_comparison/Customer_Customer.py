import os  # noqa: E402, I001, UP015, F401
from datetime import datetime  # noqa: E402, I001, UP015, F401

import openpyxl  # noqa: E402, I001, UP015, F401
import polars as pl  # noqa: E402, I001, UP015, F401
from app.services.asset_comparison.const import CUSTOMER_CUSTOMER_SAVE_PATH  # noqa: E402, I001, UP015, F401
from app.services.asset_comparison.excel_writer import new_workbook, safe_cell
from loguru import logger  # noqa: E402, I001, UP015, F401


class Customer_Customer:
    def __init__(self):
        self.input_catalog = None
        self.last_Customer_assets = None
        self.this_Customer_assets = None
        self.this_Customer_path = None
        self.last_Customer_path = None
        self.Custodian_DRI_path = None

        self.Customer_DRI_data = None
        self.this_Customer_data = None
        self.last_Customer_data = None

        self.new_Customer_assets = None
        self.removed_Customer_assets = None

    def safe_read_excel(self, path):
        """安全讀取Excel並轉為LazyFrame"""
        required_columns = {"資產編號", "RFID", "DRI"}

        try:
            # 嘗試使用 Polars 讀取 Excel
            # 注意：设置 infer_schema_length=0 让Polars将所有列读取为字符串，避免类型推断错误
            if self.input_catalog is not None:
                df_polars = self.input_catalog.read_excel(path, infer_schema_length=0)
            else:
                df_polars = pl.read_excel(path, infer_schema_length=0)

            # 檢查是否包含必需的列
            header_columns = set(df_polars.columns)
            if required_columns.intersection(header_columns):
                # 自动去除资产编号列末尾的点号
                if "資產編號" in df_polars.columns:
                    df_polars = df_polars.with_columns(
                        [pl.col("資產編號").str.strip_chars_end(".").alias("資產編號")]
                    )
                # 移除全空列並返回 DataFrame
                non_null_cols = [
                    col
                    for col in df_polars.columns
                    if not df_polars[col].is_null().all()
                ]
                return df_polars.select(non_null_cols)

            # 如果表頭不正確，嘗試查找正確的表頭行
            for i in range(min(5, len(df_polars))):
                row_values = df_polars.row(i)
                clean_columns = {
                    str(col).strip() for col in row_values if col is not None
                }
                if required_columns.intersection(clean_columns):
                    new_columns = [
                        str(col).strip() if col is not None else f"col_{j}"
                        for j, col in enumerate(row_values)
                    ]
                    df_polars = df_polars.slice(i + 1).rename(
                        {
                            old: new
                            for old, new in zip(
                                df_polars.columns, new_columns, strict=False
                            )
                        }
                    )
                    # 自动去除资产编号列末尾的点号
                    if "資產編號" in df_polars.columns:
                        df_polars = df_polars.with_columns(
                            [
                                pl.col("資產編號")
                                .str.strip_chars_end(".")
                                .alias("資產編號")
                            ]
                        )
                    # 移除全空列並返回 DataFrame
                    non_null_cols = [
                        col
                        for col in df_polars.columns
                        if not df_polars[col].is_null().all()
                    ]
                    return df_polars.select(non_null_cols)

            raise ValueError(f"未在文件 {os.path.basename(path)} 中找到完整表頭")

        except Exception as e:
            logger.exception(f"读取 {os.path.basename(path)} 失败: {e}")
            raise

    def get_Customer_DRI(self):
        self.Customer_DRI_data = None
        if self.input_catalog is not None:
            self.Customer_DRI_data = self.input_catalog.read_text_lines(
                self.Custodian_DRI_path
            )
        else:
            with open(self.Custodian_DRI_path, encoding="utf-8") as file:
                lines = file.readlines()  # 读取所有行
                self.Customer_DRI_data = [line.strip() for line in lines]

    def read_this_Customer_data(self):
        """讀取本月客戶數據（使用LazyFrame）"""
        self.this_Customer_data = None
        if not self.this_Customer_path:
            return
        df_lazy = self.safe_read_excel(self.this_Customer_path)
        self.this_Customer_data = df_lazy

    def read_last_Customer_data(self):
        """讀取上月客戶數據（使用LazyFrame）"""
        self.last_Customer_data = None
        if not self.last_Customer_path:
            return
        df_lazy = self.safe_read_excel(self.last_Customer_path)
        self.last_Customer_data = df_lazy

    def Customer_Customer_Comparison(self):
        """客戶與客戶數據比較（使用LazyFrame優化）"""
        self.new_Customer_assets = None
        self.removed_Customer_assets = None

        # 檢查 DataFrame 是否為空
        if (
            self.this_Customer_data.head(1).is_empty()
            or self.last_Customer_data.head(1).is_empty()
        ):
            return
        else:
            try:
                # 使用 DataFrame 操作，過濾符合 DRI 條件的客戶資產
                this_filtered_query = self.this_Customer_data.filter(
                    pl.col("DRI").is_in(self.Customer_DRI_data)
                ).select("Asset ID")

                last_filtered_query = self.last_Customer_data.filter(
                    pl.col("DRI").is_in(self.Customer_DRI_data)
                ).select("Asset ID")

                # 直接執行查詢
                self.this_Customer_assets = this_filtered_query.to_series().to_list()
                self.last_Customer_assets = last_filtered_query.to_series().to_list()

                self.new_Customer_assets = set(self.this_Customer_assets) - set(
                    self.last_Customer_assets
                )
                self.removed_Customer_assets = set(self.last_Customer_assets) - set(
                    self.this_Customer_assets
                )

            except Exception as e:
                logger.exception(e)

    def Save_Customer_Customer_Comparison(self):
        """保存客戶與客戶比較結果為Excel文件"""
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            if not self.new_Customer_assets and not self.removed_Customer_assets:
                return
            else:
                wb, ws = new_workbook("對比結果")
                # 查詢數據
                new_df = (
                    self.this_Customer_data.filter(
                        pl.col("Asset ID").is_in(list(self.new_Customer_assets))
                    )
                    .select(["Model Number", "Serial Number", "Asset ID", "DRI"])
                    .unique()
                )

                remove_df = (
                    self.last_Customer_data.filter(
                        pl.col("Asset ID").is_in(list(self.removed_Customer_assets))
                    )
                    .select(["Model Number", "Serial Number", "Asset ID", "DRI"])
                    .unique()
                )

                # 標題
                ws.merge_cells("A1:E1")
                ws["A1"] = f"本月客户资产_VS_上月客户资产 (对比时间{current_time})"
                ws["A1"].font = openpyxl.styles.Font(bold=True, size=14)
                ws["A1"].alignment = openpyxl.styles.Alignment(horizontal="center")
                # 合并并居中 A2:D2 和 A3:D3
                ws.merge_cells("A2:D2")
                ws.merge_cells("A3:D3")
                ws["A2"].alignment = openpyxl.styles.Alignment(horizontal="center")
                ws["A3"].alignment = openpyxl.styles.Alignment(horizontal="center")

                row = 3
                # 新增
                if not new_df.is_empty():
                    ws[f"A{row}"] = (
                        f"本月比上月新增客户资产{len(self.new_Customer_assets)}笔"
                    )
                    ws[f"A{row}"].font = openpyxl.styles.Font(bold=True, size=12)
                    row += 1
                    headers = [
                        "No.",
                        "Model Number",
                        "Serial Number",
                        "RFID",
                        "DRI",
                    ]
                    for i, h in enumerate(headers, 1):
                        c = ws.cell(row=row, column=i, value=h)
                        c.font = openpyxl.styles.Font(bold=True)
                        c.fill = openpyxl.styles.PatternFill(
                            start_color="E6F3FF",
                            end_color="E6F3FF",
                            fill_type="solid",
                        )
                    row += 1
                    for i, row_dict in enumerate(new_df.iter_rows(named=True), 1):
                        ws.cell(row=row, column=1, value=i)
                        safe_cell(ws, row, 2, row_dict["Model Number"])
                        safe_cell(ws, row, 3, row_dict["Serial Number"])
                        safe_cell(ws, row, 4, row_dict["Asset ID"])
                        safe_cell(ws, row, 5, row_dict["DRI"])
                        row += 1
                    row += 1

                # 減少
                if not remove_df.is_empty():
                    ws[f"A{row}"] = (
                        f"本月比上月减少客户资产{len(self.removed_Customer_assets)}笔"
                    )
                    ws[f"A{row}"].font = openpyxl.styles.Font(bold=True, size=12)
                    row += 1
                    headers = [
                        "No.",
                        "Model Number",
                        "Serial Number",
                        "Asset ID",
                        "DRI",
                    ]
                    for i, h in enumerate(headers, 1):
                        c = ws.cell(row=row, column=i, value=h)
                        c.font = openpyxl.styles.Font(bold=True)
                        c.fill = openpyxl.styles.PatternFill(
                            start_color="FFE6E6",
                            end_color="FFE6E6",
                            fill_type="solid",
                        )
                    row += 1
                    for i, row_dict in enumerate(remove_df.iter_rows(named=True), 1):
                        ws.cell(row=row, column=1, value=i)
                        safe_cell(ws, row, 2, row_dict["Model Number"])
                        safe_cell(ws, row, 3, row_dict["Serial Number"])
                        safe_cell(ws, row, 4, row_dict["Asset ID"])
                        safe_cell(ws, row, 5, row_dict["DRI"])
                        row += 1

                # 列寬與邊框
                from openpyxl.styles import Border, Side  # noqa: E402, I001, UP015, F401

                thin = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
                widths = [8, 30, 30, 28, 18]
                for i, w in enumerate(widths, 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
                for r in ws.iter_rows(
                    min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
                ):
                    for cell in r:
                        cell.border = thin

                wb.save(CUSTOMER_CUSTOMER_SAVE_PATH)
        except Exception as e:
            logger.exception(e)

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any

import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ATTENDANCE_HEADERS = (
    "编号",
    "姓名",
    "部門代碼",
    "部門名稱",
    "设备分组",
    "设备名称",
    "名单类型",
    "体温",
    "抓拍时间",
    "记录时间",
)
SHIFT_SHEET_NAME = "辦公室人員及隨線人員班別明細1"
OUTPUT_HEADERS = (
    "通行记录ID",
    "编号",
    "姓名",
    "部門代碼",
    "部門名稱",
    "设备分组",
    "设备名称",
    "名单类型",
    "体温",
    "抓拍时间",
    "记录时间",
    "差異時間",
    "備註",
    "備註2",
)


class AttendanceValidationError(ValueError):
    """出勤文件无法安全处理时抛出的校验错误。"""


@dataclass(frozen=True)
class ShiftSchedule:
    shift: str
    lunch_start: str
    lunch_end: str
    dinner_start: str
    dinner_end: str


@dataclass
class AttendanceRecord:
    source_row: int
    sheet_name: str
    original_data: list[Any]
    emp_id: str
    name: str
    dept_code: str
    device_name: str
    capture_time: datetime | None
    record_time: datetime
    direction: str

    @property
    def effective_time(self) -> datetime:
        return self.capture_time or self.record_time

    @property
    def key(self) -> tuple[str, int]:
        return self.sheet_name, self.source_row


@dataclass(frozen=True)
class LeaveResult:
    duration: int
    status_text: str


@dataclass(frozen=True)
class AttendanceOutputRow:
    key: str
    values: tuple[Any, ...]
    display_values: tuple[str, ...]
    status_text: str
    anomaly_text: str
    flags: tuple[str, ...]
    tone: str
    attention: bool


@dataclass(frozen=True)
class AttendanceOutputSheet:
    name: str
    rows: tuple[AttendanceOutputRow, ...]


@dataclass(frozen=True)
class AttendanceSummary:
    total_records: int
    employee_count: int
    sheet_count: int
    leave_event_count: int
    attention_record_count: int
    overtime_leave_count: int
    meal_overtime_count: int
    capture_time_anomaly_count: int
    missing_entry_count: int


@dataclass(frozen=True)
class AttendanceAnalysis:
    summary: AttendanceSummary
    sheets: tuple[AttendanceOutputSheet, ...]


class AttendanceService:
    max_leave_minutes = 30
    debounce_seconds = 60
    meal_buffer_minutes = 5
    max_pair_minutes = 480

    def process(
        self,
        attendance_content: bytes,
        attendance_suffix: str,
        shift_content: bytes,
    ) -> bytes:
        analysis = self.analyze(
            attendance_content,
            attendance_suffix,
            shift_content,
        )
        return self.export(analysis)

    def analyze(
        self,
        attendance_content: bytes,
        attendance_suffix: str,
        shift_content: bytes,
    ) -> AttendanceAnalysis:
        shifts = self._load_shifts(shift_content)
        sheet_order, records = self._load_attendance(
            attendance_content, attendance_suffix
        )

        missing_shift_ids = sorted(
            {record.emp_id for record in records} - shifts.keys()
        )
        if missing_shift_ids:
            raise AttendanceValidationError(
                "班别表缺少以下员工的有效餐时：" + "、".join(missing_shift_ids)
            )

        grouped_records: dict[str, list[AttendanceRecord]] = defaultdict(list)
        for record in records:
            grouped_records[record.emp_id].append(record)

        leave_results: dict[tuple[str, int], LeaveResult] = {}
        missing_records: set[tuple[str, int]] = set()

        for emp_id, employee_records in grouped_records.items():
            employee_records.sort(key=lambda item: item.effective_time)
            deduplicated = self._deduplicate_records(employee_records)
            leave_results.update(self._analyze_employee(deduplicated, shifts[emp_id]))
            missing_records.update(self._detect_missing_entries(deduplicated))

        sheets = self._build_output_sheets(
            sheet_order,
            records,
            leave_results,
            missing_records,
        )
        output_rows = [row for sheet in sheets for row in sheet.rows]
        return AttendanceAnalysis(
            summary=AttendanceSummary(
                total_records=len(output_rows),
                employee_count=len(grouped_records),
                sheet_count=len(sheets),
                leave_event_count=sum(bool(row.status_text) for row in output_rows),
                attention_record_count=sum(row.attention for row in output_rows),
                overtime_leave_count=sum(
                    row.status_text == "超时离岗" for row in output_rows
                ),
                meal_overtime_count=sum(
                    row.status_text in {"午餐超时", "晚餐超时"} for row in output_rows
                ),
                capture_time_anomaly_count=sum(
                    "time_anomaly" in row.flags for row in output_rows
                ),
                missing_entry_count=sum(
                    "missing_entry" in row.flags for row in output_rows
                ),
            ),
            sheets=tuple(sheets),
        )

    def export(self, analysis: AttendanceAnalysis) -> bytes:
        return self._export(analysis.sheets)

    def _load_shifts(self, content: bytes) -> dict[str, ShiftSchedule]:
        try:
            workbook = load_workbook(
                BytesIO(content),
                read_only=True,
                data_only=True,
            )
        except Exception as exc:
            raise AttendanceValidationError(
                "班别文件无法读取，请确认文件未损坏且格式为 .xlsx"
            ) from exc

        try:
            if SHIFT_SHEET_NAME not in workbook.sheetnames:
                raise AttendanceValidationError(
                    f"班别文件缺少工作表“{SHIFT_SHEET_NAME}”"
                )

            worksheet = workbook[SHIFT_SHEET_NAME]
            headers = [
                self._normalize_header(value)
                for value in next(
                    worksheet.iter_rows(
                        min_row=2,
                        max_row=2,
                        values_only=True,
                    ),
                    (),
                )
            ]
            required_headers = {
                1: "工號",
                13: "班別",
                18: "中餐（午休）時間",
                19: "晚餐時間",
            }
            for index, expected in required_headers.items():
                actual = headers[index] if index < len(headers) else ""
                if actual != expected:
                    raise AttendanceValidationError(
                        f"班别文件第 {index + 1} 列应为“{expected}”，实际为“{actual or '空'}”"
                    )

            shifts: dict[str, ShiftSchedule] = {}
            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=3, values_only=True),
                start=3,
            ):
                emp_id = self._normalize_employee_id(row[1] if len(row) > 1 else None)
                if not emp_id:
                    continue

                shift = self._string_value(row[13] if len(row) > 13 else None)
                if not shift:
                    raise AttendanceValidationError(
                        f"班别文件第 {row_number} 行员工 {emp_id} 缺少班别"
                    )

                lunch_start, lunch_end = self._parse_time_range(
                    row[18] if len(row) > 18 else None,
                    row_number,
                    emp_id,
                    "中餐（午休）時間",
                )
                dinner_start, dinner_end = self._parse_time_range(
                    row[19] if len(row) > 19 else None,
                    row_number,
                    emp_id,
                    "晚餐時間",
                )
                shifts[emp_id] = ShiftSchedule(
                    shift=shift,
                    lunch_start=lunch_start,
                    lunch_end=lunch_end,
                    dinner_start=dinner_start,
                    dinner_end=dinner_end,
                )

            if not shifts:
                raise AttendanceValidationError("班别文件中没有有效员工记录")
            return shifts
        finally:
            workbook.close()

    def _load_attendance(
        self,
        content: bytes,
        suffix: str,
    ) -> tuple[list[str], list[AttendanceRecord]]:
        if suffix == ".xls":
            return self._load_xls_attendance(content)
        if suffix == ".xlsx":
            return self._load_xlsx_attendance(content)
        raise AttendanceValidationError("通行记录仅支持 .xls 或 .xlsx 格式")

    def _load_xls_attendance(
        self,
        content: bytes,
    ) -> tuple[list[str], list[AttendanceRecord]]:
        try:
            workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
        except Exception as exc:
            raise AttendanceValidationError(
                "通行记录无法读取，请确认文件未损坏且格式与扩展名一致"
            ) from exc

        try:
            sheet_order = workbook.sheet_names()
            records: list[AttendanceRecord] = []
            for worksheet in workbook.sheets():
                if worksheet.nrows < 2:
                    raise AttendanceValidationError(f"工作表“{worksheet.name}”缺少表头")
                self._validate_attendance_headers(
                    worksheet.row_values(1),
                    worksheet.name,
                )
                for row_index in range(2, worksheet.nrows):
                    row = worksheet.row_values(row_index)
                    if len(row) <= 1 or not row[1]:
                        continue
                    records.append(
                        self._record_from_row(
                            row,
                            worksheet.name,
                            row_index + 1,
                            workbook.datemode,
                        )
                    )
        finally:
            workbook.release_resources()

        if not records:
            raise AttendanceValidationError("通行记录中没有有效数据")
        return sheet_order, records

    def _load_xlsx_attendance(
        self,
        content: bytes,
    ) -> tuple[list[str], list[AttendanceRecord]]:
        try:
            workbook = load_workbook(
                BytesIO(content),
                read_only=True,
                data_only=True,
            )
        except Exception as exc:
            raise AttendanceValidationError(
                "通行记录无法读取，请确认文件未损坏且格式与扩展名一致"
            ) from exc

        try:
            sheet_order = list(workbook.sheetnames)
            records: list[AttendanceRecord] = []
            for worksheet in workbook.worksheets:
                header_row = next(
                    worksheet.iter_rows(
                        min_row=2,
                        max_row=2,
                        values_only=True,
                    ),
                    (),
                )
                self._validate_attendance_headers(header_row, worksheet.title)
                for row_number, row in enumerate(
                    worksheet.iter_rows(min_row=3, values_only=True),
                    start=3,
                ):
                    if len(row) <= 1 or not row[1]:
                        continue
                    records.append(
                        self._record_from_row(
                            list(row),
                            worksheet.title,
                            row_number,
                        )
                    )
        finally:
            workbook.close()

        if not records:
            raise AttendanceValidationError("通行记录中没有有效数据")
        return sheet_order, records

    def _record_from_row(
        self,
        row: list[Any] | tuple[Any, ...],
        sheet_name: str,
        row_number: int,
        datemode: int | None = None,
    ) -> AttendanceRecord:
        values = list(row)
        while len(values) < 14:
            values.append("")

        emp_id = self._normalize_employee_id(values[1])
        if not emp_id:
            raise AttendanceValidationError(
                f"工作表“{sheet_name}”第 {row_number} 行缺少员工编号"
            )

        record_time = self._parse_datetime(
            values[10],
            sheet_name,
            row_number,
            "记录时间",
            datemode,
            required=True,
        )
        capture_time = self._parse_datetime(
            values[9],
            sheet_name,
            row_number,
            "抓拍时间",
            datemode,
            required=False,
        )
        device_name = self._string_value(values[6])

        return AttendanceRecord(
            source_row=row_number,
            sheet_name=sheet_name,
            original_data=values,
            emp_id=emp_id,
            name=self._string_value(values[2]),
            dept_code=self._string_value(values[3]),
            device_name=device_name,
            capture_time=capture_time,
            record_time=record_time,
            direction=self._parse_direction(device_name),
        )

    def _validate_attendance_headers(
        self,
        row: list[Any] | tuple[Any, ...],
        sheet_name: str,
    ) -> None:
        headers = [self._normalize_header(value) for value in row]
        actual = tuple(headers[1:11])
        if actual != ATTENDANCE_HEADERS:
            raise AttendanceValidationError(
                f"工作表“{sheet_name}”第 2 行表头不符合通行记录格式"
            )

    def _deduplicate_records(
        self,
        records: list[AttendanceRecord],
    ) -> list[AttendanceRecord]:
        if not records:
            return []

        deduplicated = [records[0]]
        for record in records[1:]:
            previous = deduplicated[-1]
            interval = (record.effective_time - previous.effective_time).total_seconds()
            if (
                record.direction
                and record.direction == previous.direction
                and interval < self.debounce_seconds
            ):
                deduplicated[-1] = record
            else:
                deduplicated.append(record)
        return deduplicated

    def _analyze_employee(
        self,
        records: list[AttendanceRecord],
        shift: ShiftSchedule,
    ) -> dict[tuple[str, int], LeaveResult]:
        results: dict[tuple[str, int], LeaveResult] = {}
        index = 0

        while index < len(records) - 1:
            if records[index].direction != "出":
                index += 1
                continue

            leave_record = records[index]
            return_index = index + 1
            while return_index < len(records):
                return_record = records[return_index]
                elapsed_minutes = (
                    return_record.effective_time - leave_record.effective_time
                ).total_seconds() / 60
                if elapsed_minutes > self.max_pair_minutes:
                    index += 1
                    break
                if return_record.direction == "进":
                    duration = int(elapsed_minutes)
                    if duration >= 0:
                        results[leave_record.key] = LeaveResult(
                            duration=duration,
                            status_text=self._get_leave_status(
                                leave_record.effective_time,
                                duration,
                                shift,
                            ),
                        )
                    index = return_index + 1
                    break
                return_index += 1
            else:
                index += 1
                continue

        return results

    def _detect_missing_entries(
        self,
        records: list[AttendanceRecord],
    ) -> set[tuple[str, int]]:
        missing: set[tuple[str, int]] = set()
        for current, following in zip(records, records[1:], strict=False):
            if current.direction != "出" or following.direction != "出":
                continue
            interval_minutes = (
                following.effective_time - current.effective_time
            ).total_seconds() / 60
            if self.max_leave_minutes < interval_minutes <= self.max_pair_minutes:
                missing.add(current.key)
        return missing

    def _get_leave_status(
        self,
        leave_time: datetime,
        duration: int,
        shift: ShiftSchedule,
    ) -> str:
        buffer = timedelta(minutes=self.meal_buffer_minutes)
        lunch_start, lunch_end = self._meal_range(
            leave_time.date(),
            shift.lunch_start,
            shift.lunch_end,
        )
        dinner_start, dinner_end = self._meal_range(
            leave_time.date(),
            shift.dinner_start,
            shift.dinner_end,
        )

        if lunch_start - buffer <= leave_time <= lunch_end + buffer:
            lunch_minutes = int((lunch_end - lunch_start).total_seconds() / 60)
            return "午餐超时" if duration > lunch_minutes else "午餐时间"

        if dinner_start - buffer <= leave_time <= dinner_end + buffer:
            dinner_minutes = int((dinner_end - dinner_start).total_seconds() / 60)
            return "晚餐超时" if duration > dinner_minutes else "晚餐时间"

        return "正常离岗" if duration <= self.max_leave_minutes else "超时离岗"

    def _meal_range(
        self,
        day: date,
        start_value: str,
        end_value: str,
    ) -> tuple[datetime, datetime]:
        start = datetime.combine(
            day,
            datetime.strptime(start_value, "%H:%M").time(),
        )
        end = datetime.combine(
            day,
            datetime.strptime(end_value, "%H:%M").time(),
        )
        if end <= start:
            end += timedelta(days=1)
        return start, end

    def _build_output_sheets(
        self,
        sheet_order: list[str],
        records: list[AttendanceRecord],
        leave_results: dict[tuple[str, int], LeaveResult],
        missing_records: set[tuple[str, int]],
    ) -> list[AttendanceOutputSheet]:
        records_by_sheet: dict[str, list[AttendanceRecord]] = defaultdict(list)
        for record in records:
            records_by_sheet[record.sheet_name].append(record)

        output_sheets: list[AttendanceOutputSheet] = []
        for sheet_name in sheet_order:
            sheet_records = records_by_sheet[sheet_name]
            sheet_records.sort(
                key=lambda record: (
                    record.dept_code,
                    record.emp_id,
                    record.effective_time,
                )
            )

            output_rows: list[AttendanceOutputRow] = []
            for record in sheet_records:
                row_data = list(record.original_data[:14])
                while len(row_data) < 14:
                    row_data.append("")

                leave_result = leave_results.get(record.key)
                if record.direction == "出" and leave_result:
                    row_data[11] = self._format_duration(leave_result.duration)
                    row_data[12] = leave_result.status_text

                anomaly_messages: list[str] = []
                flags: list[str] = []
                if record.capture_time:
                    time_difference = abs(
                        (record.record_time - record.capture_time).total_seconds()
                    )
                    if time_difference > 60:
                        flags.append("time_anomaly")
                        anomaly_messages.append(
                            "数据异常（抓拍与记录相差"
                            f"{self._format_time_difference(time_difference)}）"
                        )
                if record.key in missing_records:
                    flags.append("missing_entry")
                    anomaly_messages.append("缺少进入时间数据")
                row_data[13] = "；".join(anomaly_messages)

                status_text = self._string_value(row_data[12])
                if status_text == "超时离岗":
                    flags.append("overtime")
                elif status_text in {"午餐超时", "晚餐超时"}:
                    flags.extend(("overtime", "meal_overtime"))

                if row_data[13]:
                    tone = "warning"
                elif "overtime" in flags:
                    tone = "danger"
                elif status_text in {"正常离岗", "午餐时间", "晚餐时间"}:
                    tone = "success"
                else:
                    tone = "default"

                output_rows.append(
                    AttendanceOutputRow(
                        key=f"{sheet_name}:{record.source_row}",
                        values=tuple(row_data),
                        display_values=tuple(
                            self._display_value(value) for value in row_data
                        ),
                        status_text=status_text,
                        anomaly_text=self._string_value(row_data[13]),
                        flags=tuple(flags),
                        tone=tone,
                        attention=bool(flags),
                    )
                )

            output_sheets.append(
                AttendanceOutputSheet(
                    name=sheet_name,
                    rows=tuple(output_rows),
                )
            )
        return output_sheets

    def _export(
        self,
        sheets: tuple[AttendanceOutputSheet, ...],
    ) -> bytes:
        workbook = Workbook()
        for sheet_index, output_sheet in enumerate(sheets):
            worksheet = workbook.active if sheet_index == 0 else workbook.create_sheet()
            worksheet.title = output_sheet.name
            worksheet.merge_cells("A1:N1")
            worksheet["A1"] = "通行记录"
            worksheet["A1"].font = Font(bold=True, size=12)

            for column, header in enumerate(OUTPUT_HEADERS, start=1):
                cell = worksheet.cell(row=2, column=column, value=header)
                cell.fill = PatternFill(
                    start_color="34495E",
                    end_color="34495E",
                    fill_type="solid",
                )
                cell.font = Font(color="FFFFFF", bold=True)

            for output_row_number, output_row in enumerate(
                output_sheet.rows,
                start=3,
            ):
                for column, value in enumerate(output_row.values, start=1):
                    cell = worksheet.cell(
                        row=output_row_number,
                        column=column,
                        value=value,
                    )
                    if output_row.tone == "success":
                        cell.fill = PatternFill(
                            start_color="C6EFCE",
                            end_color="C6EFCE",
                            fill_type="solid",
                        )
                    elif output_row.tone == "danger":
                        cell.fill = PatternFill(
                            start_color="FFC7CE",
                            end_color="FFC7CE",
                            fill_type="solid",
                        )
                    elif output_row.tone == "warning":
                        cell.fill = PatternFill(
                            start_color="FFEB9C",
                            end_color="FFEB9C",
                            fill_type="solid",
                        )

        sheet_names = {sheet.name for sheet in sheets}
        if "Sheet" in workbook.sheetnames and "Sheet" not in sheet_names:
            del workbook["Sheet"]

        for worksheet in workbook.worksheets:
            for column in range(1, worksheet.max_column + 1):
                width = max(
                    (
                        len(str(cell.value))
                        for cell in next(
                            worksheet.iter_cols(
                                min_col=column,
                                max_col=column,
                            )
                        )
                        if cell.value is not None
                    ),
                    default=0,
                )
                worksheet.column_dimensions[get_column_letter(column)].width = width + 2

        output = BytesIO()
        workbook.save(output)
        workbook.close()
        return output.getvalue()

    def _parse_datetime(
        self,
        value: Any,
        sheet_name: str,
        row_number: int,
        field_name: str,
        datemode: int | None,
        *,
        required: bool,
    ) -> datetime | None:
        if value in (None, ""):
            if required:
                raise AttendanceValidationError(
                    f"工作表“{sheet_name}”第 {row_number} 行缺少{field_name}"
                )
            return None

        try:
            if isinstance(value, datetime):
                return value
            if isinstance(value, (int, float)) and datemode is not None:
                return xlrd.xldate_as_datetime(value, datemode)
            return datetime.fromisoformat(str(value).strip())
        except (TypeError, ValueError, xlrd.XLDateError) as exc:
            raise AttendanceValidationError(
                f"工作表“{sheet_name}”第 {row_number} 行{field_name}格式无效"
            ) from exc

    def _parse_time_range(
        self,
        value: Any,
        row_number: int,
        emp_id: str,
        field_name: str,
    ) -> tuple[str, str]:
        normalized = self._string_value(value).replace("::", ":")
        parts = [part.strip() for part in normalized.split("-")]
        if len(parts) != 2:
            raise AttendanceValidationError(
                f"班别文件第 {row_number} 行员工 {emp_id} 的{field_name}格式无效"
            )

        try:
            start = datetime.strptime(parts[0], "%H:%M").strftime("%H:%M")
            end = datetime.strptime(parts[1], "%H:%M").strftime("%H:%M")
        except ValueError as exc:
            raise AttendanceValidationError(
                f"班别文件第 {row_number} 行员工 {emp_id} 的{field_name}格式无效"
            ) from exc
        return start, end

    def _format_duration(self, total_minutes: int) -> str:
        days, remaining = divmod(total_minutes, 1440)
        hours, minutes = divmod(remaining, 60)
        parts = []
        if days:
            parts.append(f"{days}天")
        if hours:
            parts.append(f"{hours}小时")
        if minutes or not parts:
            parts.append(f"{minutes}分钟")
        return "".join(parts)

    def _format_time_difference(self, total_seconds: float) -> str:
        days, remaining = divmod(int(total_seconds), 86400)
        hours, remaining = divmod(remaining, 3600)
        minutes, seconds = divmod(remaining, 60)
        parts = []
        if days:
            parts.append(f"{days}天")
        if hours:
            parts.append(f"{hours}小时")
        if minutes:
            parts.append(f"{minutes}分钟")
        if seconds or not parts:
            parts.append(f"{seconds}秒")
        return "".join(parts)

    def _normalize_employee_id(self, value: Any) -> str:
        if value in (None, ""):
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        if isinstance(value, int):
            return str(value)
        text = str(value).strip()
        try:
            number = float(text)
            if number.is_integer():
                return str(int(number))
        except ValueError:
            pass
        return text

    def _normalize_header(self, value: Any) -> str:
        return self._string_value(value).replace("_x000c_", "").strip()

    def _display_value(self, value: Any) -> str:
        if value in (None, ""):
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)

    def _string_value(self, value: Any) -> str:
        return "" if value is None else str(value).strip()

    def _parse_direction(self, device_name: str) -> str:
        normalized = device_name.rstrip()
        if normalized.endswith(("進", "进", "入")):
            return "进"
        if normalized.endswith("出"):
            return "出"
        return ""


def validate_upload_extensions(
    attendance_filename: str,
    shift_filename: str,
) -> str:
    attendance_suffix = Path(attendance_filename).suffix.lower()
    shift_suffix = Path(shift_filename).suffix.lower()
    if attendance_suffix not in {".xls", ".xlsx"}:
        raise AttendanceValidationError("通行记录仅支持 .xls 或 .xlsx 格式")
    if shift_suffix != ".xlsx":
        raise AttendanceValidationError("班别文件仅支持 .xlsx 格式")
    return attendance_suffix

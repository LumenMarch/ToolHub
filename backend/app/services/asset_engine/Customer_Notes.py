
class pyqtSignal:
    def __init__(self, *args, **kwargs): pass
    def connect(self, *args, **kwargs): pass
    def emit(self, *args, **kwargs): pass

class QThread:
    pass

class QObject:
    pass

class QWidget:
    pass

def safe_thread_run(func):
    def wrapper(*args, **kwargs):
        return func(*args, **kwargs)
    return wrapper

import os
import re
from datetime import datetime

import openpyxl
import pandas as pd
import polars as pl
from app.services.asset_engine.const import CUSTOMER_NOTES_SAVE_PATH
from loguru import logger
from app.services.asset_engine.mod import safe_thread_run

Notes_RFID_rex = re.compile(r"(A15.*)")
Notes_RFID_rex_2 = re.compile(r"\S.*(A15.*)")
No_CheckRFID = ["A1300011C5C3", "A13000103933", "A1300010E606"]


class Customer_Notes(QThread):
    _unlock_signal = pyqtSignal()
    _update_check_signal = pyqtSignal(int, int, int, int)
    _Error_signal = pyqtSignal()
    _Save_signal = pyqtSignal(object)
    _progress_signal = pyqtSignal(int, str)
    _Update_Message_signal = pyqtSignal(str)

    # 進度信號：進度值, 進度文本
    def __init__(self, ui):
        super().__init__()
        self.ui = ui
        self.this_Customer_path = None
        self.this_Notes_path = None
        self.this_Customer_DRI_path = None

        self.this_Customer_data = None
        self.this_Notes_data = None

        self.this_Customer_assets = None
        self.this_Notes_assets = None
        self.this_Customer_DRI_data = None

        self.new_assets = None
        self.remove_assets = None

    @safe_thread_run
    def run(self):
        if (
            not self.this_Customer_path
            or not self.this_Notes_path
            or not self.this_Customer_DRI_path
        ):
            self._Update_Message_signal.emit("请选择文件")
            return

        self._Update_Message_signal.emit("Start")
        try:
            self._progress_signal.emit(10, "開始讀取客戶數據...")
            self.read_this_Customer_data()
            self._progress_signal.emit(30, "開始讀取Notes數據...")
            self.read_this_Notes_data()
            self._progress_signal.emit(60, "開始讀取DRI數據...")
            self.read_Customer_DRI()
            self._progress_signal.emit(80, "開始進行數據對比...")
            self.Customer_Notes_Comparison()
            self._progress_signal.emit(100, "對比完成")
            self._unlock_signal.emit()
        except Exception as e:
            logger.exception(f"读取数据失败: {e}")
            self._unlock_signal.emit()
            self._Error_signal.emit()
            raise

    def safe_read_excel(self, path):
        """安全讀取Excel並轉為LazyFrame"""
        required_columns = {"資產編號", "RFID", "DRI"}
        try:
            # 嘗試使用 Polars 讀取 Excel
            # 注意：设置 infer_schema_length=0 让Polars将所有列读取为字符串，避免类型推断错误
            df_polars = pl.read_excel(path, infer_schema_length=0)

            # 檢查是否包含必需的列
            header_columns = set(df_polars.columns)
            if required_columns.intersection(header_columns):
                # 自动去除资产编号列末尾的点号
                if "資產編號" in df_polars.columns:
                    df_polars = df_polars.with_columns(
                        [pl.col("資產編號").str.strip_chars_end(".").alias("資產編號")]
                    )
                # 移除全空列並返回 DataFrame
                non_null_cols = [
                    col
                    for col in df_polars.columns
                    if not df_polars[col].is_null().all()
                ]
                return df_polars.select(non_null_cols)

            # 如果表頭不正確，嘗試查找正確的表頭行
            for i in range(min(5, len(df_polars))):
                row_values = df_polars.row(i)
                clean_columns = {
                    str(col).strip() for col in row_values if col is not None
                }
                if required_columns.intersection(clean_columns):
                    # 使用找到的行作為表頭
                    new_columns = [
                        str(col).strip() if col is not None else f"col_{j}"
                        for j, col in enumerate(row_values)
                    ]
                    df_polars = df_polars.slice(i + 1).rename(
                        {old: new for old, new in zip(df_polars.columns, new_columns)}
                    )
                    # 自动去除资产编号列末尾的点号
                    if "資產編號" in df_polars.columns:
                        df_polars = df_polars.with_columns(
                            [
                                pl.col("資產編號")
                                .cast(pl.String)
                                .str.strip_chars_end(".")
                                .alias("資產編號")
                            ]
                        )
                    non_null_cols = [
                        col
                        for col in df_polars.columns
                        if not df_polars[col].is_null().all()
                    ]
                    return df_polars.select(non_null_cols)

            # 如果仍然找不到，拋出錯誤
            raise ValueError(f"未在文件 {os.path.basename(path)} 中找到完整表頭")

        except Exception as e:
            logger.exception(f"读取 {os.path.basename(path)} 失败: {e}")
            self._unlock_signal.emit()
            self._Error_signal.emit()
            raise

    def read_this_Customer_data(self):
        """讀取客戶數據（使用LazyFrame）"""
        try:
            self.this_Customer_data = None
            if not self.this_Customer_path:
                return

            df_lazy = self.safe_read_excel(self.this_Customer_path)

            # 處理 RFID 列的前導零問題
            if "RFID" in df_lazy.columns:
                df_lazy = df_lazy.with_columns(
                    [
                        pl.when(
                            pl.col("RFID")
                            .cast(pl.String)
                            .str.starts_with("000000000000")
                        )
                        .then(pl.col("RFID").cast(pl.String).str.strip_chars_start("0"))
                        .otherwise(pl.col("RFID").cast(pl.String))
                        .alias("RFID")
                    ]
                )

            self.this_Customer_data = df_lazy

        except Exception as e:
            logger.exception(
                f"读取 {os.path.basename(self.this_Customer_path)} 失败: {e}"
            )
            self._unlock_signal.emit()
            self._Error_signal.emit()
            return

    def read_this_Notes_data(self):
        """讀取Notes數據（使用LazyFrame）"""
        try:
            self.this_Notes_data = None
            if not self.this_Notes_path:
                return

            df_lazy = self.safe_read_excel(self.this_Notes_path)

            # 處理 RFID（Tag） 列的前導零問題
            if "RFID（Tag）" in df_lazy.columns:
                df_lazy = df_lazy.with_columns(
                    [
                        pl.when(
                            pl.col("RFID（Tag）")
                            .cast(pl.String)
                            .str.starts_with("000000000000")
                        )
                        .then(
                            pl.col("RFID（Tag）")
                            .cast(pl.String)
                            .str.strip_chars_start("0")
                        )
                        .otherwise(pl.col("RFID（Tag）").cast(pl.String))
                        .alias("RFID（Tag）")
                    ]
                )

                # 過濾掉指定的RFID，不參與對比
                df_lazy = df_lazy.filter(
                    ~pl.col("RFID（Tag）").cast(pl.String).is_in(No_CheckRFID)
                )

            self.this_Notes_data = df_lazy

        except Exception as e:
            logger.exception(f"读取 {os.path.basename(self.this_Notes_path)} 失败: {e}")
            self._unlock_signal.emit()
            self._Error_signal.emit()
            return

    def read_Customer_DRI(self):
        """讀取客戶DRI數據"""
        self.this_Customer_DRI_data = None
        if not self.this_Customer_DRI_path:
            return
        with open(self.this_Customer_DRI_path, "r", encoding="utf-8") as file:
            lines = file.readlines()
            self.this_Customer_DRI_data = [line.strip() for line in lines]

    def Customer_Notes_Comparison(self):
        """Customer與Notes數據比較（使用LazyFrame優化）"""
        if self.this_Customer_data is None or self.this_Notes_data is None:
            self._unlock_signal.emit()
            return
        try:
            # 初始化Notes資產列表
            self.this_Notes_assets = []

            # 1. 從RFID列提取符合A1開頭的RFID
            notes_rfid_query = self.this_Notes_data.filter(
                pl.col("RFID（Tag）").is_not_null()
                & pl.col("RFID（Tag）").cast(pl.String).str.contains(r"^A1.*")
            ).select("RFID（Tag）")

            rfid_assets = notes_rfid_query.to_series().to_list()
            self.this_Notes_assets.extend(rfid_assets)

            # 2. 從備注説明列中提取符合正則表達式的RFID（優先級較低）
            self._extract_rfid_from_remarks()

            # 使用 LazyFrame 操作，過濾出符合 DRI 條件的客戶資產
            customer_rfid_query = self.this_Customer_data.filter(
                pl.col("DRI").is_in(self.this_Customer_DRI_data)
            ).select("RFID")

            self.this_Customer_assets = customer_rfid_query.to_series().to_list()

            self.new_assets = set(self.this_Notes_assets) - set(
                self.this_Customer_assets
            )
            self.remove_assets = set(self.this_Customer_assets) - set(
                self.this_Notes_assets
            )

            if self.new_assets or self.remove_assets:
                self._update_check_signal.emit(
                    len(self.this_Notes_assets),
                    len(self.this_Customer_assets),
                    len(self.new_assets),
                    len(self.remove_assets),
                )
            else:
                self._update_check_signal.emit(
                    len(self.this_Notes_assets), len(self.this_Customer_assets), 0, 0
                )

        except Exception as e:
            logger.exception(f"Customer_Notes-读取数据对比失败: {e}")
            self._unlock_signal.emit()
            raise

    def _extract_rfid_from_remarks(self):
        """從備注説明列中提取符合正則表達式的RFID（優先級較低，如果RFID列已有則跳過）"""
        try:
            if (
                self.this_Notes_data is None
                or "備注説明" not in self.this_Notes_data.columns
            ):
                return

            # 獲取備注説明列中符合正則表達式的數據
            remarks_data = self.this_Notes_data.filter(
                pl.col("備注説明").is_not_null()
                & (
                    pl.col("備注説明")
                    .cast(pl.String)
                    .str.contains(Notes_RFID_rex.pattern)
                    | pl.col("備注説明")
                    .cast(pl.String)
                    .str.contains(Notes_RFID_rex_2.pattern)
                )
            ).select("備注説明")

            extracted_rfids = []
            for remark in remarks_data.to_series():
                if remark:
                    # 使用第一個正則表達式提取
                    match1 = Notes_RFID_rex.search(str(remark))
                    if match1:
                        extracted_rfids.append(match1.group(1))
                        continue

                    # 使用第二個正則表達式提取
                    match2 = Notes_RFID_rex_2.search(str(remark))
                    if match2:
                        extracted_rfids.append(match2.group(1))

            # 將提取的RFID添加到Notes資產列表中（避免重複，優先使用RFID列的數據）
            if extracted_rfids:
                existing_rfids = set(self.this_Notes_assets)
                for rfid in extracted_rfids:
                    # 如果RFID列中已經存在該RFID，則跳過備注列中的相同RFID
                    if rfid not in existing_rfids:
                        self.this_Notes_assets.append(rfid)

        except Exception as e:
            logger.exception(f"從備注説明列提取RFID失敗: {e}")
            raise

    def Save_Customer_Notes_Comparison(self):
        """保存Customer與Notes比較結果為Excel文件"""
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            if not self.new_assets and not self.remove_assets:
                return
            else:
                # 創建Excel寫入器
                excel_path = CUSTOMER_NOTES_SAVE_PATH
                with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
                    # 構建新增資產的DataFrame
                    new_assets_list = list(self.new_assets)
                    new_df_rows = []

                    for rfid in new_assets_list:
                        # 查找該RFID在Notes數據中的記錄
                        matching_rows = self.this_Notes_data.filter(
                            pl.col("RFID（Tag）") == rfid
                        )

                        if not matching_rows.is_empty():
                            # 如果RFID在RFID列中找到
                            row = matching_rows.select(
                                [
                                    "資產名稱",
                                    "資產編號",
                                    "RFID（Tag）",
                                    "保管人",
                                    "備注説明",
                                ]
                            ).row(0)
                            new_df_rows.append(
                                {
                                    "資產名稱": row[0] if len(row) > 0 else "",
                                    "資產編號": row[1] if len(row) > 1 else "",
                                    "RFID（Tag）": row[2] if len(row) > 2 else "",
                                    "保管人": row[3] if len(row) > 3 else "",
                                    "備注説明": row[4] if len(row) > 4 else "",
                                }
                            )
                        else:
                            # 如果RFID是從備注説明列提取的，創建一個特殊記錄
                            new_df_rows.append(
                                {
                                    "資產名稱": "從備注提取",
                                    "資產編號": "",
                                    "RFID（Tag）": rfid,
                                    "保管人": "",
                                    "備注説明": f"從備注説明列提取的RFID: {rfid}",
                                }
                            )

                    # 創建新增資產DataFrame
                    if new_df_rows:
                        new_df = pd.DataFrame(new_df_rows)
                    else:
                        new_df = pd.DataFrame()

                    # 構建減少資產的DataFrame
                    remove_df = self.this_Customer_data.filter(
                        pl.col("RFID").is_in(list(self.remove_assets))
                    ).select(["Model Number", "Serial Number", "RFID", "DRI"])

                    remove_df_pandas = remove_df.to_pandas()

                    # 創建一個空的DataFrame來初始化工作表
                    empty_df = pd.DataFrame()
                    empty_df.to_excel(writer, sheet_name="對比結果", index=False)

                    # 獲取工作表
                    worksheet = writer.sheets["對比結果"]

                    # 寫入標題信息
                    worksheet.merge_cells("A1:F1")
                    worksheet["A1"] = (
                        f"本月Notes客户资产_VS_本月系统客户资产 (对比时间{current_time})"
                    )
                    worksheet["A1"].font = openpyxl.styles.Font(bold=True, size=14)
                    worksheet["A1"].alignment = openpyxl.styles.Alignment(
                        horizontal="center"
                    )
                    # 合并并居中 A2:D2 和 A3:D3
                    worksheet.merge_cells("A2:D2")
                    worksheet.merge_cells("A3:D3")
                    worksheet["A2"].alignment = openpyxl.styles.Alignment(
                        horizontal="center"
                    )
                    worksheet["A3"].alignment = openpyxl.styles.Alignment(
                        horizontal="center"
                    )

                    # 寫入新增資產數據
                    if not new_df.empty:
                        # 新增資產標題
                        worksheet["A3"] = (
                            f"本月Notes比系统新增资产 {len(self.new_assets)}笔"
                        )
                        worksheet["A3"].font = openpyxl.styles.Font(bold=True, size=12)

                        # 添加列標題
                        headers = [
                            "No.",
                            "資產名稱",
                            "資產編號",
                            "RFID（Tag）",
                            "保管人",
                            "備注説明",
                        ]
                        for i, header in enumerate(headers, 1):
                            cell = worksheet.cell(row=4, column=i, value=header)
                            cell.font = openpyxl.styles.Font(bold=True)
                            cell.fill = openpyxl.styles.PatternFill(
                                start_color="E6F3FF",
                                end_color="E6F3FF",
                                fill_type="solid",
                            )

                        # 寫入數據
                        for i in range(len(new_df)):
                            worksheet.cell(row=i + 5, column=1, value=i + 1)  # No.
                            worksheet.cell(
                                row=i + 5, column=2, value=new_df.iloc[i]["資產名稱"]
                            )
                            worksheet.cell(
                                row=i + 5, column=3, value=new_df.iloc[i]["資產編號"]
                            )
                            worksheet.cell(
                                row=i + 5, column=4, value=new_df.iloc[i]["RFID（Tag）"]
                            )
                            worksheet.cell(
                                row=i + 5, column=5, value=new_df.iloc[i]["保管人"]
                            )
                            worksheet.cell(
                                row=i + 5, column=6, value=new_df.iloc[i]["備注説明"]
                            )

                    # 寫入減少資產數據
                    if not remove_df_pandas.empty:
                        start_row = len(new_df) + 7 if not new_df.empty else 5

                        # 減少資產標題
                        worksheet[f"A{start_row}"] = (
                            f"本月Notes比系统减少资产 {len(self.remove_assets)}笔"
                        )
                        worksheet[f"A{start_row}"].font = openpyxl.styles.Font(
                            bold=True, size=12
                        )

                        # 添加列標題
                        headers = [
                            "No.",
                            "Model Number",
                            "Serial Number",
                            "RFID",
                            "DRI",
                        ]
                        for i, header in enumerate(headers, 1):
                            cell = worksheet.cell(
                                row=start_row + 1, column=i, value=header
                            )
                            cell.font = openpyxl.styles.Font(bold=True)
                            cell.fill = openpyxl.styles.PatternFill(
                                start_color="FFE6E6",
                                end_color="FFE6E6",
                                fill_type="solid",
                            )

                        # 寫入數據
                        for i in range(len(remove_df_pandas)):
                            worksheet.cell(
                                row=start_row + 2 + i, column=1, value=i + 1
                            )  # No.
                            worksheet.cell(
                                row=start_row + 2 + i,
                                column=2,
                                value=remove_df_pandas.iloc[i]["Model Number"],
                            )
                            worksheet.cell(
                                row=start_row + 2 + i,
                                column=3,
                                value=remove_df_pandas.iloc[i]["Serial Number"],
                            )
                            worksheet.cell(
                                row=start_row + 2 + i,
                                column=4,
                                value=remove_df_pandas.iloc[i]["RFID"],
                            )
                            worksheet.cell(
                                row=start_row + 2 + i,
                                column=5,
                                value=remove_df_pandas.iloc[i]["DRI"],
                            )

                    # 設置列寬和邊框
                    from openpyxl.styles import Border, Side

                    thin_border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"),
                    )

                    # 設置列寬
                    column_widths = [8, 30, 20, 25, 15, 40]  # 根據內容調整
                    for i, width in enumerate(column_widths, 1):
                        worksheet.column_dimensions[
                            openpyxl.utils.get_column_letter(i)
                        ].width = width

                    # 添加邊框
                    for row in worksheet.iter_rows(
                        min_row=1,
                        max_row=worksheet.max_row,
                        min_col=1,
                        max_col=worksheet.max_column,
                    ):
                        for cell in row:
                            cell.border = thin_border

                self._Save_signal.emit(self.ui.frame_7)

        except Exception as e:
            logger.exception(f"保存数据对比失败: {e}")
            self._unlock_signal.emit()
            raise

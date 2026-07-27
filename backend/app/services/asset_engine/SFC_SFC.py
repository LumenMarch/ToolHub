class pyqtSignal:
    def __init__(self, *args, **kwargs):
        pass

    def connect(self, *args, **kwargs):
        pass

    def emit(self, *args, **kwargs):
        pass


class QThread:
    pass


class QObject:
    pass


class QWidget:
    pass


def safe_thread_run(func):  # noqa: F811
    def wrapper(*args, **kwargs):
        return func(*args, **kwargs)

    return wrapper


import os  # noqa: E402, I001, UP015, F401
from datetime import datetime  # noqa: E402, I001, UP015, F401

import openpyxl  # noqa: E402, I001, UP015, F401
import pandas as pd  # noqa: E402, I001, UP015, F401
from app.services.asset_engine.const import SFC_SFC_SAVE_PATH  # noqa: E402, I001, UP015, F401
from loguru import logger  # noqa: E402, I001, UP015, F401
from app.services.asset_engine.TableParser import (  # noqa: E402
    TableParser,
)  # 使用自定義的 HTML 表格解析器  # noqa: E402, I001, UP015, F401


class SFC_SFC(QThread):
    _SFC_SFC_Comparison = pyqtSignal(int, int, int, int)
    _unlock_signal = pyqtSignal()
    _Error_signal = pyqtSignal()
    _Save_signal = pyqtSignal(object)
    _progress_signal = pyqtSignal(int, str)  # 進度信號：進度值, 進度文本
    _Update_Message_signal = pyqtSignal(str)

    def __init__(self, ui):
        super().__init__()
        self.ui = ui
        self.removed_count = None
        self.new_count = None
        self.removed_assets = None
        self.new_assets = None
        self.this_SFC_data = None
        self.last_SFC_data = None

        self.This_data_Path = None
        self.Last_data_path = None

    @safe_thread_run
    def run(self):
        if not self.This_data_Path or not self.Last_data_path:
            self._Update_Message_signal.emit("请选择文件")
            return

        self._Update_Message_signal.emit("Start")
        try:
            self._progress_signal.emit(10, "读取本月SFC数据...")
            self.This_SFC_date()
            self._progress_signal.emit(30, "读取上月SFC数据...")
            self.Last_SFC_date()
            self._progress_signal.emit(60, "开始进行数据对比...")
            self.SFC_SFC_Comparison()
            self._progress_signal.emit(100, "对比完成")
            self._unlock_signal.emit()
        except Exception as e:
            logger.exception(e)
            self._Error_signal.emit()
            self._unlock_signal.emit()

    def safe_read_excel(self, path):
        """安全讀取Excel或HTML偽裝的Excel，並返回 pandas DataFrame"""
        required_columns = {"设备编号", "资产编号", "資產編號"}

        # 智能檢測文件格式並相應處理
        def is_html_file(file_path):
            """檢測文件是否為HTML格式"""
            try:
                with open(file_path, encoding="utf-8", errors="ignore") as f:
                    first_line = f.readline().strip().lower()
                    return (
                        first_line.startswith("<")
                        or "html" in first_line
                        or "table" in first_line
                    )
            except Exception:
                return False

        # 首先檢測文件格式
        if is_html_file(path):
            logger.info(f"檢測到HTML格式文件，使用HTML解析: {os.path.basename(path)}")
            # 使用 HTML 解析方式
            try:
                with open(path, encoding="utf-8", errors="ignore") as f:
                    html_content = f.read()

                # 使用自定義的 TableParser 解析 HTML
                parser = TableParser()
                parser.feed(html_content)

                if not parser.tables:
                    raise ValueError(
                        f"未在 HTML 文件中找到表格: {os.path.basename(path)}"
                    )

                # 使用第一個表格
                table_data = parser.tables[0]
                if not table_data:
                    raise ValueError(f"表格為空: {os.path.basename(path)}")

                # 查找表頭行
                header_row = None
                for idx, row in enumerate(table_data):
                    # 處理空值：將None轉換為空字符串，然後清理
                    clean_columns = {
                        str(col).strip() for col in row if col and str(col).strip()
                    }
                    if required_columns.intersection(clean_columns):
                        header_row = idx
                        break

                if header_row is not None:
                    # 使用找到的行作為表頭
                    headers = table_data[header_row]
                    data_rows = table_data[header_row + 1 :]
                else:
                    # 默認使用第一行為表頭
                    logger.warning(
                        f"HTML中未找到完整表頭，默認使用第一行: {os.path.basename(path)}"
                    )
                    headers = table_data[0] if table_data else []
                    data_rows = table_data[1:] if len(table_data) > 1 else []

                # 確保所有行都有相同的列數
                max_cols = len(headers) if headers else 0
                if max_cols == 0:
                    raise ValueError(f"無法確定表格結構: {os.path.basename(path)}")

                # 補齊缺失的列並處理空值
                normalized_data = []
                for row in data_rows:
                    # 處理每一行的空值：將None轉換為空字符串
                    cleaned_row = [
                        str(cell).strip() if cell is not None else "" for cell in row
                    ]
                    cleaned_row + [""] * (max_cols - len(cleaned_row))
                    normalized_data.append(cleaned_row[:max_cols])

                # 創建 pandas DataFrame
                if normalized_data:
                    df_dict = {}
                    for i, header in enumerate(headers):
                        # 處理表頭的空值
                        column_name = (
                            str(header).strip()
                            if header is not None and str(header).strip()
                            else f"col_{i}"
                        )
                        # 處理數據列的空值
                        column_data = []
                        for row in normalized_data:
                            if i < len(row):
                                cell_value = row[i]
                                # 確保空值統一處理為空字符串
                                column_data.append(
                                    cell_value
                                    if cell_value and str(cell_value).strip()
                                    else ""
                                )
                            else:
                                column_data.append("")
                        df_dict[column_name] = column_data

                    df_pandas = pd.DataFrame(df_dict)
                else:
                    # 如果沒有數據行，創建空的 DataFrame
                    df_dict = {}
                    for i, header in enumerate(headers):
                        column_name = (
                            str(header).strip()
                            if header is not None and str(header).strip()
                            else f"col_{i}"
                        )
                        df_dict[column_name] = []
                    df_pandas = pd.DataFrame(df_dict)

                logger.info(f"HTML解析成功，原始数据行数: {len(df_pandas)}")

            except Exception as html_error:
                logger.exception(f"HTML解析失敗: {html_error}")
                self._unlock_signal.emit()
                self._Error_signal.emit()
                return None

        else:
            logger.info(
                f"檢測到Excel格式文件，使用pandas讀取: {os.path.basename(path)}"
            )
            # 使用 pandas 讀取真正的 Excel 文件
            try:
                df_pandas = pd.read_excel(path, engine="openpyxl")
                logger.info(f"Excel文件讀取成功，原始数据行数: {len(df_pandas)}")

            except Exception as excel_error:
                logger.exception(f"pandas讀取Excel失敗: {excel_error}")
                self._unlock_signal.emit()
                self._Error_signal.emit()
                return None

        # 統一的数据清理和过滤逻辑
        try:
            # 檢查是否包含必需的列
            df_columns = set(df_pandas.columns.astype(str))
            if not required_columns.intersection(df_columns):
                logger.warning(
                    f"文件中未找到必需的列，嘗試查找包含關鍵字的列: {os.path.basename(path)}"
                )
                # 嘗試查找包含關鍵字的列
                found_columns = []
                for col in df_pandas.columns:
                    col_str = str(col).lower()
                    if any(
                        keyword in col_str
                        for keyword in ["设备编号", "资产编号", "資產編號", "設備編號"]
                    ):
                        found_columns.append(col)

                if not found_columns:
                    logger.warning(
                        f"未找到資產編號相關列，但繼續處理: {os.path.basename(path)}"
                    )
                else:
                    logger.info(f"找到相關列: {found_columns}")

            # 🎯 加强空数据过滤
            # 自动去除资产编号列末尾的点号
            if "资产编号" in df_pandas.columns:
                df_pandas["资产编号"] = (
                    df_pandas["资产编号"].astype(str).str.rstrip(".")
                )
            if "資產編號" in df_pandas.columns:
                df_pandas["資產編號"] = (
                    df_pandas["資產編號"].astype(str).str.rstrip(".")
                )

            # 将所有列转换为字符串类型以避免数据类型推断错误
            df_pandas = df_pandas.astype(str)

            # 🎯 过滤完全空白的行
            # 检查每一行是否所有列都是空值或只包含空格
            def is_empty_row(row):
                return all(str(cell).strip() == "" for cell in row)

            # 过滤掉完全空白的行
            df_pandas = df_pandas[~df_pandas.apply(is_empty_row, axis=1)]

            # 🎯 过滤资产编号为空的行
            asset_cols = ["资产编号", "資產編號"]
            for col in asset_cols:
                if col in df_pandas.columns:
                    # 过滤掉资产编号为空、NaN、None的行
                    df_pandas = df_pandas[
                        (df_pandas[col].astype(str).str.strip() != "")
                        & (df_pandas[col].astype(str).str.lower() != "nan")
                        & (df_pandas[col].astype(str).str.lower() != "none")
                        & (df_pandas[col].astype(str).str.lower() != "null")
                    ]

            logger.info(f"过滤后数据行数: {len(df_pandas)}")
            if len(df_pandas) > 0:
                logger.info(f"数据样本: {df_pandas.head(3).to_dict('records')}")

            return df_pandas

        except Exception as e:
            logger.exception(f"数据清理和过滤失败: {e}")
            self._unlock_signal.emit()
            self._Error_signal.emit()
            return None

    def This_SFC_date(self):
        """讀取本期SFC數據（使用 pandas DataFrame）"""
        self.this_SFC_data = None
        try:
            self.this_SFC_data = self.safe_read_excel(self.This_data_Path)
        except Exception as e:
            self._unlock_signal.emit()
            self._Error_signal.emit()
            logger.exception(e)

    def Last_SFC_date(self):
        """讀取上期SFC數據（使用 pandas DataFrame）"""
        self.last_SFC_data = None
        try:
            self.last_SFC_data = self.safe_read_excel(self.Last_data_path)
        except Exception as e:
            logger.error(e)

    def _get_comparison_keys(self, df):
        """获取用于对比的唯一标识键。
        资产编号有效时用资产编号，为 NA 时降级使用设备编号。
        返回 (keys_set, asset_col, device_col)"""
        if df is None or df.empty:
            return set(), None, None

        asset_col = next((c for c in ["资产编号", "資產編號"] if c in df.columns), None)
        device_col = next(
            (c for c in ["设备编号", "設備編號"] if c in df.columns), None
        )

        invalid_values = {"nan", "none", "null", "na", ""}
        keys = set()
        for _, row in df.iterrows():
            asset_val = str(row.get(asset_col, "")).strip() if asset_col else ""
            device_val = str(row.get(device_col, "")).strip() if device_col else ""

            if asset_val and asset_val.lower() not in invalid_values:
                keys.add(f"A:{asset_val}")
            elif device_val and device_val.lower() not in invalid_values:
                keys.add(f"D:{device_val}")

        return keys, asset_col, device_col

    def SFC_SFC_Comparison(self):
        """SFC數據比較（使用 pandas DataFrame）"""
        self.new_assets = None
        self.removed_assets = None
        if self.last_SFC_data is None or self.this_SFC_data is None:
            return
        else:
            try:
                # 获取对比标识键（资产编号降级设备编号）
                this_keys, self._asset_col, self._device_col = (
                    self._get_comparison_keys(self.this_SFC_data)
                )
                last_keys, _, _ = self._get_comparison_keys(self.last_SFC_data)

                logger.info(f"本月有效资产数量: {len(this_keys)}")
                logger.info(f"上月有效资产数量: {len(last_keys)}")
                if this_keys:
                    logger.info(f"本月资产样本: {list(this_keys)[:3]}")
                if last_keys:
                    logger.info(f"上月资产样本: {list(last_keys)[:3]}")

                self.new_assets = this_keys - last_keys
                self.removed_assets = last_keys - this_keys

                new_count = len(self.new_assets)
                removed_count = len(self.removed_assets)

                logger.info(f"新增资产数量: {new_count}")
                logger.info(f"减少资产数量: {removed_count}")

                if new_count > 0 or removed_count > 0:
                    self._SFC_SFC_Comparison.emit(
                        len(this_keys), len(last_keys), new_count, removed_count
                    )
                else:
                    self._SFC_SFC_Comparison.emit(len(this_keys), len(last_keys), 0, 0)
            except Exception as e:
                self._unlock_signal.emit()
                logger.exception(e)

    def Save_SFC_SFC_Comparison(self):
        """保存SFC比較結果為Excel文件"""
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            if not self.new_assets and not self.removed_assets:
                return
            else:
                # 检测列名
                asset_col = getattr(self, "_asset_col", None)
                device_col = getattr(self, "_device_col", None)
                if not asset_col:
                    for c in ["资产编号", "資產編號"]:
                        if c in self.this_SFC_data.columns:
                            asset_col = c
                            break
                if not device_col:
                    for c in ["设备编号", "設備編號"]:
                        if c in self.this_SFC_data.columns:
                            device_col = c
                            break

                # 分离 A:/D: 前缀标识
                def split_keys(keys):
                    by_asset = {k[2:] for k in keys if k.startswith("A:")}
                    by_device = {k[2:] for k in keys if k.startswith("D:")}
                    return by_asset, by_device

                new_by_asset, new_by_device = split_keys(self.new_assets)
                removed_by_asset, removed_by_device = split_keys(self.removed_assets)

                invalid_values = {"nan", "none", "null", "na"}

                def filter_by_keys(df, asset_keys, device_keys):
                    """用资产编号或设备编号过滤 DataFrame"""
                    mask = pd.Series([False] * len(df))
                    if asset_col and asset_keys:
                        am = df[asset_col].astype(str).str.strip().isin(asset_keys)
                        valid = ~df[asset_col].astype(str).str.strip().str.lower().isin(
                            invalid_values
                        )
                        mask = mask | (am & valid)
                    if device_col and device_keys:
                        dm = df[device_col].astype(str).str.strip().isin(device_keys)
                        mask = mask | dm
                    return mask

                # 选列：name, asset, keeper + device(用于NA替补)
                name_col = None
                for c in ["设备名称", "設備名称", "資產名稱", "资产名称"]:
                    if c in self.this_SFC_data.columns:
                        name_col = c
                        break
                if not name_col:
                    name_col = self.this_SFC_data.columns[0]

                keeper_col = "保管人"
                if keeper_col not in self.this_SFC_data.columns:
                    for c in self.this_SFC_data.columns:
                        if "保管" in c or "管理" in c or "负责" in c:
                            keeper_col = c
                            break
                    else:
                        keeper_col = self.this_SFC_data.columns[-1]

                select_cols = [name_col, asset_col, keeper_col]
                if device_col and device_col not in select_cols:
                    select_cols.append(device_col)

                new_mask = filter_by_keys(
                    self.this_SFC_data, new_by_asset, new_by_device
                )
                new_df = self.this_SFC_data[new_mask][select_cols].drop_duplicates()

                removed_mask = filter_by_keys(
                    self.last_SFC_data, removed_by_asset, removed_by_device
                )
                removed_df = self.last_SFC_data[removed_mask][
                    select_cols
                ].drop_duplicates()

                # 资产编号为 NA 时用设备编号替代显示
                def fill_na_asset(df):
                    if device_col and asset_col:
                        na_mask = (
                            df[asset_col]
                            .astype(str)
                            .str.strip()
                            .str.lower()
                            .isin(invalid_values)
                        )
                        df = df.copy()
                        df.loc[na_mask, asset_col] = df.loc[na_mask, device_col]
                    return df

                new_df = fill_na_asset(new_df)
                removed_df = fill_na_asset(removed_df)

                # 统一输出列名
                out_cols_map = {
                    name_col: "设备名称",
                    asset_col: "资产编号",
                    keeper_col: "保管人",
                }
                new_df = new_df.rename(columns=out_cols_map)
                removed_df = removed_df.rename(columns=out_cols_map)
                # 只保留输出用的三列
                new_df = new_df[["设备名称", "资产编号", "保管人"]]
                removed_df = removed_df[["设备名称", "资产编号", "保管人"]]

                # 創建Excel寫入器
                with pd.ExcelWriter(SFC_SFC_SAVE_PATH, engine="openpyxl") as writer:
                    # 創建一個空的DataFrame來初始化工作表
                    empty_df = pd.DataFrame()
                    empty_df.to_excel(writer, sheet_name="對比結果", index=False)

                    # 獲取工作表
                    worksheet = writer.sheets["對比結果"]

                    # 寫入標題信息
                    worksheet.merge_cells("A1:D1")
                    worksheet["A1"] = (
                        f"本月SFC资产_VS_上月SFC资产 (对比时间{current_time})"
                    )
                    worksheet["A1"].font = openpyxl.styles.Font(bold=True, size=14)
                    worksheet["A1"].alignment = openpyxl.styles.Alignment(
                        horizontal="center"
                    )
                    # 合并并居中 A2:D2 和 A3:D3
                    worksheet.merge_cells("A2:D2")
                    worksheet.merge_cells("A3:D3")
                    worksheet["A2"].alignment = openpyxl.styles.Alignment(
                        horizontal="center"
                    )
                    worksheet["A3"].alignment = openpyxl.styles.Alignment(
                        horizontal="center"
                    )

                    current_row = 3

                    # 寫入新增資產數據
                    if not new_df.empty:
                        # 標題
                        worksheet[f"A{current_row}"] = (
                            f"本月比上月新增资产 {len(self.new_assets)}笔"
                        )
                        worksheet[f"A{current_row}"].font = openpyxl.styles.Font(
                            bold=True, size=12
                        )
                        current_row += 1

                        # 列標題
                        headers = ["No.", "设备名称", "资产编号", "保管人"]
                        for i, header in enumerate(headers, 1):
                            cell = worksheet.cell(
                                row=current_row, column=i, value=header
                            )
                            cell.font = openpyxl.styles.Font(bold=True)
                            cell.fill = openpyxl.styles.PatternFill(
                                start_color="E6F3FF",
                                end_color="E6F3FF",
                                fill_type="solid",
                            )
                        current_row += 1

                        # 數據
                        for i in range(len(new_df)):
                            worksheet.cell(row=current_row + i, column=1, value=i + 1)
                            worksheet.cell(
                                row=current_row + i,
                                column=2,
                                value=new_df.iloc[i]["设备名称"],
                            )
                            worksheet.cell(
                                row=current_row + i,
                                column=3,
                                value=new_df.iloc[i]["资产编号"],
                            )
                            worksheet.cell(
                                row=current_row + i,
                                column=4,
                                value=new_df.iloc[i]["保管人"],
                            )
                        current_row += len(new_df) + 2

                    # 寫入減少資產數據
                    if not removed_df.empty:
                        worksheet[f"A{current_row}"] = (
                            f"本月比上月减少资产 {len(self.removed_assets)}笔"
                        )
                        worksheet[f"A{current_row}"].font = openpyxl.styles.Font(
                            bold=True, size=12
                        )
                        current_row += 1

                        headers = ["No.", "设备名称", "资产编号", "保管人"]
                        for i, header in enumerate(headers, 1):
                            cell = worksheet.cell(
                                row=current_row, column=i, value=header
                            )
                            cell.font = openpyxl.styles.Font(bold=True)
                            cell.fill = openpyxl.styles.PatternFill(
                                start_color="FFE6E6",
                                end_color="FFE6E6",
                                fill_type="solid",
                            )
                        current_row += 1

                        for i in range(len(removed_df)):
                            worksheet.cell(row=current_row + i, column=1, value=i + 1)
                            worksheet.cell(
                                row=current_row + i,
                                column=2,
                                value=removed_df.iloc[i]["设备名称"],
                            )
                            worksheet.cell(
                                row=current_row + i,
                                column=3,
                                value=removed_df.iloc[i]["资产编号"],
                            )
                            worksheet.cell(
                                row=current_row + i,
                                column=4,
                                value=removed_df.iloc[i]["保管人"],
                            )

                    # 設置列寬和邊框
                    from openpyxl.styles import Border, Side  # noqa: E402, I001, UP015, F401

                    thin_border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"),
                    )

                    # 設置列寬
                    column_widths = [8, 40, 25, 20]  # 根據內容調整
                    for i, width in enumerate(column_widths, 1):
                        worksheet.column_dimensions[
                            openpyxl.utils.get_column_letter(i)
                        ].width = width

                    # 添加邊框
                    for row in worksheet.iter_rows(
                        min_row=1,
                        max_row=worksheet.max_row,
                        min_col=1,
                        max_col=worksheet.max_column,
                    ):
                        for cell in row:
                            cell.border = thin_border

                self._Save_signal.emit(self.ui.frame_2)
        except Exception as e:
            self._Error_signal.emit()
            self._unlock_signal.emit()
            logger.exception(e)

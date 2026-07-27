
class pyqtSignal:
    def __init__(self, *args, **kwargs): pass
    def connect(self, *args, **kwargs): pass
    def emit(self, *args, **kwargs): pass

class QThread:
    pass

class QObject:
    pass

class QWidget:
    pass

def safe_thread_run(func):
    def wrapper(*args, **kwargs):
        return func(*args, **kwargs)
    return wrapper

import os
from datetime import datetime

import openpyxl
import pandas as pd
import polars as pl
from app.services.asset_engine.const import NOTES_NOTES_SAVE_PATH
from loguru import logger
from app.services.asset_engine.mod import safe_thread_run

No_CheckRFID = ["A1300011C5C3", "A13000103933", "A1300010E606"]


class Notes_Notes(QThread):
    _This_Assets_Notes_signal = pyqtSignal(int, int, int)
    _Last_Assets_Notes_signal = pyqtSignal(int, int, int)
    _This_Last_Assets_Notes_signal = pyqtSignal(int, int, int, int, int, int)
    _unlock_signal = pyqtSignal()
    _Error_signal = pyqtSignal()
    _Save_signal = pyqtSignal(object)
    _Update_Message_signal = pyqtSignal(str)
    _progress_signal = pyqtSignal(int, str)

    # 進度信號：進度值, 進度文本

    def __init__(self, ui):
        super().__init__()
        self.ui = ui

        self.this_invalid_all_rows = None
        self.last_assets_filtered = None
        self.last_All_Notes_ = None
        self.this_assets_filtered = None
        self.this_All_Notes_ = None
        self.last_invalid_all_rows = None

        self.removed_No_assets = []
        self.new_No_assets = []
        self.removed_assets = []
        self.new_assets = []
        self.last_Notes_data = []
        self.this_Notes_data = []

        self.This_Notes_path = None
        self.Last_Notes_path = None

    @safe_thread_run
    def run(self):
        if not self.This_Notes_path or not self.Last_Notes_path:
            self.ui.this_Notes_edit.setText("请选择文件")
            return
        try:
            self._Update_Message_signal.emit("Start")
            self._progress_signal.emit(10, "開始讀取本月Notes數據...")
            self.This_Notes_date()
            self._progress_signal.emit(40, "開始讀取上月Notes數據...")
            self.Last_Notes_date()
            self._progress_signal.emit(70, "開始進行數據對比...")
            self.Notes_Notes_Comparison()
            self._progress_signal.emit(100, "對比完成")
        except Exception as e:
            logger.exception(e)
            self._unlock_signal.emit()
            self._Error_signal.emit()

    def safe_read_excel(self, path):
        """安全讀取Excel並返回 DataFrame（不是 LazyFrame）"""
        required_columns = {"資產編號"}

        try:
            # 嘗試使用 Polars 讀取 Excel
            # 注意：设置 infer_schema_length=0 让Polars将所有列读取为字符串，避免类型推断错误
            df_polars = pl.read_excel(path, infer_schema_length=0)

            # 檢查是否包含必需的列
            header_columns = set(df_polars.columns)
            if required_columns.intersection(header_columns):
                # 自动去除资产编号列末尾的点号
                if "資產編號" in df_polars.columns:
                    df_polars = df_polars.with_columns(
                        [pl.col("資產編號").str.strip_chars_end(".").alias("資產編號")]
                    )
                # 移除全空列並返回 DataFrame
                non_null_cols = [
                    col
                    for col in df_polars.columns
                    if not df_polars[col].is_null().all()
                ]
                return df_polars.select(non_null_cols)

            # 如果表頭不正確，嘗試查找正確的表頭行
            for i in range(min(5, len(df_polars))):
                row_values = df_polars.row(i)
                clean_columns = {
                    str(col).strip() for col in row_values if col is not None
                }
                if required_columns.intersection(clean_columns):
                    new_columns = [
                        str(col).strip() if col is not None else f"col_{j}"
                        for j, col in enumerate(row_values)
                    ]
                    df_polars = df_polars.slice(i + 1).rename(
                        {old: new for old, new in zip(df_polars.columns, new_columns)}
                    )
                    # 自动去除资产编号列末尾的点号
                    if "資產編號" in df_polars.columns:
                        df_polars = df_polars.with_columns(
                            [
                                pl.col("資產編號")
                                .str.strip_chars_end(".")
                                .alias("資產編號")
                            ]
                        )
                    # 移除全空列並返回 DataFrame
                    non_null_cols = [
                        col
                        for col in df_polars.columns
                        if not df_polars[col].is_null().all()
                    ]
                    return df_polars.select(non_null_cols)

            raise ValueError(f"未在文件 {os.path.basename(path)} 中找到完整表頭")

        except Exception as e:
            logger.error(f"读取 {os.path.basename(path)} 失败: {e}")
            self._unlock_signal.emit()
            self._Error_signal.emit()
            raise

    def This_Notes_date(self):
        """讀取本月Notes數據（使用LazyFrame）"""
        self.this_Notes_data = None
        if not self.This_Notes_path:
            return
        try:
            df_lazy = self.safe_read_excel(self.This_Notes_path)
            self.this_Notes_data = df_lazy
        except Exception as e:
            self._unlock_signal.emit()
            self._Error_signal.emit()
            logger.error(e)

    def Last_Notes_date(self):
        """讀取上月Notes數據（使用LazyFrame）"""
        self.last_Notes_data = None
        if not self.Last_Notes_path:
            return
        try:
            df_lazy = self.safe_read_excel(self.Last_Notes_path)
            self.last_Notes_data = df_lazy
        except Exception as e:
            self._unlock_signal.emit()
            logger.error(e)

    def Notes_Notes_Comparison(self):
        """Notes與Notes數據比較（使用LazyFrame優化）"""
        if self.this_Notes_data is None or self.last_Notes_data is None:
            logger.error("文件讀取錯誤，請檢查文件")
            self._unlock_signal.emit()
            self._Error_signal.emit()
            return
        else:
            try:
                # 使用 DataFrame 操作，獲取所有資產名稱
                this_all_notes_query = self.this_Notes_data.filter(
                    pl.col("資產名稱").is_not_null()
                ).select("資產名稱")
                self.this_All_Notes_ = this_all_notes_query.to_series().to_list()

                last_all_notes_query = self.last_Notes_data.filter(
                    pl.col("資產名稱").is_not_null()
                ).select("資產名稱")
                self.last_All_Notes_ = last_all_notes_query.to_series().to_list()

                # 過濾和匹配資產編號（使用DataFrame）- 只识别以18-或13-开头的为有效资产编号
                this_assets_query = self.this_Notes_data.filter(
                    pl.col("資產編號").is_not_null()
                    & (
                        pl.col("資產編號")
                        .cast(pl.String)
                        .str.strip_chars()
                        .str.len_chars()
                        > 0
                    )
                    & (
                        pl.col("資產編號").cast(pl.String).str.starts_with("18-")
                        | pl.col("資產編號").cast(pl.String).str.starts_with("13-")
                    )
                ).select("資產編號")
                self.this_assets_filtered = this_assets_query.to_series().to_list()

                last_assets_query = self.last_Notes_data.filter(
                    pl.col("資產編號").is_not_null()
                    & (
                        pl.col("資產編號")
                        .cast(pl.String)
                        .str.strip_chars()
                        .str.len_chars()
                        > 0
                    )
                    & (
                        pl.col("資產編號").cast(pl.String).str.starts_with("18-")
                        | pl.col("資產編號").cast(pl.String).str.starts_with("13-")
                    )
                ).select("資產編號")
                self.last_assets_filtered = last_assets_query.to_series().to_list()

                self.new_assets = set(self.this_assets_filtered) - set(
                    self.last_assets_filtered
                )
                self.removed_assets = set(self.last_assets_filtered) - set(
                    self.this_assets_filtered
                )

                # 處理無資產記錄（資產編號為空或不以18-或13-开头的记录）
                # 使用機身SN来匹配无资产记录的变化

                # 獲取本月無資產記錄（带机身SN）
                this_no_asset_df = self.this_Notes_data.filter(
                    pl.col("資產編號").is_null()
                    | (
                        pl.col("資產編號")
                        .cast(pl.String)
                        .str.strip_chars()
                        .str.len_chars()
                        == 0
                    )
                    | (
                        ~(
                            pl.col("資產編號").cast(pl.String).str.starts_with("18-")
                            | pl.col("資產編號").cast(pl.String).str.starts_with("13-")
                        )
                    )
                ).select(["資產名稱", "機身SN", "保管人"])

                # 獲取上月無資產記錄（带机身SN）
                last_no_asset_df = self.last_Notes_data.filter(
                    pl.col("資產編號").is_null()
                    | (
                        pl.col("資產編號")
                        .cast(pl.String)
                        .str.strip_chars()
                        .str.len_chars()
                        == 0
                    )
                    | (
                        ~(
                            pl.col("資產編號").cast(pl.String).str.starts_with("18-")
                            | pl.col("資產編號").cast(pl.String).str.starts_with("13-")
                        )
                    )
                ).select(["資產名稱", "機身SN", "保管人"])

                self.this_invalid_all_rows = this_no_asset_df.height
                self.last_invalid_all_rows = last_no_asset_df.height
                logger.info(
                    f"Notes_Notes本月无资产记录数量: {self.this_invalid_all_rows}"
                )
                logger.info(
                    f"Notes_Notes上月无资产记录数量: {self.last_invalid_all_rows}"
                )

                # 使用機身SN来识别新增和减少的无资产记录
                # 创建用于比对的唯一标识（资产名称 + 机身SN）
                this_no_asset_keys = (
                    this_no_asset_df.select(
                        pl.concat_str(
                            [pl.col("資產名稱"), pl.col("機身SN")], separator="||"
                        ).alias("key")
                    )
                    .to_series()
                    .to_list()
                )

                last_no_asset_keys = (
                    last_no_asset_df.select(
                        pl.concat_str(
                            [pl.col("資產名稱"), pl.col("機身SN")], separator="||"
                        ).alias("key")
                    )
                    .to_series()
                    .to_list()
                )

                # 找出新增和减少的无资产记录
                new_no_asset_keys_set = set(this_no_asset_keys) - set(
                    last_no_asset_keys
                )
                removed_no_asset_keys_set = set(last_no_asset_keys) - set(
                    this_no_asset_keys
                )

                # 保存新增和减少的无资产记录的key列表（用于后续保存时过滤）
                self.new_No_assets = list(new_no_asset_keys_set)
                self.removed_No_assets = list(removed_no_asset_keys_set)

                logger.info(f"Notes_Notes新增无资产记录: {len(self.new_No_assets)} 笔")
                logger.info(
                    f"Notes_Notes减少无资产记录: {len(self.removed_No_assets)} 笔"
                )

                self._This_Assets_Notes_signal.emit(
                    len(self.this_All_Notes_),
                    len(self.this_assets_filtered),
                    self.this_invalid_all_rows,
                )
                self._Last_Assets_Notes_signal.emit(
                    len(self.last_All_Notes_),
                    len(self.last_assets_filtered),
                    self.last_invalid_all_rows,
                )

                if (
                    len(self.new_assets) > 0
                    or len(self.removed_assets) > 0
                    or len(self.new_No_assets) > 0
                    or len(self.removed_No_assets) > 0
                ):
                    self._This_Last_Assets_Notes_signal.emit(
                        abs(len(self.new_assets) - len(self.removed_assets)),
                        len(self.new_assets),
                        len(self.removed_assets),
                        abs(len(self.new_No_assets) - len(self.removed_No_assets)),
                        len(self.new_No_assets),
                        len(self.removed_No_assets),
                    )
                else:
                    self._This_Last_Assets_Notes_signal.emit(0, 0, 0, 0, 0, 0)

            except Exception as e:
                self._unlock_signal.emit()
                self._Error_signal.emit()
                logger.error(e)

    def Save_Notes_Notes_Comparison(self):
        """保存Notes與Notes比較結果為Excel文件"""
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if (
            not self.new_assets
            and not self.removed_assets
            and not self.new_No_assets
            and not self.removed_No_assets
        ):
            return
        else:
            try:
                # 創建Excel寫入器
                with pd.ExcelWriter(NOTES_NOTES_SAVE_PATH, engine="openpyxl") as writer:
                    # 使用 DataFrame 構建查詢
                    new_df = (
                        self.this_Notes_data.filter(
                            pl.col("資產編號").is_in(list(self.new_assets))
                        )
                        .select(["資產名稱", "資產編號", "保管人"])
                        .unique()
                    )

                    removed_df = (
                        self.last_Notes_data.filter(
                            pl.col("資產編號").is_in(list(self.removed_assets))
                        )
                        .select(["資產名稱", "資產編號", "保管人"])
                        .unique()
                    )

                    # 轉換為pandas DataFrame
                    new_df_pandas = new_df.to_pandas()
                    removed_df_pandas = removed_df.to_pandas()

                    # 創建一個空的DataFrame來初始化工作表
                    empty_df = pd.DataFrame()
                    empty_df.to_excel(writer, sheet_name="對比結果", index=False)

                    # 獲取工作表
                    worksheet = writer.sheets["對比結果"]

                    # 寫入標題信息
                    worksheet.merge_cells("A1:D1")
                    worksheet["A1"] = f"本月Notes_VS_上月Notes (对比时间{current_time})"
                    worksheet["A1"].font = openpyxl.styles.Font(bold=True, size=14)
                    worksheet["A1"].alignment = openpyxl.styles.Alignment(
                        horizontal="center"
                    )
                    # 合并并居中 A2:D2 和 A3:D3
                    worksheet.merge_cells("A2:D2")
                    worksheet.merge_cells("A3:D3")
                    worksheet["A2"].alignment = openpyxl.styles.Alignment(
                        horizontal="center"
                    )
                    worksheet["A3"].alignment = openpyxl.styles.Alignment(
                        horizontal="center"
                    )

                    current_row = 3

                    # 寫入新增資產數據
                    if not new_df_pandas.empty:
                        # 標題
                        worksheet[f"A{current_row}"] = (
                            f"本月比上月新增资产 {len(self.new_assets)}笔"
                        )
                        worksheet[f"A{current_row}"].font = openpyxl.styles.Font(
                            bold=True, size=12
                        )
                        current_row += 1

                        # 列標題
                        headers = ["No.", "資產名稱", "資產編號", "保管人"]
                        for i, header in enumerate(headers, 1):
                            cell = worksheet.cell(
                                row=current_row, column=i, value=header
                            )
                            cell.font = openpyxl.styles.Font(bold=True)
                            cell.fill = openpyxl.styles.PatternFill(
                                start_color="E6F3FF",
                                end_color="E6F3FF",
                                fill_type="solid",
                            )
                        current_row += 1

                        # 數據
                        for i in range(len(new_df_pandas)):
                            worksheet.cell(row=current_row + i, column=1, value=i + 1)
                            worksheet.cell(
                                row=current_row + i,
                                column=2,
                                value=new_df_pandas.iloc[i]["資產名稱"],
                            )
                            worksheet.cell(
                                row=current_row + i,
                                column=3,
                                value=new_df_pandas.iloc[i]["資產編號"],
                            )
                            worksheet.cell(
                                row=current_row + i,
                                column=4,
                                value=new_df_pandas.iloc[i]["保管人"],
                            )
                        current_row += len(new_df_pandas) + 2

                    # 寫入減少資產數據
                    if not removed_df_pandas.empty:
                        worksheet[f"A{current_row}"] = (
                            f"本月比上月减少资产 {len(self.removed_assets)}笔"
                        )
                        worksheet[f"A{current_row}"].font = openpyxl.styles.Font(
                            bold=True, size=12
                        )
                        current_row += 1

                        headers = ["No.", "資產名稱", "資產編號", "保管人"]
                        for i, header in enumerate(headers, 1):
                            cell = worksheet.cell(
                                row=current_row, column=i, value=header
                            )
                            cell.font = openpyxl.styles.Font(bold=True)
                            cell.fill = openpyxl.styles.PatternFill(
                                start_color="FFE6E6",
                                end_color="FFE6E6",
                                fill_type="solid",
                            )
                        current_row += 1

                        for i in range(len(removed_df_pandas)):
                            worksheet.cell(row=current_row + i, column=1, value=i + 1)
                            worksheet.cell(
                                row=current_row + i,
                                column=2,
                                value=removed_df_pandas.iloc[i]["資產名稱"],
                            )
                            worksheet.cell(
                                row=current_row + i,
                                column=3,
                                value=removed_df_pandas.iloc[i]["資產編號"],
                            )
                            worksheet.cell(
                                row=current_row + i,
                                column=4,
                                value=removed_df_pandas.iloc[i]["保管人"],
                            )
                        current_row += len(removed_df_pandas) + 2

                    # 處理無資產記錄的變化
                    if len(self.new_No_assets) > 0:
                        # 根据key列表过滤出新增的无资产记录
                        new_No_df = (
                            self.this_Notes_data.filter(
                                pl.col("資產編號").is_null()
                                | (
                                    pl.col("資產編號")
                                    .cast(pl.String)
                                    .str.strip_chars()
                                    .str.len_chars()
                                    == 0
                                )
                                | (
                                    ~(
                                        pl.col("資產編號")
                                        .cast(pl.String)
                                        .str.starts_with("18-")
                                        | pl.col("資產編號")
                                        .cast(pl.String)
                                        .str.starts_with("13-")
                                    )
                                )
                            )
                            .select(["資產名稱", "機身SN", "保管人"])
                            .with_columns(
                                pl.concat_str(
                                    [pl.col("資產名稱"), pl.col("機身SN")],
                                    separator="||",
                                ).alias("key")
                            )
                            .filter(pl.col("key").is_in(self.new_No_assets))
                            .drop("key")
                        )

                        if not new_No_df.is_empty():
                            new_No_df_pandas = new_No_df.to_pandas()
                            worksheet[f"A{current_row}"] = (
                                f"本月比上月新增无资产记录 {len(self.new_No_assets)}笔"
                            )
                            worksheet[f"A{current_row}"].font = openpyxl.styles.Font(
                                bold=True, size=12
                            )
                            current_row += 1

                            headers = ["No.", "資產名稱", "機身SN", "保管人"]
                            for i, header in enumerate(headers, 1):
                                cell = worksheet.cell(
                                    row=current_row, column=i, value=header
                                )
                                cell.font = openpyxl.styles.Font(bold=True)
                                cell.fill = openpyxl.styles.PatternFill(
                                    start_color="E6F3FF",
                                    end_color="E6F3FF",
                                    fill_type="solid",
                                )
                            current_row += 1

                            for i in range(len(new_No_df_pandas)):
                                worksheet.cell(
                                    row=current_row + i, column=1, value=i + 1
                                )
                                worksheet.cell(
                                    row=current_row + i,
                                    column=2,
                                    value=new_No_df_pandas.iloc[i]["資產名稱"],
                                )
                                worksheet.cell(
                                    row=current_row + i,
                                    column=3,
                                    value=new_No_df_pandas.iloc[i]["機身SN"],
                                )
                                worksheet.cell(
                                    row=current_row + i,
                                    column=4,
                                    value=new_No_df_pandas.iloc[i]["保管人"],
                                )
                            current_row += len(new_No_df_pandas) + 2

                    if len(self.removed_No_assets) > 0:
                        # 根据key列表过滤出减少的无资产记录
                        removed_No_df = (
                            self.last_Notes_data.filter(
                                pl.col("資產編號").is_null()
                                | (
                                    pl.col("資產編號")
                                    .cast(pl.String)
                                    .str.strip_chars()
                                    .str.len_chars()
                                    == 0
                                )
                                | (
                                    ~(
                                        pl.col("資產編號")
                                        .cast(pl.String)
                                        .str.starts_with("18-")
                                        | pl.col("資產編號")
                                        .cast(pl.String)
                                        .str.starts_with("13-")
                                    )
                                )
                            )
                            .select(["資產名稱", "機身SN", "保管人"])
                            .with_columns(
                                pl.concat_str(
                                    [pl.col("資產名稱"), pl.col("機身SN")],
                                    separator="||",
                                ).alias("key")
                            )
                            .filter(pl.col("key").is_in(self.removed_No_assets))
                            .drop("key")
                        )

                        if not removed_No_df.is_empty():
                            removed_No_df_pandas = removed_No_df.to_pandas()
                            worksheet[f"A{current_row}"] = (
                                f"本月比上月减少无资产记录 {len(self.removed_No_assets)}笔"
                            )
                            worksheet[f"A{current_row}"].font = openpyxl.styles.Font(
                                bold=True, size=12
                            )
                            current_row += 1

                            headers = ["No.", "資產名稱", "機身SN", "保管人"]
                            for i, header in enumerate(headers, 1):
                                cell = worksheet.cell(
                                    row=current_row, column=i, value=header
                                )
                                cell.font = openpyxl.styles.Font(bold=True)
                                cell.fill = openpyxl.styles.PatternFill(
                                    start_color="FFE6E6",
                                    end_color="FFE6E6",
                                    fill_type="solid",
                                )
                            current_row += 1

                            for i in range(len(removed_No_df_pandas)):
                                worksheet.cell(
                                    row=current_row + i, column=1, value=i + 1
                                )
                                worksheet.cell(
                                    row=current_row + i,
                                    column=2,
                                    value=removed_No_df_pandas.iloc[i]["資產名稱"],
                                )
                                worksheet.cell(
                                    row=current_row + i,
                                    column=3,
                                    value=removed_No_df_pandas.iloc[i]["機身SN"],
                                )
                                worksheet.cell(
                                    row=current_row + i,
                                    column=4,
                                    value=removed_No_df_pandas.iloc[i]["保管人"],
                                )

                    # 設置列寬和邊框
                    from openpyxl.styles import Border, Side

                    thin_border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"),
                    )

                    # 設置列寬
                    column_widths = [8, 40, 25, 20]  # 根據內容調整
                    for i, width in enumerate(column_widths, 1):
                        worksheet.column_dimensions[
                            openpyxl.utils.get_column_letter(i)
                        ].width = width

                    # 添加邊框
                    for row in worksheet.iter_rows(
                        min_row=1,
                        max_row=worksheet.max_row,
                        min_col=1,
                        max_col=worksheet.max_column,
                    ):
                        for cell in row:
                            cell.border = thin_border

                self._Save_signal.emit(self.ui.frame_3)

            except Exception as e:
                self._unlock_signal.emit()
                logger.error(e)

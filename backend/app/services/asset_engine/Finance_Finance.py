
class pyqtSignal:
    def __init__(self, *args, **kwargs): pass
    def connect(self, *args, **kwargs): pass
    def emit(self, *args, **kwargs): pass

class QThread:
    pass

class QObject:
    pass

class QWidget:
    pass

def safe_thread_run(func):  # noqa: F811
    def wrapper(*args, **kwargs):
        return func(*args, **kwargs)
    return wrapper

import os  # noqa: E402, I001, UP015, F401
from datetime import datetime  # noqa: E402, I001, UP015, F401

import openpyxl  # noqa: E402, I001, UP015, F401
import pandas as pd  # noqa: E402, I001, UP015, F401
import polars as pl  # noqa: E402, I001, UP015, F401
from app.services.asset_engine.const import FINANCE_FINANCE_SAVE_PATH  # noqa: E402, I001, UP015, F401
from loguru import logger  # noqa: E402, I001, UP015, F401


class Finance_Finance(QThread):
    _Update_Message_signal = pyqtSignal(str)
    _Update_this_pysignal = pyqtSignal(int, int)
    _Update_last_pysignal = pyqtSignal(int, int)
    _update_check_signal = pyqtSignal(int, int, int, int, int, int, int, int)
    _unlock_signal = pyqtSignal()
    _Error_signal = pyqtSignal()
    _Save_signal = pyqtSignal(object)
    _progress_signal = pyqtSignal(int, str)  # 進度信號：進度值, 進度文本

    def __init__(self, ui):
        super().__init__()
        self.ui = ui
        self.Department_Difference = []
        self.Custodian_Difference = []
        self.check_Custodian = []
        self.check_Department = []

        self.this_Finance_path = None
        self.last_Finance_path = None
        self.Custodian_path = None
        self.Department_path = None

        self.this_Finance_data = None
        self.last_Finance_data = None
        self.Custodian_data = None
        self.Department_data = None

        self.this_Custodian_assets = None
        self.last_Custodian_assets = None
        self.this_Department_assets = None
        self.last_Department_assets = None

        self.new_Custodian_assets = []
        self.new_Department_assets = []
        self.removed_Custodian_assets = []
        self.removed_Department_assets = []

    @safe_thread_run
    def run(self):
        if not self.this_Finance_path or not self.last_Finance_path:
            self._Update_Message_signal.emit("请选择文件")
            return
        if not self.Custodian_path and not self.Department_path:
            self._Update_Message_signal.emit("请选择保管人或\n部门代码")
            return
        self._Update_Message_signal.emit("Start")
        try:
            self._progress_signal.emit(10, "讀取保管人數據...")
            self.read_Custodian_data()
            self._progress_signal.emit(20, "讀取部門數據...")
            self.read_Department_data()
            self._progress_signal.emit(40, "讀取本月財務數據...")
            self.read_this_Finance_data()
            self._progress_signal.emit(60, "讀取上月財務數據...")
            self.read_last_Finance_data()
            self._progress_signal.emit(80, "開始財務數據對比...")
            self.Finance_check()
            self._progress_signal.emit(100, "對比完成")
        except Exception as e:
            logger.exception(e)
            self._unlock_signal.emit()
            self._Error_signal.emit()

    def read_Custodian_data(self):
        self.Custodian_data = []
        if self.Custodian_path is None:
            self.Custodian_data = []
        else:
            with open(self.Custodian_path, encoding="utf-8") as file:
                lines = file.readlines()  # 读取所有行
                self.Custodian_data = [line.strip() for line in lines]

    def read_Department_data(self):
        self.Department_data = []
        if self.Department_path is None:
            self.Department_data = []
        else:
            with open(self.Department_path, encoding="utf-8") as file:
                lines = file.readlines()  # 读取所有行
                self.Department_data = [line.strip() for line in lines]

    def safe_read_excel(self, path):
        """使用 polars 讀取 Excel 並返回 DataFrame（不是 LazyFrame）"""
        try:
            # 先用 read_excel 讀取數據
            # 注意：设置 infer_schema_length=0 让Polars将所有列读取为字符串，避免类型推断错误
            df_eager = pl.read_excel(path, infer_schema_length=0)

            # 檢查表頭中是否包含空值
            header_row = df_eager.columns
            if any(
                cell is None or "__UNNAMED__" in str(cell).strip()
                for cell in header_row
            ):
                # 如果需要重新設置表頭，使用第一行數據作為表頭
                new_columns = [
                    str(col).strip() if col is not None else f"col_{i}"
                    for i, col in enumerate(df_eager.row(0))
                ]
                df_eager = df_eager.slice(1).rename(
                    {old: new for old, new in zip(df_eager.columns, new_columns,strict=False)}
                )
                print(f"文件 {os.path.basename(path)} 的表頭包含空值，已重新設置")

            # 自动去除资产编号列末尾的点号
            if "資產編號" in df_eager.columns:
                df_eager = df_eager.with_columns(
                    [pl.col("資產編號").str.strip_chars_end(".").alias("資產編號")]
                )

            # 返回 DataFrame 而不是 LazyFrame
            return df_eager

        except Exception as e:
            logger.error(f"读取 {os.path.basename(path)} 失败: {e}")
            self._unlock_signal.emit()
            raise

    def read_this_Finance_data(self):
        """讀取本月財務數據（使用 DataFrame）"""
        self.this_Finance_data = None
        if not self.this_Finance_path:
            return
        self.this_Finance_data = self.safe_read_excel(self.this_Finance_path)

    def read_last_Finance_data(self):
        """讀取上月財務數據（使用 DataFrame）"""
        self.last_Finance_data = None
        if not self.last_Finance_path:
            return
        self.last_Finance_data = self.safe_read_excel(self.last_Finance_path)

    def Finance_check(self):
        try:
            # 先檢查 DataFrame 是否為空
            if self.last_Finance_data.is_empty() or self.this_Finance_data.is_empty():
                logger.error("数据为空")
                return

            print("開始處理財務數據比較...")

            # 使用 DataFrame 構建查詢，在最後一次性執行
            this_Custodian_query = self.this_Finance_data.filter(
                pl.col("保管人員").is_in(self.Custodian_data)
            )

            this_Department_query = self.this_Finance_data.filter(
                pl.col("資產所屬部門代號")
                .cast(pl.String)
                .str.zfill(4)
                .is_in(self.Department_data)
            )

            last_Custodian_query = self.last_Finance_data.filter(
                pl.col("保管人員").is_in(self.Custodian_data)
            )

            last_Department_query = self.last_Finance_data.filter(
                pl.col("資產所屬部門代號")
                .cast(pl.String)
                .str.zfill(4)
                .is_in(self.Department_data)
            )

            # 直接獲得過濾後的數據（不需要 collect）
            this_Custodian_filtered = this_Custodian_query
            this_Department_filtered = this_Department_query
            last_Custodian_filtered = last_Custodian_query
            last_Department_filtered = last_Department_query

            if len(this_Custodian_filtered) > 0:
                # 獲取部門數據並檢查異常
                current_departments = this_Custodian_filtered.select(
                    pl.col("資產所屬部門代號").cast(pl.String).str.zfill(4)
                ).to_series()

                # 找出不在部門列表中的記錄
                invalid_department_mask = ~current_departments.is_in(
                    self.Department_data
                )
                invalid_department_df = this_Custodian_filtered.filter(
                    invalid_department_mask
                )

                self.check_Department = (
                    invalid_department_df.select("資產編號").to_series().to_list()
                    if len(invalid_department_df) > 0
                    else []
                )
                logger.info(f"发现{len(invalid_department_df)}条异常部门记录")
            else:
                self.check_Department = []

            if len(this_Department_filtered) > 0:
                # 獲取保管人數據並檢查異常
                current_custodians = this_Department_filtered.select(
                    "保管人員"
                ).to_series()

                # 找出不在保管人列表中的記錄
                invalid_custodian_mask = ~current_custodians.is_in(self.Custodian_data)
                invalid_custodian_df = this_Department_filtered.filter(
                    invalid_custodian_mask
                )

                self.check_Custodian = (
                    invalid_custodian_df.select("資產編號").to_series().to_list()
                    if len(invalid_custodian_df) > 0
                    else []
                )
                logger.info(f"发现{len(invalid_custodian_df)}条异常保管人记录")
            else:
                self.check_Custodian = []

            # 獲取資產編號列表
            self.this_Custodian_assets = (
                this_Custodian_filtered.select("資產編號").to_series().to_list()
            )
            self.last_Custodian_assets = (
                last_Custodian_filtered.select("資產編號").to_series().to_list()
            )

            self.this_Department_assets = (
                this_Department_filtered.select("資產編號").to_series().to_list()
            )
            self.last_Department_assets = (
                last_Department_filtered.select("資產編號").to_series().to_list()
            )

            self.new_Custodian_assets = set(self.this_Custodian_assets) - set(
                self.last_Custodian_assets
            )
            self.new_Department_assets = set(self.this_Department_assets) - set(
                self.last_Department_assets
            )
            self.removed_Custodian_assets = set(self.last_Custodian_assets) - set(
                self.this_Custodian_assets
            )
            self.removed_Department_assets = set(self.last_Department_assets) - set(
                self.this_Department_assets
            )

            self._Update_this_pysignal.emit(
                len(self.this_Department_assets), len(self.this_Custodian_assets)
            )
            self._Update_last_pysignal.emit(
                len(self.last_Department_assets), len(self.last_Custodian_assets)
            )

            self.Custodian_Difference = abs(
                len(self.new_Custodian_assets) - len(self.removed_Custodian_assets)
            )
            self.Department_Difference = abs(
                len(self.new_Department_assets) - len(self.removed_Department_assets)
            )

            self._update_check_signal.emit(
                self.Custodian_Difference,
                len(self.new_Custodian_assets),
                len(self.removed_Custodian_assets),
                self.Department_Difference,
                len(self.new_Department_assets),
                len(self.removed_Department_assets),
                len(
                    self.check_Custodian
                    if isinstance(self.check_Custodian, (list, pl.Series))
                    else 0
                ),
                len(
                    self.check_Department
                    if isinstance(self.check_Department, (list, pl.Series))
                    else 0
                ),
            )

        except Exception as e:
            self._Error_signal.emit()
            self._unlock_signal.emit()
            logger.exception(e)

    def Save_Check(self):
        """保存財務對比結果為Excel文件"""
        try:
            if (
                not self.new_Custodian_assets
                and not self.removed_Custodian_assets
                and not self.new_Department_assets
                and not self.removed_Department_assets
                and not self.check_Custodian
                and not self.check_Department
            ):
                return

            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # 使用 DataFrame 構建查詢
            new_df_query = (
                self.this_Finance_data.filter(
                    pl.col("資產編號").is_in(list(self.new_Custodian_assets))
                )
                .select(["資產名稱", "資產編號", "資產所屬部門代號", "保管人員"])
                .unique()
            )

            removed_df_query = (
                self.last_Finance_data.filter(
                    pl.col("資產編號").is_in(list(self.removed_Custodian_assets))
                )
                .select(["資產名稱", "資產編號", "資產所屬部門代號", "保管人員"])
                .unique()
            )

            new_df_Department_query = (
                self.this_Finance_data.filter(
                    pl.col("資產編號").is_in(list(self.new_Department_assets))
                )
                .select(["資產名稱", "資產編號", "資產所屬部門代號", "保管人員"])
                .unique()
            )

            removed_df_Department_query = (
                self.last_Finance_data.filter(
                    pl.col("資產編號").is_in(list(self.removed_Department_assets))
                )
                .select(["資產名稱", "資產編號", "資產所屬部門代號", "保管人員"])
                .unique()
            )

            Error_Custodian_query = (
                self.this_Finance_data.filter(
                    pl.col("資產編號").is_in(self.check_Custodian)
                )
                .select(["資產名稱", "資產所屬部門代號", "資產編號", "保管人員"])
                .unique()
            )

            Error_Department_query = (
                self.this_Finance_data.filter(
                    pl.col("資產編號").is_in(self.check_Department)
                )
                .select(["資產名稱", "資產所屬部門代號", "資產編號", "保管人員"])
                .unique()
            )

            # 直接獲取數據
            new_df = new_df_query
            removed_df = removed_df_query
            new_df_Department = new_df_Department_query
            removed_df_Department = removed_df_Department_query
            Error_Custodian = Error_Custodian_query
            Error_Department = Error_Department_query

            # 创建Excel文件
            with pd.ExcelWriter(FINANCE_FINANCE_SAVE_PATH, engine="openpyxl") as writer:
                # 创建一个空的DataFrame来初始化工作表
                empty_df = pd.DataFrame()
                empty_df.to_excel(writer, sheet_name="對比結果", index=False)

                # 获取工作表对象
                worksheet = writer.sheets["對比結果"]

                # 设置标题
                worksheet.merge_cells("A1:E1")
                worksheet["A1"] = f"本月财务_VS_上月财务 (对比时间{current_time})"
                worksheet["A1"].font = openpyxl.styles.Font(bold=True, size=14)
                worksheet["A1"].alignment = openpyxl.styles.Alignment(
                    horizontal="center"
                )
                # 合并并居中 A2:D2 和 A3:D3
                worksheet.merge_cells("A2:D2")
                worksheet.merge_cells("A3:D3")
                worksheet["A2"].alignment = openpyxl.styles.Alignment(
                    horizontal="center"
                )
                worksheet["A3"].alignment = openpyxl.styles.Alignment(
                    horizontal="center"
                )

                current_row = 3

                # 处理保管人新增资产
                if not new_df.is_empty():
                    current_row = self._add_section_to_excel(
                        worksheet,
                        new_df.to_pandas(),
                        f"依保管人本月比上月新增资产 {len(self.new_Custodian_assets)}笔",
                        current_row,
                        "new",
                    )

                # 处理保管人减少资产
                if not removed_df.is_empty():
                    current_row = self._add_section_to_excel(
                        worksheet,
                        removed_df.to_pandas(),
                        f"依保管人本月比上月减少资产 {len(self.removed_Custodian_assets)}笔",
                        current_row,
                        "removed",
                    )

                # 处理部门新增资产
                if not new_df_Department.is_empty():
                    current_row = self._add_section_to_excel(
                        worksheet,
                        new_df_Department.to_pandas(),
                        f"依保管部门本月比上月新增资产 {len(self.new_Department_assets)}笔",
                        current_row,
                        "new",
                    )

                # 处理部门减少资产
                if not removed_df_Department.is_empty():
                    current_row = self._add_section_to_excel(
                        worksheet,
                        removed_df_Department.to_pandas(),
                        f"依保管部门本月比上月减少资产 {len(self.removed_Department_assets)}笔",
                        current_row,
                        "removed",
                    )

                # 处理保管人错误资产
                if not Error_Custodian.is_empty():
                    current_row = self._add_section_to_excel(
                        worksheet,
                        Error_Custodian.to_pandas(),
                        f"本月财务_VS_上月财务_保管人错误资产 {len(self.check_Custodian)}笔",
                        current_row,
                        "error",
                    )

                # 处理部门错误资产
                if not Error_Department.is_empty():
                    current_row = self._add_section_to_excel(
                        worksheet,
                        Error_Department.to_pandas(),
                        f"本月财务_VS_上月财务_部门错误资产 {len(self.check_Department)}笔",
                        current_row,
                        "error",
                    )

                # 设置列宽
                for col in ["A", "B", "C", "D", "E"]:
                    worksheet.column_dimensions[col].width = 20

            self._Save_signal.emit(self.ui.frame)

        except Exception as e:
            self._unlock_signal.emit()
            logger.exception(e)

    def _add_section_to_excel(self, worksheet, df, title, start_row, section_type):
        """添加一个数据段到Excel工作表"""
        # 添加段标题
        worksheet.cell(row=start_row, column=1, value=title)
        worksheet.cell(row=start_row, column=1).font = openpyxl.styles.Font(bold=True)
        start_row += 2

        if df.empty:
            worksheet.cell(row=start_row, column=1, value="（无数据）")
            return start_row + 2

        # 添加表头
        headers = ["No."] + list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=start_row, column=col_idx, value=header)
            cell.font = openpyxl.styles.Font(bold=True)

            # 根据类型设置颜色
            if section_type == "new":
                cell.fill = openpyxl.styles.PatternFill(
                    start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"
                )
            elif section_type == "removed":
                cell.fill = openpyxl.styles.PatternFill(
                    start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"
                )
            elif section_type == "error":
                cell.fill = openpyxl.styles.PatternFill(
                    start_color="FFFF99", end_color="FFFF99", fill_type="solid"
                )

            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style="thin"),
                right=openpyxl.styles.Side(style="thin"),
                top=openpyxl.styles.Side(style="thin"),
                bottom=openpyxl.styles.Side(style="thin"),
            )

        start_row += 1

        # 添加数据行
        for row_idx, (_, row) in enumerate(df.iterrows(), 1):
            # No. 列
            cell = worksheet.cell(row=start_row, column=1, value=row_idx)
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style="thin"),
                right=openpyxl.styles.Side(style="thin"),
                top=openpyxl.styles.Side(style="thin"),
                bottom=openpyxl.styles.Side(style="thin"),
            )

            # 数据列
            for col_idx, value in enumerate(row, 2):
                cell = worksheet.cell(
                    row=start_row,
                    column=col_idx,
                    value=str(value) if value is not None else "",
                )
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style="thin"),
                    right=openpyxl.styles.Side(style="thin"),
                    top=openpyxl.styles.Side(style="thin"),
                    bottom=openpyxl.styles.Side(style="thin"),
                )
            start_row += 1

        return start_row + 1
